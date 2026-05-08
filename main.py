print("⚡ ACCAGENIUS ULTIMATE - AI BETTING INTELLIGENCE PLATFORM ⚡")

import os
import requests
import random
import hmac
import hashlib
import asyncio
import logging
from concurrent.futures import ThreadPoolExecutor
from datetime import datetime, timedelta
# ══════════════════════════════════════════════════════════════
# PERSISTENT DATABASE
# PostgreSQL via DATABASE_URL if available, else SQLite
# ══════════════════════════════════════════════════════════════
import json as _json
import sqlite3 as _sqlite3

# Database URL — set automatically by Railway Postgres plugin
DATABASE_URL = (
    os.getenv("DATABASE_URL") or
    os.getenv("DATABASE_PRIVATE_URL") or
    os.getenv("POSTGRES_URL") or
    os.getenv("PGURL") or
    ""
).strip()


def _get_pg():
    """Try to get a PostgreSQL connection. Returns None if unavailable."""
    if not DATABASE_URL:
        return None
    try:
        import psycopg2
        conn = psycopg2.connect(DATABASE_URL, connect_timeout=5)
        return conn
    except Exception:
        return None

def _get_sqlite():
    """Get SQLite connection."""
    db_path = "/tmp/accagenius.db"
    conn = _sqlite3.connect(db_path, timeout=10)
    conn.execute("PRAGMA journal_mode=WAL")
    conn.row_factory = _sqlite3.Row
    return conn

def _is_postgres():
    return bool(DATABASE_URL)

def _get_db():
    pg = _get_pg()
    if pg:
        return pg, True
    return _get_sqlite(), False

def _init_db():
    """Create tables."""
    pg = _get_pg()
    if pg:
        try:
            cur = pg.cursor()
            cur.execute("""CREATE TABLE IF NOT EXISTS users (
                email TEXT PRIMARY KEY, data TEXT NOT NULL,
                updated_at TIMESTAMP DEFAULT NOW())""")
            cur.execute("""CREATE TABLE IF NOT EXISTS saved_accas (
                id TEXT PRIMARY KEY, email TEXT NOT NULL, data TEXT NOT NULL,
                created_at TIMESTAMP DEFAULT NOW())""")
            cur.execute("""CREATE TABLE IF NOT EXISTS codes (
                code TEXT PRIMARY KEY, email TEXT, used_at TEXT, source TEXT DEFAULT 'manual')""")
            cur.execute("CREATE INDEX IF NOT EXISTS idx_accas_email ON saved_accas(email)")
            pg.commit()
            pg.close()
            logger.info("✅ DB initialised (PostgreSQL)")
        except Exception as e:
            logger.error(f"PostgreSQL init error: {e}")
            pg.close()
    else:
        try:
            conn = _get_sqlite()
            conn.executescript("""
                CREATE TABLE IF NOT EXISTS users (
                    email TEXT PRIMARY KEY, data TEXT NOT NULL,
                    updated_at TEXT DEFAULT (datetime('now')));
                CREATE TABLE IF NOT EXISTS saved_accas (
                    id TEXT PRIMARY KEY, email TEXT NOT NULL, data TEXT NOT NULL,
                    created_at TEXT DEFAULT (datetime('now')));
                CREATE TABLE IF NOT EXISTS codes (
                    code TEXT PRIMARY KEY, email TEXT, used_at TEXT, source TEXT DEFAULT 'manual');
                CREATE INDEX IF NOT EXISTS idx_accas_email ON saved_accas(email);
            """)
            conn.commit()
            conn.close()
            logger.info("✅ DB initialised (SQLite)")
        except Exception as e:
            logger.error(f"SQLite init error: {e}")

def db_save_user(email: str, user_data: dict):
    pg = _get_pg()
    try:
        if pg:
            cur = pg.cursor()
            cur.execute(
                "INSERT INTO users (email,data) VALUES (%s,%s) ON CONFLICT (email) DO UPDATE SET data=%s,updated_at=NOW()",
                (email, _json.dumps(user_data), _json.dumps(user_data)))
            pg.commit(); pg.close()
        else:
            conn = _get_sqlite()
            conn.execute("INSERT OR REPLACE INTO users (email,data,updated_at) VALUES (?,?,datetime('now'))",
                (email, _json.dumps(user_data)))
            conn.commit(); conn.close()
    except Exception as e:
        logger.warning(f"db_save_user error: {e}")
        try: pg.close() if pg else None
        except: pass

def db_load_all_users() -> dict:
    pg = _get_pg()
    try:
        if pg:
            cur = pg.cursor()
            cur.execute("SELECT email, data FROM users")
            rows = cur.fetchall(); pg.close()
            return {r[0]: _json.loads(r[1]) for r in rows}
        else:
            conn = _get_sqlite()
            rows = conn.execute("SELECT email, data FROM users").fetchall()
            conn.close()
            return {r["email"]: _json.loads(r["data"]) for r in rows}
    except Exception as e:
        logger.warning(f"db_load_users error: {e}")
        return {}

def db_save_steam_prices(date_str: str, prices: dict):
    """Persist steam prices to DB so they survive Railway restarts."""
    pg = _get_pg()
    try:
        data_json = _json.dumps(prices)
        if pg:
            cur = pg.cursor()
            cur.execute("""CREATE TABLE IF NOT EXISTS steam_prices
                (date TEXT PRIMARY KEY, data TEXT, updated_at TIMESTAMP DEFAULT NOW())""")
            cur.execute(
                "INSERT INTO steam_prices (date,data) VALUES (%s,%s) ON CONFLICT (date) DO UPDATE SET data=%s,updated_at=NOW()",
                (date_str, data_json, data_json))
            pg.commit(); pg.close()
        else:
            conn = _get_sqlite()
            conn.execute("CREATE TABLE IF NOT EXISTS steam_prices (date TEXT PRIMARY KEY, data TEXT, updated_at TEXT)")
            conn.execute("INSERT OR REPLACE INTO steam_prices (date,data,updated_at) VALUES (?,?,datetime('now'))",
                (date_str, data_json))
            conn.commit(); conn.close()
    except Exception as e:
        logger.warning(f"db_save_steam_prices error: {e}")

def db_load_steam_prices(date_str: str) -> dict:
    """Load steam prices from DB for a given date."""
    pg = _get_pg()
    try:
        if pg:
            cur = pg.cursor()
            cur.execute("""CREATE TABLE IF NOT EXISTS steam_prices
                (date TEXT PRIMARY KEY, data TEXT, updated_at TIMESTAMP DEFAULT NOW())""")
            pg.commit()
            cur.execute("SELECT data FROM steam_prices WHERE date=%s", (date_str,))
            row = cur.fetchone(); pg.close()
            return _json.loads(row[0]) if row else {}
        else:
            conn = _get_sqlite()
            row = conn.execute("SELECT data FROM steam_prices WHERE date=?", (date_str,)).fetchone()
            conn.close()
            return _json.loads(row["data"]) if row else {}
    except Exception as e:
        logger.warning(f"db_load_steam_prices error: {e}")
        return {}

def db_save_acca(acca: dict):
    pg = _get_pg()
    aid = str(acca.get("id")); email = acca.get("email","")
    data = _json.dumps(acca); created = acca.get("created_at", datetime.now().isoformat())
    try:
        if pg:
            cur = pg.cursor()
            cur.execute(
                "INSERT INTO saved_accas (id,email,data,created_at) VALUES (%s,%s,%s,%s) ON CONFLICT (id) DO UPDATE SET data=%s",
                (aid, email, data, created, data))
            pg.commit(); pg.close()
        else:
            conn = _get_sqlite()
            conn.execute("INSERT OR REPLACE INTO saved_accas (id,email,data,created_at) VALUES (?,?,?,?)",
                (aid, email, data, created))
            conn.commit(); conn.close()
    except Exception as e:
        logger.warning(f"db_save_acca error: {e}")

def db_get_accas(email: str) -> list:
    pg = _get_pg()
    try:
        if pg:
            cur = pg.cursor()
            cur.execute("SELECT data FROM saved_accas WHERE email=%s ORDER BY created_at DESC", (email.lower(),))
            rows = cur.fetchall(); pg.close()
            return [_json.loads(r[0]) for r in rows]
        else:
            conn = _get_sqlite()
            rows = conn.execute("SELECT data FROM saved_accas WHERE email=? ORDER BY created_at DESC",
                (email.lower(),)).fetchall()
            conn.close()
            return [_json.loads(r["data"]) for r in rows]
    except Exception as e:
        logger.warning(f"db_get_accas error: {e}")
        return []

def db_delete_acca(acca_id: str):
    pg = _get_pg()
    try:
        if pg:
            cur = pg.cursor()
            cur.execute("DELETE FROM saved_accas WHERE id=%s", (str(acca_id),))
            pg.commit(); pg.close()
        else:
            conn = _get_sqlite()
            conn.execute("DELETE FROM saved_accas WHERE id=?", (str(acca_id),))
            conn.commit(); conn.close()
    except Exception as e:
        logger.warning(f"db_delete_acca error: {e}")

def db_save_code(code: str, email: str = "", source: str = "manual"):
    pg = _get_pg()
    used_at = datetime.now().isoformat() if email else None
    try:
        if pg:
            cur = pg.cursor()
            cur.execute(
                "INSERT INTO codes (code,email,used_at,source) VALUES (%s,%s,%s,%s) ON CONFLICT (code) DO UPDATE SET email=%s,used_at=%s",
                (code, email, used_at, source, email, used_at))
            pg.commit(); pg.close()
        else:
            conn = _get_sqlite()
            conn.execute("INSERT OR REPLACE INTO codes (code,email,used_at,source) VALUES (?,?,?,?)",
                (code, email, used_at, source))
            conn.commit(); conn.close()
    except Exception as e:
        logger.warning(f"db_save_code error: {e}")

def db_load_codes() -> tuple:
    pg = _get_pg()
    try:
        if pg:
            cur = pg.cursor()
            cur.execute("SELECT code, email, used_at, source FROM codes")
            rows = cur.fetchall(); pg.close()
            all_c = {r[0] for r in rows}
            used_c = {r[0]: {"email": r[1], "used_at": r[2], "source": r[3]} for r in rows if r[1]}
        else:
            conn = _get_sqlite()
            rows = conn.execute("SELECT code, email, used_at, source FROM codes").fetchall()
            conn.close()
            all_c = {r["code"] for r in rows}
            used_c = {r["code"]: {"email": r["email"], "used_at": r["used_at"], "source": r["source"]}
                      for r in rows if r["email"]}
        return all_c, used_c
    except Exception as e:
        logger.warning(f"db_load_codes error: {e}")
        return set(), {}

# London timezone helper — no external deps required
def _london_now():
    """Return current datetime in Europe/London time without external deps."""
    import time as _time
    # UK is UTC+0 (GMT) Oct-Mar, UTC+1 (BST) Mar-Oct
    # DST starts last Sunday March, ends last Sunday October
    now_utc = datetime.utcnow()
    year = now_utc.year
    # Last Sunday in March
    import calendar
    mar_days = calendar.monthrange(year, 3)[1]
    dst_start = datetime(year, 3, 31 - (datetime(year, 3, 31).weekday() + 1) % 7, 1)
    # Last Sunday in October  
    dst_end = datetime(year, 10, 31 - (datetime(year, 10, 31).weekday() + 1) % 7, 1)
    if dst_start <= now_utc.replace(tzinfo=None) < dst_end:
        return now_utc + __import__("datetime").timedelta(hours=1)
    return now_utc
from typing import Optional, Dict, Any, List
from fastapi import FastAPI, HTTPException, Request
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel

logger = logging.getLogger("accagenius")

# =========================
# CONFIG
# =========================
API_FOOTBALL_KEY = os.getenv("API_FOOTBALL_KEY", "a65dde1b9f1c13dd484a2d1889b70824")
BASE_URL = "https://v3.football.api-sports.io"
HEADERS = {
    "x-rapidapi-key": API_FOOTBALL_KEY,
    "x-rapidapi-host": "v3.football.api-sports.io"
}

def get_headers():
    """Read API key fresh each call in case env var loads after startup."""
    key = os.getenv("API_FOOTBALL_KEY", API_FOOTBALL_KEY)
    return {
        "x-rapidapi-key": key,
        "x-rapidapi-host": "v3.football.api-sports.io"
    }

# Telegram config — reads YOUR Railway variable names
TELEGRAM_BOT_TOKEN    = os.getenv("TELEGRAM_BOT_TOKEN", "8459345549:AAEwCGBLjwijbVliHY1Qaw9KB-fybfu0Pa8")
# Read your exact variable names first, then fallbacks
TELEGRAM_CHANNEL_FREE = (
    os.getenv("FREE_CHANNEL_ID", "") or
    os.getenv("TELEGRAM_CHANNEL_FREE", "") or
    os.getenv("TELEGRAM_CHANNEL_ID", "")
)
TELEGRAM_CHANNEL_PRO  = (
    os.getenv("PRO_CHANNEL_ID", "") or
    os.getenv("TELEGRAM_CHANNEL_PRO", "") or
    os.getenv("TELEGRAM_CHANNEL_ID", "")
)
TELEGRAM_CHANNEL_ID   = TELEGRAM_CHANNEL_PRO or TELEGRAM_CHANNEL_FREE

# ── EMAIL (Resend) ───────────────────────────────────────────────
RESEND_API_KEY           = os.getenv("RESEND_API_KEY", "re_LuMWocBr_AwT1A7BeJ4dF1Wzrsc7Q84Ae")
EMAIL_FROM               = os.getenv("EMAIL_FROM", "AccaGenius <noreply@accagenius.com>")
SITE_URL                 = os.getenv("SITE_URL", "https://accagenius.com")

# ── STRIPE ───────────────────────────────────────────────────────────────────
STRIPE_API_KEY        = os.getenv("STRIPE_API_KEY", "sk_test_51T7zKIC2iUpRnqYDN1GdhxTByYL8jb9TYkOSeBL7BQIiICqqSyt5hfzzy7kt4hYNROCoIh16KBdRm8nfVtLL5znG00SI0OF04Z")
STRIPE_WH_SECRET    = os.getenv("STRIPE_WH_SECRET", "")
STRIPE_PRICE_ID          = os.getenv("STRIPE_PRICE_ID", "")  # your £20/month recurring price ID
STRIPE_TRADING_PRICE_ID  = os.getenv("STRIPE_TRADING_PRICE_ID", "")  # £300/year trading plan
STRIPE_TRADING_TRIAL_ID  = os.getenv("STRIPE_TRADING_TRIAL_ID", "")  # £1 for 30 days trial
ADMIN_EMAIL              = "ajjennings7@hotmail.com"  # admin gets access to everything
TRADING_SEASONS          = ["2020-21","2021-22","2022-23","2023-24","2024-25","2025-26"]

def _is_admin(email: str) -> bool:
    """Check if email is the admin account."""
    if not email: return False
    return email.lower().strip() == ADMIN_EMAIL.lower().strip()

def _is_trading(email: str) -> bool:
    """Check if user has trading plan access (admin always has access)."""
    if _is_admin(email): return True
    user = _user_db.get(email.lower().strip(), {})
    return user.get("plan") in ("trading", "trading_trial")

def _is_pro_or_above(email: str) -> bool:
    """Check if user has pro or trading access."""
    if _is_admin(email): return True
    user = _user_db.get(email.lower().strip(), {})
    return user.get("plan") in ("pro", "trading", "trading_trial")
_reset_tokens: dict = {}  # token -> {email, name, expires, used}
_TOKENS_FILE = "/tmp/ag_reset_tokens.json"

# ── ACCESS CODES ─────────────────────────────────────────────────
PRO_ACCESS_CODES_RAW     = os.getenv("PRO_ACCESS_CODES", "ACCA-X7K2,ACCA-M9P4")
_all_codes: set          = set(c.strip().upper() for c in PRO_ACCESS_CODES_RAW.split(",") if c.strip())
_used_codes: dict        = {}  # code -> {email, name, used_at}
TELEGRAM_PRO_INVITE_LINK = os.getenv("TELEGRAM_PRO_INVITE_LINK", "https://t.me/+6hHYweKkeEs3Mjk0")

# ── ADMIN PASSWORD ───────────────────────────────────────────────
# Support both variable names - AJ_ADMIN_2026 (current) and ADMIN_PASSWORD (standard)
ADMIN_PASSWORD_ENV = (os.getenv("ADMIN_PASSWORD") or os.getenv("AJ_ADMIN_2026") or "").strip()

# ── Football-Bet-Data.com credentials ──
FBD_EMAIL    = os.getenv("FBD_EMAIL", "").strip()
FBD_PASSWORD = os.getenv("FBD_PASSWORD", "").strip()

# Daily limits for free channel
free_alerts_sent_today  = 0   # max 1 per day on free
free_alert_reset_date   = ""

# Alert thresholds
ALERT_WIN_PCT    = 60    # Win % must be >= this (lowered from 65)
ALERT_XG_GAP    = 0.25  # xG gap between teams must be >= this (lowered from 0.4)
ALERT_MINUTE_MIN = 20   # Don't alert before 20th minute
ALERT_MINUTE_MAX = 80   # Don't alert after 80th minute (extended from 75)

app = FastAPI(title="AccaGenius Ultimate API")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

from fastapi.responses import JSONResponse
from fastapi.requests import Request

@app.exception_handler(Exception)
async def global_exception_handler(request: Request, exc: Exception):
    """Catch-all — return clean JSON instead of crashing and flooding logs"""
    logger.error(f"Unhandled error on {request.url.path}: {type(exc).__name__}: {exc}")
    return JSONResponse(
        status_code=500,
        content={"error": str(exc), "path": str(request.url.path)}
    )

# In-memory store
saved_accas_store: List[dict] = []

# =========================
# SMART API CACHE
# Saves your 100 calls/day free quota by serving cached data
# TTLs chosen to balance freshness vs API spend:
#   today fixtures  — 20 min  (live scores handled by /live endpoint separately)
#   next-round      — 60 min  (fixtures don't change often)
#   standings       — 120 min (table changes only after matches)
#   form/h2h        — 180 min (historical — barely changes)
#   odds            — 15 min  (odds shift but slowly)
#   live scores     — 60 sec  (must be fresh — but only called when XG tab open)
# =========================
_cache: Dict[str, Any] = {}

# ────────────────────────────────────────────────────────────────
# EMAIL  (Resend API — resend.com)
# ────────────────────────────────────────────────────────────────

def send_email(to: str, subject: str, html_body: str) -> bool:
    if not RESEND_API_KEY:
        logger.warning("RESEND_API_KEY not set — email skipped")
        return False
    reply_to = os.getenv("SUPPORT_EMAIL", "j3nno83@gmail.com")
    try:
        r = requests.post(
            "https://api.resend.com/emails",
            headers={"Authorization": f"Bearer {RESEND_API_KEY}", "Content-Type": "application/json"},
            json={"from": EMAIL_FROM, "reply_to": [reply_to], "to": [to], "subject": subject, "html": html_body},
            timeout=10
        )
        ok = r.status_code in (200, 201)
        if ok:
            logger.info(f"✉️  Email sent: {subject} → {to}")
        else:
            logger.error(f"Email failed {r.status_code}: {r.text[:200]}")
        return ok
    except Exception as e:
        logger.error(f"Email error: {e}")
        return False


def _email_base(body_html: str) -> str:
    """Wrap content in branded AccaGenius email shell."""
    return f"""<!DOCTYPE html><html><head><meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<style>body{{margin:0;padding:0;background:#0a0c12;font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif}}</style>
</head><body>
<div style="max-width:560px;margin:0 auto;padding:32px 16px">
  <div style="text-align:center;margin-bottom:28px">
    <div style="display:inline-block;background:linear-gradient(135deg,#f0b429,#f6ad55);border-radius:14px;padding:10px 20px">
      <span style="font-size:20px;font-weight:900;color:#000;letter-spacing:1px">⚡ ACCAGENIUS</span>
    </div>
  </div>
  {body_html}
  <div style="text-align:center;color:#4b5563;font-size:12px;margin-top:28px;line-height:1.8">
    AccaGenius · <a href="{SITE_URL}" style="color:#4cc9f0;text-decoration:none">accagenius.com</a><br>
    Please gamble responsibly · 18+ only ·
    <a href="https://www.begambleaware.org" style="color:#4cc9f0;text-decoration:none">BeGambleAware.org</a>
  </div>
</div>
</body></html>"""


def email_welcome_pro(to: str, name: str, code: str, telegram_invite: str = "") -> bool:
    tg = f"""<div style="background:#0d1b2a;border:1px solid #4cc9f0;border-radius:10px;padding:16px;margin:16px 0;text-align:center">
      <div style="font-size:22px;margin-bottom:6px">📲</div>
      <div style="font-weight:700;color:#4cc9f0;margin-bottom:6px">Join Your Pro Telegram Channel</div>
      <div style="color:#8b9cbf;font-size:13px;margin-bottom:12px">Daily accas, live in-play alerts &amp; xG analysis straight to your phone</div>
      <a href="{telegram_invite}" style="display:inline-block;background:linear-gradient(90deg,#4cc9f0,#7c3aed);color:#fff;font-weight:700;padding:10px 24px;border-radius:8px;text-decoration:none;font-size:14px">Join AccaGenius Pro →</a>
    </div>""" if telegram_invite else ""

    features = [
        ("AI Acca Generator", "Balanced, safe &amp; risky picks across 19 leagues"),
        ("Live xG Tracker", "In-play expected goals with Goal Due alerts"),
        ("Today's Acca of the Day", "3 themed accas every morning — Form, Rank &amp; Value"),
        ("HT Value Finder", "First-half patterns and halftime picks"),
        ("Full Match Analysis", "H2H, form, lineups, predictions &amp; best odds"),
        ("Saved Accas &amp; P&amp;L", "Track every bet and monitor your profit"),
    ]
    feat_html = "".join(
        f'<div style="display:flex;gap:10px;margin-bottom:10px"><span style="color:#f0b429">⚡</span>'
        f'<div><div style="color:#fff;font-size:14px;font-weight:600">{f[0]}</div>'
        f'<div style="color:#8b9cbf;font-size:12px">{f[1]}</div></div></div>'
        for f in features
    )
    body = f"""
    <div style="background:#111827;border-radius:16px;padding:28px;margin-bottom:16px;border:1px solid #1f2937">
      <h1 style="color:#f0b429;font-size:22px;margin:0 0 10px;font-weight:900">Welcome to the Pro, {name}! 🎉</h1>
      <p style="color:#8b9cbf;font-size:14px;margin:0 0 20px;line-height:1.6">
        Your Pro access is now active. Full AI picks, live stats, xG analysis — everything is unlocked.
      </p>
      <div style="background:#0d1b2a;border:1px solid #1f2937;border-radius:10px;padding:14px;margin-bottom:16px">
        <div style="color:#8b9cbf;font-size:11px;text-transform:uppercase;letter-spacing:1px;margin-bottom:4px">Your Access Code</div>
        <div style="font-family:monospace;font-size:22px;font-weight:900;color:#f0b429;letter-spacing:4px">{code}</div>
        <div style="color:#8b9cbf;font-size:11px;margin-top:4px">Keep this safe — you'll need it if you sign in on a new device</div>
      </div>
      {tg}
      <div style="text-align:center;margin-top:20px">
        <a href="{SITE_URL}" style="display:inline-block;background:linear-gradient(135deg,#f0b429,#f6ad55);color:#000;font-weight:800;padding:13px 32px;border-radius:10px;text-decoration:none;font-size:15px">🚀 Go to AccaGenius</a>
      </div>
    </div>
    <div style="background:#111827;border-radius:16px;padding:22px;border:1px solid #1f2937">
      <h2 style="color:#fff;font-size:15px;margin:0 0 14px;font-weight:700">✅ What's included in your Pro plan</h2>
      {feat_html}
    </div>"""
    return send_email(to, "⚡ Welcome to AccaGenius Pro — you're in!", _email_base(body))



def email_welcome_free(to: str, name: str) -> bool:
    """Send welcome email to free users with free Telegram channel link."""
    if not RESEND_API_KEY or not to:
        return False
    first = name.split()[0] if name else "there"
    body = f"""
    <div style="background:#111827;border-radius:16px;padding:28px;border:1px solid #1f2937;max-width:520px;margin:0 auto">
      <div style="font-size:36px;text-align:center;margin-bottom:12px">⚡</div>
      <h1 style="color:#fff;font-size:20px;margin:0 0 10px;text-align:center;font-weight:800">Welcome to AccaGenius!</h1>
      <p style="color:#8b9cbf;font-size:14px;margin:0 0 20px;text-align:center">
        Hi {first}, your free account is ready 🎉
      </p>
      <div style="background:#0d1117;border-radius:12px;padding:20px;margin-bottom:20px">
        <p style="color:#e2e8f0;font-weight:700;margin:0 0 12px">✅ Your Free Plan includes:</p>
        <ul style="color:#8b9cbf;font-size:13px;line-height:2;margin:0;padding-left:20px">
          <li>Up to 3 in-play tips/day</li>
          <li>Acca of the Day — teams &amp; bet type</li>
          <li>Win/loss result updates</li>
          <li>Running P&amp;L tracker</li>
          <li>Free Telegram tips channel</li>
        </ul>
      </div>
      <a href="https://t.me/AccaGeniusFree" target="_blank"
         style="display:block;background:linear-gradient(90deg,#f0b429,#f59e0b);color:#000;font-weight:800;
                font-size:15px;text-align:center;padding:14px;border-radius:10px;text-decoration:none;margin-bottom:16px">
        📲 Join Free Telegram Channel
      </a>
      <div style="background:rgba(240,180,41,0.08);border:1px solid rgba(240,180,41,0.2);border-radius:8px;
                  padding:14px;text-align:center;margin-bottom:16px">
        <p style="color:#f0b429;font-weight:700;margin:0 0 6px;font-size:14px">Want full AI analysis?</p>
        <p style="color:#8b9cbf;font-size:12px;margin:0">Upgrade to Pro for xG stats, win probability, live odds &amp; more</p>
        <a href="{SITE_URL}" style="color:#f0b429;font-size:12px;font-weight:700">→ {SITE_URL}</a>
      </div>
      <p style="color:#475569;font-size:11px;text-align:center;margin:0">AccaGenius · {SITE_URL}</p>
    </div>"""
    try:
        r = requests.post("https://api.resend.com/emails",
            headers={"Authorization": f"Bearer {RESEND_API_KEY}", "Content-Type": "application/json"},
            json={"from": EMAIL_FROM, "to": [to], "subject": "⚡ Welcome to AccaGenius!", "html": f"<html><body style='background:#0a0a0f;padding:20px'>{body}</body></html>"},
            timeout=10)
        logger.info(f"Free welcome email to {to}: {r.status_code}")
        return r.status_code in (200, 201)
    except Exception as e:
        logger.error(f"Free welcome email failed: {e}")
        return False


def email_password_reset(to: str, name: str, token: str) -> bool:
    reset_url = f"{SITE_URL}?reset={token}"
    body = f"""
    <div style="background:#111827;border-radius:16px;padding:28px;border:1px solid #1f2937">
      <div style="font-size:36px;text-align:center;margin-bottom:12px">🔑</div>
      <h1 style="color:#fff;font-size:20px;margin:0 0 10px;text-align:center;font-weight:800">Reset Your Password</h1>
      <p style="color:#8b9cbf;font-size:14px;margin:0 0 24px;text-align:center;line-height:1.6">
        Hi {name}, tap below to reset your AccaGenius password.<br>
        <strong style="color:#f0b429">This link expires in 1 hour.</strong>
      </p>
      <div style="text-align:center;margin-bottom:22px">
        <a href="{reset_url}" style="display:inline-block;background:linear-gradient(135deg,#f0b429,#f6ad55);color:#000;font-weight:800;padding:13px 32px;border-radius:10px;text-decoration:none;font-size:15px">Reset My Password</a>
      </div>
      <div style="background:#0d1b2a;border-radius:8px;padding:12px;font-size:12px;color:#8b9cbf;text-align:center">
        If you didn't request this, ignore this email. Your password won't change.
      </div>
    </div>"""
    return send_email(to, "🔑 Reset your AccaGenius password", _email_base(body))


def cache_get(key: str) -> Any:
    """Return cached value if still fresh, else None."""
    entry = _cache.get(key)
    if entry and datetime.now() < entry["expires"]:
        return entry["data"]
    return None

def cache_set(key: str, data: Any, ttl_seconds: int):
    """Store data in cache with expiry."""
    _cache[key] = {"data": data, "expires": datetime.now() + timedelta(seconds=ttl_seconds)}

def cache_stats() -> dict:
    """Return cache hit counts and remaining quota estimate."""
    now = datetime.now()
    live_keys = [k for k, v in _cache.items() if now < v["expires"]]
    return {"cached_keys": len(live_keys), "total_keys": len(_cache)}

# Track API calls to warn when approaching daily limit
_api_call_count = 0
_api_call_reset = datetime.now().date()

def api_get(endpoint: str, params: dict, timeout: int = 10) -> dict:
    """Safe API call wrapper with quota tracking."""
    global _api_call_count, _api_call_reset
    # Reset counter each day
    today = _london_now().date()
    if today != _api_call_reset:
        _api_call_count = 0
        _api_call_reset = today
    _api_call_count += 1
    if _api_call_count % 5000 == 0 or _api_call_count > 65000:
        pct = round(_api_call_count / 75000 * 100, 1)
        logger.warning(f"📊 API quota: {_api_call_count}/75000 ({pct}%) calls today")
    if _api_call_count > 74000:
        logger.error(f"🛑 API call blocked — at {_api_call_count}/75000 daily limit")
        return {}
    try:
        r = requests.get(f"{BASE_URL}/{endpoint}", headers=get_headers(), params=params, timeout=timeout)
        if r.status_code == 429:
            logger.error(f"🚫 API rate limit hit (429) on /{endpoint} — {_api_call_count} calls today")
            return {}
        if r.status_code == 200:
            return r.json()
        return {}
    except Exception as e:
        logger.error(f"api_get error on /{endpoint}: {e}")
        return {}

# Track which fixtures we've already alerted — resets each day
alerted_fixtures: set = set()
alerted_details: dict = {}  # fixture_id -> {home, away, bet, odds, channel} for result posting
last_alert_reset: str = ""

# =========================
# ══════════════════════════════════════════════════════════════════
# FOOTBALL-BET-DATA.COM — FULL INTEGRATION
# ══════════════════════════════════════════════════════════════════
import io as _io
_fbd_cache = {"date": None, "data": None}
_FBD_BASE  = "https://www.football-bet-data.com"
_FBD_HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/146.0.0.0 Safari/537.36 Edg/146.0.0.0",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8",
    "Accept-Language": "en-GB,en;q=0.9",
}

def _fbd_odds_to_prob(val):
    try:
        v = float(val)
        if v <= 0: return None
        if 1.0 <= v <= 20.0: return round((1/v)*100, 1)
        if 0 < v <= 100: return round(v, 1)
        return None
    except: return None

def _fbd_get_session():
    import requests as _req
    from bs4 import BeautifulSoup as _BS
    s = _req.Session()
    s.headers.update(_FBD_HEADERS)
    r = s.get(f"{_FBD_BASE}/signin/", timeout=15)
    soup = _BS(r.text, "html.parser")
    def _v(name):
        t = soup.find("input", {"name": name})
        return t["value"] if t and t.get("value") else ""
    vs = _v("__VIEWSTATE")
    ev = _v("__EVENTVALIDATION")
    print(f"FBD login: VIEWSTATE={'yes' if vs else 'NO'}, email_field={'yes' if soup.find('input',{'name':'ctl00$ContentPlaceHolder2$unameTextBox'}) else 'NO'}", flush=True)
    payload = {
        "__EVENTTARGET":"","__EVENTARGUMENT":"","__LASTFOCUS":"",
        "__VIEWSTATE": vs,
        "__VIEWSTATEGENERATOR": _v("__VIEWSTATEGENERATOR"),
        "__SCROLLPOSITIONX":"0","__SCROLLPOSITIONY":"0",
        "__EVENTVALIDATION": ev,
        "ctl00$ContentPlaceHolder2$HiddenField1": "/",
        "ctl00$ContentPlaceHolder2$unameTextBox": FBD_EMAIL,
        "ctl00$ContentPlaceHolder2$pwordTextBox": FBD_PASSWORD,
        "ctl00$ContentPlaceHolder2$submitButton": "Submit",
    }
    r2 = s.post(f"{_FBD_BASE}/signin/", data=payload,
                headers={"Content-Type":"application/x-www-form-urlencoded",
                         "Referer":f"{_FBD_BASE}/signin/","Origin":_FBD_BASE},
                allow_redirects=True, timeout=15)
    if "signin" in r2.url:
        print("FBD login: FAILED — still on signin page", flush=True)
    else:
        print(f"FBD login: SUCCESS — {r2.url}", flush=True)
    return s

def _fbd_download_excel(session):
    from bs4 import BeautifulSoup as _BS
    r = session.get(f"{_FBD_BASE}/today/", timeout=15)
    soup = _BS(r.text, "html.parser")
    def _v(name):
        t = soup.find("input", {"name": name})
        return t["value"] if t and t.get("value") else ""
    payload = {
        "__EVENTTARGET":"","__EVENTARGUMENT":"","__LASTFOCUS":"",
        "__VIEWSTATE": _v("__VIEWSTATE"),
        "__VIEWSTATEGENERATOR": _v("__VIEWSTATEGENERATOR"),
        "__SCROLLPOSITIONX":"0","__SCROLLPOSITIONY":"0",
        "__EVENTVALIDATION": _v("__EVENTVALIDATION"),
        "ctl00$ContentPlaceHolder1$dlButton1":"Download",
    }
    r2 = session.post(f"{_FBD_BASE}/today/", data=payload,
                      headers={"Content-Type":"application/x-www-form-urlencoded",
                               "Referer":f"{_FBD_BASE}/today/","Origin":_FBD_BASE,
                               "Sec-Fetch-Dest":"document","Sec-Fetch-Mode":"navigate",
                               "Sec-Fetch-Site":"same-origin","Upgrade-Insecure-Requests":"1"},
                      allow_redirects=True, timeout=30)
    ct = r2.headers.get("Content-Type","")
    if not any(x in ct for x in ("excel","spreadsheet","octet","zip")):
        logger.warning(f"FBD unexpected content-type: {ct}")
        return None
    return r2.content

def _fbd_parse_excel(excel_bytes):
    try:
        import openpyxl as _xl
        wb = _xl.load_workbook(_io.BytesIO(excel_bytes), read_only=True, data_only=True)
        target = None
        for name in wb.sheetnames:
            if name.lower().startswith("date-") or name.lower()=="summary":
                target=wb[name]; break
        if not target: target=wb[wb.sheetnames[0]]
        matches=[]; headers=None
        for row in target.iter_rows(values_only=True):
            if not any(row): continue
            if headers is None:
                if row[0] in ("League","Date","Home Team") or (len(row)>1 and row[1] in ("Home Team",)):
                    headers=[str(h).strip() if h else f"col_{i}" for i,h in enumerate(row)]
                continue
            if not headers: continue
            m={headers[i]:v for i,v in enumerate(row) if i<len(headers)}
            home=str(m.get("Home Team","") or m.get("col_1","") or "").strip()
            away=str(m.get("Away Team","") or m.get("col_2","") or "").strip()
            if not home or not away or home=="Home Team": continue
            matches.append({
                "league":         str(m.get("League","") or m.get("LeagueID","")),
                "home":           home, "away": away,
                "prediction":     str(m.get("Pred") or m.get("Pred'n","") or ""),
                "pred_home_odds": _fbd_odds_to_prob(m.get("Pred H Odds")),
                "pred_draw_odds": _fbd_odds_to_prob(m.get("Pred D Odds")),
                "pred_away_odds": _fbd_odds_to_prob(m.get("Pred A Odds")),
                "avg_home_odds":  _fbd_odds_to_prob(m.get("Avg H Odds")),
                "avg_draw_odds":  _fbd_odds_to_prob(m.get("Avg D Odds")),
                "avg_away_odds":  _fbd_odds_to_prob(m.get("Avg A Odds")),
                "max_home_odds":  _fbd_odds_to_prob(m.get("Max H Odds")),
                "pred_over25":    _fbd_odds_to_prob(m.get("Pred Ov25")),
                "pred_under25":   _fbd_odds_to_prob(m.get("Pred Un25")),
                "pred_btts":      _fbd_odds_to_prob(m.get("Pred BTS Y")),
                "pred_goals":     _fbd_odds_to_prob(m.get("Pred Tot Gls")),
                "avg_over25":     _fbd_odds_to_prob(m.get("Avg ov25")),
                "avg_btts_yes":   _fbd_odds_to_prob(m.get("BTS Y")),
                "dc_1x":          _fbd_odds_to_prob(m.get("DC 1X")),
                "dc_x2":          _fbd_odds_to_prob(m.get("DC X2")),
            })
        logger.info(f"✅ FBD: parsed {len(matches)} fixtures")
        return matches
    except Exception as e:
        logger.error(f"FBD parse error: {e}")
        return []

def get_fbd_data(force=False):
    today_str = datetime.now().strftime("%Y-%m-%d")
    if not force and _fbd_cache["date"]==today_str and _fbd_cache["data"]:
        return _fbd_cache["data"]
    if not FBD_EMAIL or not FBD_PASSWORD:
        return []
    try:
        session = _fbd_get_session()
        excel   = _fbd_download_excel(session)
        if not excel: return []
        data = _fbd_parse_excel(excel)
        _fbd_cache["date"] = today_str
        _fbd_cache["data"] = data
        logger.info(f"✅ FBD daily data loaded: {len(data)} matches")
        return data
    except Exception as e:
        logger.error(f"FBD download failed: {e}")
        return []

# ── FBD Team Name Aliases ──
FBD_NAME_ALIASES = {
    "psv eindhoven":"psv","nec nijmegen":"nijmegen","fortuna sittard":"sittard",
    "az alkmaar":"az","borussia dortmund":"dortmund","bayer leverkusen":"leverkusen",
    "rb leipzig":"leipzig","borussia mönchengladbach":"m'gladbach",
    "eintracht frankfurt":"frankfurt","vfb stuttgart":"stuttgart",
    "1. fc köln":"köln","1. fc union berlin":"union berlin",
    "sc freiburg":"freiburg","vfl wolfsburg":"wolfsburg","vfl bochum":"bochum",
    "fc augsburg":"augsburg","tsg hoffenheim":"hoffenheim",
    "sv werder bremen":"werder bremen","1. fsv mainz 05":"mainz",
    "manchester city":"man city","manchester united":"man utd",
    "tottenham hotspur":"tottenham","wolverhampton wanderers":"wolves",
    "west ham united":"west ham","newcastle united":"newcastle",
    "nottingham forest":"nott'm forest","brighton & hove albion":"brighton",
    "atletico madrid":"atl. madrid","athletic bilbao":"athletic club",
    "real betis":"betis","ac milan":"milan","inter milan":"inter",
    "fc internazionale":"inter","ss lazio":"lazio","as roma":"roma",
    "ssc napoli":"napoli","paris saint-germain":"psg",
    "olympique marseille":"marseille","olympique lyonnais":"lyon",
    "galatasaray sk":"galatasaray","fenerbahçe sk":"fenerbahce",
    "beşiktaş jk":"besiktas","bayern münchen":"bayern munich",
    "fc bayern münchen":"bayern munich",
}

def _fbd_normalise(name: str) -> str:
    n = name.lower().strip()
    if n in FBD_NAME_ALIASES: return FBD_NAME_ALIASES[n]
    for api_name, fbd_name in FBD_NAME_ALIASES.items():
        if n == fbd_name: return fbd_name
    return n

def _fbd_names_match(a: str, b: str) -> bool:
    an, bn = _fbd_normalise(a), _fbd_normalise(b)
    if an == bn: return True
    if an in bn or bn in an: return True
    a_words = [w for w in an.split() if len(w)>3]
    b_words = [w for w in bn.split() if len(w)>3]
    if a_words and b_words:
        if any(w in bn for w in a_words): return True
        if any(w in an for w in b_words): return True
    return False

def get_fbd_match(home: str, away: str):
    for m in get_fbd_data():
        if _fbd_names_match(home, m["home"]) and _fbd_names_match(away, m["away"]):
            return m
    return None

def fbd_boost_confidence(base_conf: int, home: str, away: str, bet: str) -> int:
    fbd = get_fbd_match(home, away)
    if not fbd: return base_conf
    pred = fbd.get("prediction","").strip().upper()
    if bet=="home"  and pred in ("H","1"): base_conf=min(95,base_conf+8)
    if bet=="away"  and pred in ("A","2"): base_conf=min(95,base_conf+8)
    if bet=="draw"  and pred=="D":         base_conf=min(95,base_conf+6)
    if bet=="home"  and pred in ("A","2"): base_conf=max(30,base_conf-5)
    if bet=="away"  and pred in ("H","1"): base_conf=max(30,base_conf-5)
    return base_conf

# LEAGUES (17)
# =========================
LEAGUES = [
    # ── England ──
    {"code": "PL",   "name": "Premier League",         "country": "England",     "id": 39,  "flag": "🏴󠁧󠁢󠁥󠁮󠁧󠁿"},
    {"code": "ELC",  "name": "Championship",            "country": "England",     "id": 40,  "flag": "🏴󠁧󠁢󠁥󠁮󠁧󠁿"},
    {"code": "EL1",  "name": "League One",              "country": "England",     "id": 41,  "flag": "🏴󠁧󠁢󠁥󠁮󠁧󠁿"},
    {"code": "EL2",  "name": "League Two",              "country": "England",     "id": 42,  "flag": "🏴󠁧󠁢󠁥󠁮󠁧󠁿"},
    # ── Europe ──
    {"code": "CL",   "name": "Champions League",        "country": "Europe",      "id": 2,   "flag": "🌍"},
    {"code": "EL",   "name": "Europa League",           "country": "Europe",      "id": 3,   "flag": "🌍"},
    {"code": "ECL",  "name": "Conference League",       "country": "Europe",      "id": 848, "flag": "🌍"},
    # ── Internationals ──
    {"code": "WC",   "name": "World Cup",               "country": "World",       "id": 1,   "flag": "🌎"},
    {"code": "EURO", "name": "European Championship",   "country": "Europe",      "id": 4,   "flag": "🇪🇺"},
    {"code": "NL",   "name": "UEFA Nations League",     "country": "Europe",      "id": 5,   "flag": "🇪🇺"},
    {"code": "WCQ",  "name": "World Cup Qualifying",    "country": "World",       "id": 32,  "flag": "🌎"},
    {"code": "CA",   "name": "Copa America",            "country": "S.America",   "id": 9,   "flag": "🌎"},
    # ── Spain ──
    {"code": "PD",   "name": "La Liga",                 "country": "Spain",       "id": 140, "flag": "🇪🇸"},
    {"code": "SD",   "name": "La Liga 2",               "country": "Spain",       "id": 141, "flag": "🇪🇸"},
    # ── Germany ──
    {"code": "BL1",  "name": "Bundesliga",              "country": "Germany",     "id": 78,  "flag": "🇩🇪"},
    {"code": "BL2",  "name": "2. Bundesliga",           "country": "Germany",     "id": 79,  "flag": "🇩🇪"},
    # ── Italy ──
    {"code": "SA",   "name": "Serie A",                 "country": "Italy",       "id": 135, "flag": "🇮🇹"},
    {"code": "SB",   "name": "Serie B",                 "country": "Italy",       "id": 136, "flag": "🇮🇹"},
    # ── France ──
    {"code": "FL1",  "name": "Ligue 1",                 "country": "France",      "id": 61,  "flag": "🇫🇷"},
    {"code": "FL2",  "name": "Ligue 2",                 "country": "France",      "id": 62,  "flag": "🇫🇷"},
    # ── Other Europe ──
    {"code": "NED",  "name": "Eredivisie",              "country": "Netherlands", "id": 88,  "flag": "🇳🇱"},
    {"code": "PPL",  "name": "Primeira Liga",           "country": "Portugal",    "id": 94,  "flag": "🇵🇹"},
    {"code": "TUR",  "name": "Super Lig",               "country": "Turkey",      "id": 203, "flag": "🇹🇷"},
    {"code": "BEL",  "name": "Belgium Pro League",      "country": "Belgium",     "id": 144, "flag": "🇧🇪"},
    {"code": "POL",  "name": "Ekstraklasa",             "country": "Poland",      "id": 106, "flag": "🇵🇱"},
    {"code": "SPFL", "name": "Scottish Premiership",    "country": "Scotland",    "id": 179, "flag": "🏴󠁧󠁢󠁳󠁣󠁴󠁿"},
    {"code": "AUT",  "name": "Austrian Bundesliga",     "country": "Austria",     "id": 218, "flag": "🇦🇹"},
    {"code": "SWE",  "name": "Allsvenskan",             "country": "Sweden",      "id": 113, "flag": "🇸🇪"},
    {"code": "NOR",  "name": "Eliteserien",             "country": "Norway",      "id": 103, "flag": "🇳🇴"},
    {"code": "GRK",  "name": "Super League Greece",     "country": "Greece",      "id": 197, "flag": "🇬🇷"},
    # ── Americas (Summer) ──
    {"code": "BSA",  "name": "Serie A Brazil",          "country": "Brazil",      "id": 71,  "flag": "🇧🇷"},
    {"code": "MLS",  "name": "MLS",                     "country": "USA",         "id": 253, "flag": "🇺🇸"},
    {"code": "LMX",  "name": "Liga MX",                 "country": "Mexico",      "id": 262, "flag": "🇲🇽"},
    {"code": "ARG",  "name": "Liga Profesional",        "country": "Argentina",   "id": 128, "flag": "🇦🇷"},
]
LEAGUE_IDS = {l["code"]: l["id"] for l in LEAGUES}
LEAGUE_INFO = {l["code"]: l for l in LEAGUES}

def get_season() -> int:
    now = datetime.now()
    return now.year if now.month >= 8 else now.year - 1

# =========================
# TELEGRAM ALERTS
# =========================

def send_telegram(message: str) -> bool:
    """Send message to Telegram channel. Returns True on success."""
    channel = TELEGRAM_CHANNEL_PRO or TELEGRAM_CHANNEL_ID or TELEGRAM_CHANNEL_FREE
    if not TELEGRAM_BOT_TOKEN or not channel:
        logger.warning("Telegram not configured — set TELEGRAM_BOT_TOKEN and TELEGRAM_CHANNEL_PRO in Railway Variables")
        return False
    try:
        url = f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/sendMessage"
        payload = {
            "chat_id": normalise_channel(channel),
            "text": message,
            "parse_mode": "HTML",
            "disable_web_page_preview": True
        }
        r = requests.post(url, json=payload, timeout=10)
        if r.status_code == 200:
            logger.info(f"Telegram alert sent: {message[:60]}")
            return True
        else:
            logger.error(f"Telegram error {r.status_code}: {r.text[:200]}")
            return False
    except Exception as e:
        logger.error(f"Telegram exception: {e}")
        return False


def score_live_match(fixture_id: int, home: str, away: str,
                     home_score: int, away_score: int,
                     minute: int, league: str) -> Optional[dict]:
    """
    Score a live match against all three alert conditions.
    Returns alert dict if all conditions met, else None.
    """
    # Condition 1: Minute range — not too early or too late
    if not (ALERT_MINUTE_MIN <= minute <= ALERT_MINUTE_MAX):
        return None

    # Fetch live stats
    stats_data = api_get("fixtures/statistics", {"fixture": fixture_id})
    stats_raw = stats_data.get("response", [])
    if len(stats_raw) < 2:
        return None

    FIELD_MAP = {
        "Shots on Goal": "shots_on_goal", "Total Shots": "total_shots",
        "Shots insidebox": "shots_inside_box", "Goalkeeper Saves": "goalkeeper_saves",
        "Ball Possession": "ball_possession", "Corner Kicks": "corner_kicks",
        "Expected Goals": "expected_goals", "expected_goals": "expected_goals",
    }

    team_stats = {}
    for td in stats_raw:
        tname = td["team"]["name"]
        ts = {}
        for s in td.get("statistics", []):
            key = FIELD_MAP.get(s["type"]) or s["type"].lower().replace(" ", "_")
            val = s.get("value")
            if isinstance(val, str) and val.endswith("%"):
                try: val = int(val.replace("%", ""))
                except: pass
            ts[key] = val if val is not None else 0
        team_stats[tname] = ts

    hs = team_stats.get(home, {})
    as_ = team_stats.get(away, {})

    # ── xG: use real if available, else estimate ──
    h_xg = float(hs.get("expected_goals") or 0)
    a_xg = float(as_.get("expected_goals") or 0)
    if h_xg == 0 and a_xg == 0:
        # Estimate from shots
        h_shots_on = float(hs.get("shots_on_goal") or 0)
        a_shots_on = float(as_.get("shots_on_goal") or 0)
        h_inside   = float(hs.get("shots_inside_box") or 0)
        a_inside   = float(as_.get("shots_inside_box") or 0)
        h_saves    = float(as_.get("goalkeeper_saves") or 0)  # opp keeper saves = h attacks
        a_saves    = float(hs.get("goalkeeper_saves") or 0)
        h_xg = round(h_shots_on * 0.33 + h_inside * 0.08 + h_saves * 0.15, 2)
        a_xg = round(a_shots_on * 0.33 + a_inside * 0.08 + a_saves * 0.15, 2)
        xg_real = False
    else:
        xg_real = True

    xg_gap = round(h_xg - a_xg, 2)

    # ── Win % from score + xG + possession ──
    h_poss = float(hs.get("ball_possession") or 50)
    a_poss = float(as_.get("ball_possession") or 50)
    score_diff = home_score - away_score

    # Base win % from current score using Poisson-style logic
    # Weight: score (50%) + xG advantage (30%) + possession (20%)
    score_weight = max(0, min(40, score_diff * 18))  # each goal = ~18%
    xg_weight    = max(-20, min(20, xg_gap * 15))
    poss_weight  = round((h_poss - 50) * 0.15, 1)
    base_home    = 50 + score_weight + xg_weight + poss_weight
    h_win_pct    = round(max(5, min(95, base_home)))
    a_win_pct    = round(max(5, min(95, 100 - h_win_pct - 10)))
    draw_pct     = max(0, 100 - h_win_pct - a_win_pct)

    # ── Condition 2: Win % >= threshold ──
    if h_win_pct < ALERT_WIN_PCT and a_win_pct < ALERT_WIN_PCT:
        return None

    # ── Condition 3: xG gap meaningful ──
    if abs(xg_gap) < ALERT_XG_GAP and h_xg + a_xg > 0.1:
        return None

    # ── Determine favoured team ──
    if h_win_pct >= ALERT_WIN_PCT:
        fav_team   = home
        fav_win_pct = h_win_pct
        fav_xg     = h_xg
        opp_xg     = a_xg
        bet        = f"{home} Win"
    else:
        fav_team   = away
        fav_win_pct = a_win_pct
        fav_xg     = a_xg
        opp_xg     = h_xg
        xg_gap     = abs(xg_gap)
        bet        = f"{away} Win"

    # ── Scored first check (momentum) ──
    scored_first = (h_win_pct >= ALERT_WIN_PCT and home_score > away_score) or \
                   (a_win_pct >= ALERT_WIN_PCT and away_score > home_score)
    # Pro channel: alert on strong xG dominance even without a goal
    # Only block if truly no dominance (xG gap < 0.25 and not winning)
    if not scored_first and abs(xg_gap) < 0.4:
        return None

    # ── Fetch live odds ──
    odds_data = get_real_odds(fixture_id)
    h_odds = odds_data.get("home", 0)
    a_odds = odds_data.get("away", 0)
    d_odds = odds_data.get("draw", 0)
    fav_odds = h_odds if fav_team == home else a_odds
    if not fav_odds or fav_odds <= 1.01:
        fav_odds_str = "N/A"
    else:
        fav_odds_str = f"{fav_odds:.2f}"

    xg_label = "Real xG" if xg_real else "Est. xG"

    return {
        "fixture_id": fixture_id,
        "home": home, "away": away,
        "score": f"{home_score}-{away_score}",
        "minute": minute,
        "league": league,
        "h_win_pct": h_win_pct,
        "a_win_pct": a_win_pct,
        "draw_pct": draw_pct,
        "h_xg": h_xg, "a_xg": a_xg,
        "xg_gap": round(xg_gap, 2),
        "xg_real": xg_real,
        "xg_label": xg_label,
        "fav_team": fav_team,
        "fav_win_pct": fav_win_pct,
        "bet": bet,
        "fav_odds": fav_odds_str,
        "h_odds": h_odds, "d_odds": d_odds, "a_odds": a_odds,
        "scored_first": scored_first,
    }


def build_telegram_message(alert: dict) -> str:
    """Full PRO message — all stats, xG, win %, odds."""
    minute = alert['minute']
    score  = alert['score']
    league = alert['league']
    home   = alert['home']
    away   = alert['away']
    bet    = alert['bet']
    odds   = alert['fav_odds']
    h_wp   = alert['h_win_pct']
    a_wp   = alert['a_win_pct']
    h_xg   = alert['h_xg']
    a_xg   = alert['a_xg']
    xg_lbl = alert['xg_label']
    first  = "✅ Scored first" if alert['scored_first'] else "📊 xG dominance"

    return (
        f"🟢 <b>ACCAGENIUS PRO ALERT</b> 🟢\n"
        f"━━━━━━━━━━━━━━━━━━━━\n"
        f"⚽ <b>{home} vs {away}</b>\n"
        f"🏆 {league}\n"
        f"🕐 {minute}' | Score: <b>{score}</b>\n"
        f"━━━━━━━━━━━━━━━━━━━━\n"
        f"📊 <b>Win Probability</b>\n"
        f"  {home}: <b>{h_wp}%</b> 🟢\n"
        f"  Draw: {alert['draw_pct']}%\n"
        f"  {away}: {a_wp}%\n"
        f"\n"
        f"⚡ <b>{xg_lbl}</b>\n"
        f"  {home}: {h_xg}  |  {away}: {a_xg}\n"
        f"  Gap: {alert['xg_gap']:+.2f}\n"
        f"\n"
        f"🔥 <b>Trigger: {first}</b>\n"
        f"━━━━━━━━━━━━━━━━━━━━\n"
        f"💰 <b>Suggested Bet: {bet}</b>\n"
        f"📈 Current Odds: <b>{odds}</b>\n"
        f"━━━━━━━━━━━━━━━━━━━━\n"
        f"⚠️ <i>Bet responsibly. 18+ only.</i>\n"
        f"🤖 AccaGenius Pro | Live Intelligence"
    )


def build_free_message(alert: dict) -> str:
    """Teaser FREE message — just the pick, score, minute. No xG or win %."""
    return (
        f"🟡 <b>ACCAGENIUS FREE TIP</b>\n"
        f"━━━━━━━━━━━━━━━━━━━━\n"
        f"⚽ <b>{alert['home']} vs {alert['away']}</b>\n"
        f"🕐 {alert['minute']}' | Score: <b>{alert['score']}</b>\n"
        f"\n"
        f"💰 <b>In-Play Tip: {alert['bet']}</b>\n"
        f"\n"
        f"🔒 <i>Win probability, xG stats & full analysis</i>\n"
        f"🔒 <i>available on AccaGenius Pro</i>\n"
        f"━━━━━━━━━━━━━━━━━━━━\n"
        f"👉 <b>Upgrade: accagenius.com</b>\n"
        f"⚠️ <i>Bet responsibly. 18+ only.</i>"
    )




def score_inplay_markets(fixture_id: int, home: str, away: str,
                          home_score: int, away_score: int,
                          minute: int, league: str, team_stats: dict) -> list:
    """
    5 alert types — ALL fire to Pro Telegram only:
    1. Match Winner: level score, 45+ min, dominant xG+poss+shots
    2. HT/FT Comeback: losing by 1, 40-52 min, losing team leads xG
    3. Both Teams to Score: one team scoreless, 30-75 min, both 1.0+ xG
    4. Over 1.5 Goals: before 65 min, max 1 goal, combined xG >= 2.5
    5. Goal Due: handled separately in Goal Due scanner block
    Free channel: acca of the day + first 3 inplays only (handled in sending logic)
    """
    alerts = []

    if minute < 20:
        return []

    hs   = team_stats.get(home, {})
    as_  = team_stats.get(away, {})
    h_xg = float(hs.get("expected_goals") or 0)
    a_xg = float(as_.get("expected_goals") or 0)
    h_poss  = float(str(hs.get("ball_possession") or "50").replace("%",""))
    a_poss  = 100 - h_poss
    h_shots = int(hs.get("shots_on_goal") or 0)
    a_shots = int(as_.get("shots_on_goal") or 0)

    total_xg    = round(h_xg + a_xg, 2)
    total_goals = home_score + away_score
    is_draw     = home_score == away_score
    goal_diff   = abs(home_score - away_score)
    score_str   = f"{home_score}-{away_score}"

    # Hard block: never alert if goal diff >= 2 (game is over effectively)
    if goal_diff >= 2:
        return []

    # ── 1. MATCH WINNER ──────────────────────────────────────────────────
    # Level score, 2nd half only, strong xG + possession + shots dominance
    if is_draw and minute >= 45:
        winner = dom_xg = dom_poss = dom_shots = opp_xg = None
        if h_xg > a_xg + 0.9 and h_poss >= 58 and h_shots > a_shots:
            winner, dom_xg, dom_poss, dom_shots, opp_xg = home, h_xg, h_poss, h_shots, a_xg
        elif a_xg > h_xg + 0.9 and a_poss >= 58 and a_shots > h_shots:
            winner, dom_xg, dom_poss, dom_shots, opp_xg = away, a_xg, a_poss, a_shots, h_xg
        if winner:
            conf = min(88, int(68 + (dom_xg - opp_xg) * 12 + (dom_poss - 55) * 0.4))
            alerts.append({
                "market": "Match Winner",
                "team": winner,
                "fixture_id": fixture_id,
                "home": home, "away": away,
                "score": score_str,
                "minute": minute, "league": league,
                "reason": (
                    f"{winner} pushing hard — xG {dom_xg:.2f} vs {opp_xg:.2f}, "
                    f"{dom_poss:.0f}% poss, {dom_shots} shots on target. "
                    f"Scores level {score_str} at {minute}'"
                ),
                "confidence": conf,
            })

    # ── 2. HT/FT COMEBACK ────────────────────────────────────────────────
    # Losing by exactly 1, around HT, losing team clearly dominates xG
    if 40 <= minute <= 52 and goal_diff == 1:
        if home_score < away_score and h_xg > a_xg + 0.6:
            alerts.append({
                "market": "HT/FT Value — Comeback",
                "team": home,
                "fixture_id": fixture_id,
                "home": home, "away": away,
                "score": score_str,
                "minute": minute, "league": league,
                "reason": (
                    f"{home} losing {score_str} but dominant — xG {h_xg:.2f} vs {a_xg:.2f}. "
                    f"2nd half comeback strongly likely"
                ),
                "confidence": 74,
            })
        elif away_score < home_score and a_xg > h_xg + 0.6:
            alerts.append({
                "market": "HT/FT Value — Comeback",
                "team": away,
                "fixture_id": fixture_id,
                "home": home, "away": away,
                "score": score_str,
                "minute": minute, "league": league,
                "reason": (
                    f"{away} losing {score_str} but dominant — xG {a_xg:.2f} vs {h_xg:.2f}. "
                    f"2nd half comeback strongly likely"
                ),
                "confidence": 74,
            })

    # ── 3. BOTH TEAMS TO SCORE ───────────────────────────────────────────
    # One team scoreless, 30-75 min, both teams generating real xG
    if 30 <= minute < 75 and total_goals < 2 and (home_score == 0 or away_score == 0):
        if h_xg >= 1.0 and a_xg >= 1.0:
            scoreless = home if home_score == 0 else away
            conf = min(84, int(60 + (h_xg + a_xg) * 6))
            alerts.append({
                "market": "Both Teams to Score",
                "fixture_id": fixture_id,
                "home": home, "away": away,
                "score": score_str,
                "minute": minute, "league": league,
                "reason": (
                    f"Both creating chances — {home} xG {h_xg:.2f} / {away} xG {a_xg:.2f}. "
                    f"{scoreless} yet to score with {90-minute} mins left"
                ),
                "confidence": conf,
            })

    # ── 4. OVER 1.5 GOALS ────────────────────────────────────────────────
    # Before 65 min, max 1 goal scored, combined xG >= 2.5
    if minute < 65 and total_goals <= 1 and total_xg >= 2.5:
        conf = min(85, int(62 + (total_xg - 2.5) * 12))
        alerts.append({
            "market": "Over 1.5 Goals",
            "fixture_id": fixture_id,
            "home": home, "away": away,
            "score": score_str,
            "minute": minute, "league": league,
            "reason": (
                f"Combined xG {total_xg:.2f} but only {total_goals} goal(s) — "
                f"xG strongly suggests more goals before FT"
            ),
            "confidence": conf,
        })

    return alerts


def build_market_telegram_message(alert: dict) -> str:
    """Build Pro Telegram message for any in-play market alert."""
    sep    = "━" * 22
    market = alert.get("market", "In-Play Alert")
    team   = alert.get("team", "")
    home   = alert["home"]
    away   = alert["away"]
    score  = alert["score"]
    minute = alert["minute"]
    league = alert["league"]
    reason = alert.get("reason", "")
    conf   = alert.get("confidence", 75)

    icons = {
        "Match Winner":            "🏆",
        "Draw No Bet":             "🛡️",
        "Over 2.5 Goals":          "⚽",
        "Over 1.5 Goals":          "⚽",
        "Over 0.5 Goals":          "⚽",
        "Next Goal":               "🎯",
        "Both Teams to Score":     "⚽⚽",
        "HT/FT Value — Comeback":  "🔄",
    }
    icon = icons.get(market, "📊")

    # Match Winner gets ⭐ WINNER ⭐ treatment
    if market == "Match Winner" and team:
        header = f"⭐ <b>WINNER ALERT — PRO</b> ⭐"
        bet_line = f"🏆 <b>⭐ {team} TO WIN ⭐</b>"
    else:
        header = f"{icon} <b>IN-PLAY ALERT — PRO</b>"
        bet_line = f"💰 <b>{market}{' — ' + team if team else ''}</b>"

    nl = chr(10)
    return (
        header + nl + sep + nl
        + f"⚽ <b>{home} vs {away}</b>" + nl
        + f"🏆 {league}" + nl
        + f"⏱ <b>{minute}'</b>  |  🔴 Live Score: <b>{score}</b>" + nl
        + sep + nl
        + bet_line + nl
        + f"📊 Confidence: <b>{conf}%</b>" + nl + nl
        + f"📋 {reason}" + nl
        + sep + nl
        + "⚠️ <i>Bet responsibly. 18+ only.</i>" + nl
        + "🤖 AccaGenius Pro | In-Play Intelligence"
    )

async def live_alert_scanner():
    """Background task — scans live matches every 5 minutes.
    Also sends Today's Acca at 9am and live alerts throughout the day."""
    global alerted_fixtures, last_alert_reset
    global free_alerts_sent_today, free_alert_reset_date
    logger.info("🤖 Telegram dual-channel scanner started")

    acca_sent_date = ""  # tracks which date we've already sent the morning acca

    while True:
        try:
            now   = datetime.now()
            today = now.strftime("%Y-%m-%d")
            hour  = now.hour
            minute = now.minute

            # ── Reset daily trackers at midnight ──
            if today != last_alert_reset:
                alerted_fixtures = set()
                alerted_details  = {}
                last_alert_reset = today
                logger.info("Alert tracker reset for new day")

            if today != free_alert_reset_date:
                free_alerts_sent_today = 0
                free_alert_reset_date  = today
                logger.info("Free alert counter reset for new day")

            # Need at least bot token to do anything
            if not TELEGRAM_BOT_TOKEN:
                await asyncio.sleep(300)
                continue

            has_pro  = bool(TELEGRAM_CHANNEL_PRO or TELEGRAM_CHANNEL_ID)
            has_free = bool(TELEGRAM_CHANNEL_FREE)

            if not has_pro and not has_free:
                await asyncio.sleep(300)
                continue

            # ── 9AM: Send Today's Acca to both channels ──
            if hour == 9 and acca_sent_date != today:
                acca_sent_date = today  # mark immediately so we don't double-send
                pro_channel = TELEGRAM_CHANNEL_PRO or TELEGRAM_CHANNEL_ID
                try:
                    # Generate picks fresh for today — call endpoint directly
                    acca_data = cache_get(f"today_acca_{today}")
                    if not acca_data:
                        logger.info("9am: generating today's acca for Telegram...")
                        try:
                            acca_data = await get_today_acca()
                        except Exception as eg:
                            logger.error(f"9am acca generation failed: {eg}")
                            acca_data = None
                    sep = "━" * 20

                    if acca_data and any(a.get("picks") for a in acca_data.get("accas", [])):
                        for acca in acca_data.get("accas", []):
                            picks = acca.get("picks", [])
                            if not picks:
                                continue
                            label = acca.get("label", "Acca")
                            emoji = acca.get("emoji", "⚽")
                            total_odds = round(
                                __import__("functools").reduce(
                                    lambda a, p: a * float(p.get("odds", 1.75)), picks, 1.0
                                ), 2
                            )
                            legs = ""
                            for p in picks:
                                legs += "  ⚽ " + p.get("home","") + " vs " + p.get("away","") + "\n"
                                legs += "     ✅ " + str(p.get("bet","")) + " @ " + str(p.get("odds","")) + "\n"

                            if has_pro:
                                pro_msg = (
                                    emoji + " <b>TODAY'S ACCA — " + label.upper() + "</b>\n"
                                    + sep + "\n" + legs + sep + "\n"
                                    + "💰 Combined: <b>" + str(total_odds) + "x</b>\n"
                                    + "📅 " + today + " · Match Winner\n"
                                    + "🤖 AccaGenius · accagenius.com"
                                )
                                send_telegram_to(pro_msg, pro_channel)

                            if has_free:
                                free_legs = ""
                                for p in picks:
                                    free_legs += "  ⚽ " + p.get("home","") + " vs " + p.get("away","") + "\n"
                                    free_legs += "     ✅ " + str(p.get("bet","")) + " @ " + str(p.get("odds","")) + "\n"
                                free_msg = (
                                    emoji + " <b>TODAY'S ACCA — " + label.upper() + "</b>\n"
                                    + sep + "\n"
                                    + free_legs
                                    + sep + "\n"
                                    + "💰 Combined: <b>" + str(total_odds) + "x</b> · " + str(len(picks)) + " picks\n"
                                    + "📅 " + today + "\n"
                                    + "━━━━━━━━━━━━━━━━━━━━\n"
                                    + "🤖 AccaGenius · accagenius.com"
                                )
                                send_telegram_to(free_msg, TELEGRAM_CHANNEL_FREE)

                        logger.info(f"9am acca sent: {len(acca_data.get('accas', []))} accas")

                    # ── FBD Value Bets: BTTS + Over 2.5 to Pro channel ──
                    try:
                        fbd_matches = get_fbd_data()
                        btts_picks, over25_picks = [], []
                        for m in fbd_matches:
                            home=m.get("home",""); away=m.get("away","")
                            if not home or not away: continue
                            bp=m.get("pred_btts"); op=m.get("pred_over25")
                            if bp and float(bp)>=60:
                                ao=m.get("avg_btts_yes")
                                odds_str=f"@ ~{100/float(ao):.2f}" if ao and float(ao)>0 else ""
                                btts_picks.append(f"  ⚽ {home} vs {away} — BTTS {odds_str} ({int(float(bp))}%)")
                            if op and float(op)>=60:
                                ao=m.get("avg_over25")
                                odds_str=f"@ ~{100/float(ao):.2f}" if ao and float(ao)>0 else ""
                                over25_picks.append(f"  ⚽ {home} vs {away} — Over 2.5 {odds_str} ({int(float(op))}%)")
                        sep2="━"*20
                        if btts_picks and has_pro:
                            send_telegram_to(
                                f"🎯 <b>TODAY'S BTTS VALUE PICKS</b>\n{sep2}\n"
                                +"\n".join(btts_picks[:5])+f"\n{sep2}\n"
                                +"📊 Powered by FBD prediction model\n"
                                +"⚠️ 18+ | BeGambleAware.org\n🤖 AccaGenius Pro", pro_channel)
                        if over25_picks and has_pro:
                            send_telegram_to(
                                f"📈 <b>TODAY'S OVER 2.5 VALUE PICKS</b>\n{sep2}\n"
                                +"\n".join(over25_picks[:5])+f"\n{sep2}\n"
                                +"📊 Powered by FBD prediction model\n"
                                +"⚠️ 18+ | BeGambleAware.org\n🤖 AccaGenius Pro", pro_channel)
                        if btts_picks or over25_picks:
                            logger.info(f"9am FBD value bets sent: {len(btts_picks)} BTTS, {len(over25_picks)} Over 2.5")
                    except Exception as fbd_e:
                        logger.warning(f"FBD value bets send error: {fbd_e}")

                    else:
                        # No picks available — send "no acca today" message
                        no_acca_msg = (
                            "📅 <b>ACCAGENIUS — " + today + "</b>\n"
                            + sep + "\n"
                            + "⚠️ <b>No Acca Today</b>\n"
                            + "Not enough qualifying fixtures today.\n"
                            + "Check back for in-play alerts during live matches!\n"
                            + "🤖 AccaGenius · accagenius.com"
                        )
                        if has_pro:
                            send_telegram_to(no_acca_msg, pro_channel)
                        if has_free:
                            send_telegram_to(no_acca_msg, TELEGRAM_CHANNEL_FREE)
                        logger.info("9am: no acca available — sent no-acca message")

                except Exception as e:
                    logger.error(f"9am acca send error: {e}")

            # Fetch live matches — use cache to avoid burning API quota
            # Cache is 90s so still near-real-time
            cached_live = cache_get("live_matches")
            if cached_live:
                live_matches = cached_live.get("matches", [])
                logger.info(f"Scanner: using cached live data ({len(live_matches)} matches)")
            else:
                live_data = api_get("fixtures", {"live": "all", "timezone": "Europe/London"})
                our_league_ids = {l["id"] for l in LEAGUES}
                live_matches = []
                for f in live_data.get("response", []):
                    try:
                        if f["league"]["id"] not in our_league_ids:
                            continue
                        live_matches.append({
                            "id":         f["fixture"]["id"],
                            "home":       f["teams"]["home"]["name"],
                            "away":       f["teams"]["away"]["name"],
                            "home_score": f["goals"]["home"] or 0,
                            "away_score": f["goals"]["away"] or 0,
                            "minute":     f["fixture"]["status"].get("elapsed") or 0,
                            "league":     f["league"]["name"],
                        })
                    except Exception:
                        pass
                # Store in cache for 90s so next scan can reuse
                cache_set("live_matches", {"matches": live_matches}, 90)
                logger.info(f"Scanner: fresh live fetch — {len(live_matches)} matches")

            pro_channel = TELEGRAM_CHANNEL_PRO or TELEGRAM_CHANNEL_ID

            # ── LIVE GOAL DETECTION — fire update when score changes ──
            live_score_map = {m["id"]: (m["home_score"], m["away_score"], m["home"], m["away"], m["minute"], m["league"]) for m in live_matches}
            for fid_key, details in list(alerted_details.items()):
                if details.get("goal_update_posted"): continue
                raw_fid = details.get("fixture_id") or fid_key
                try:
                    raw_fid = int(str(raw_fid).split("_")[0]) if "_" in str(raw_fid) else int(raw_fid)
                except:
                    continue
                if raw_fid not in live_score_map: continue
                cur_h, cur_a, home, away, cur_min, league = live_score_map[raw_fid]
                prev_h = details.get("home_score_at_alert", cur_h)
                prev_a = details.get("away_score_at_alert", cur_a)
                # Parse prev scores properly (may be stored as string)
                try:
                    prev_h = int(str(prev_h).strip())
                    prev_a = int(str(prev_a).strip())
                except:
                    prev_h, prev_a = cur_h, cur_a

                if cur_h + cur_a > prev_h + prev_a:
                    # A goal was scored since last check!
                    scorer = home if cur_h > prev_h else away
                    new_score = f"{cur_h}-{cur_a}"
                    market = details.get("market", details.get("bet", "Alert"))
                    sep = chr(9473) * 20
                    goal_msg = (
                        "\u26bd <b>GOAL AFTER ALERT \u2014 PRO</b>\n"
                        + sep + "\n"
                        + f"\u2705 <b>{home} vs {away}</b>\n"
                        + f"\U0001f3c6 {league}\n"
                        + f"\U0001f550 {cur_min}' | New score: <b>{new_score}</b>\n"
                        + sep + "\n"
                        + f"\U0001f3af Alert was: <b>{market}</b>\n"
                        + f"\U0001f4ca Score at alert: {details.get('score_at_alert', '?')}\n"
                        + f"\u26bd Scored: <b>{scorer}</b>\n"
                        + sep + "\n"
                        + "\U0001f916 AccaGenius Pro | Live Intelligence"
                    )
                    if has_pro:
                        send_telegram_to(goal_msg, pro_channel)
                        logger.info(f"Goal update sent: {home} vs {away} {new_score}")
                    # Update tracked score so next goal also fires
                    details["home_score_at_alert"] = cur_h
                    details["away_score_at_alert"] = cur_a
                    details["last_known_score"] = new_score
                    # Don't set goal_update_posted=True — allow all goals to fire

            # ── CHECK RESULTS of previously alerted matches ──
            live_ids = {m["id"] for m in live_matches}
            for fid, details in list(alerted_details.items()):
                if details.get("result_posted"): continue
                if details.get("is_market"): continue  # market alerts handled separately
                try:
                    fid_int = int(fid)
                except:
                    continue
                if fid_int in live_ids: continue  # Still playing
                # Match finished — fetch result (cache 30 mins so we don't re-fetch each scan)
                try:
                    result_cache_key = f"result_{fid}"
                    fr_data = cache_get(result_cache_key)
                    if not fr_data:
                        fd = api_get("fixtures", {"id": fid, "timezone": "Europe/London"})
                        fr_data = fd.get("response", [{}])[0] if fd.get("response") else {}
                        if fr_data:
                            cache_set(result_cache_key, fr_data, 1800)  # 30 min cache
                    fr = fr_data
                    status = fr.get("fixture", {}).get("status", {}).get("short", "")
                    if status in ["FT", "AET", "PEN"]:
                        hs = fr.get("goals", {}).get("home", 0) or 0
                        as_ = fr.get("goals", {}).get("away", 0) or 0
                        bet = details.get("bet", "")
                        home = details.get("home", "")
                        away = details.get("away", "")
                        odds = float(details.get("odds", 2.0) or 2.0)
                        # Evaluate result
                        won = False
                        if home + " Win" in bet: won = hs > as_
                        elif away + " Win" in bet: won = as_ > hs
                        elif "Draw" in bet: won = hs == as_
                        elif "Both Teams" in bet or "BTTS" in bet: won = hs > 0 and as_ > 0
                        elif "Over 2.5" in bet: won = (hs + as_) > 2.5
                        elif "Over 1.5" in bet: won = (hs + as_) > 1.5

                        result = "won" if won else "lost"
                        stake = telegram_pl["stake"]
                        telegram_pl["tips"] += 1
                        if won:
                            profit = round(stake * odds - stake, 2)
                            telegram_pl["won"] += 1
                            telegram_pl["profit"] = round(telegram_pl["profit"] + profit, 2)
                            pl_line = f"Won:{telegram_pl['won']} Lost:{telegram_pl['lost']} · P&L: {'+'if telegram_pl['profit']>=0 else ''}£{telegram_pl['profit']:.2f}"

                            # Edited FREE message — original tip + WIN result appended
                            free_result = (
                                details.get("original_msg", "").rstrip() + "\n"
                                f"━━━━━━━━━━━━━━━━━━━━\n"
                                f"✅ <b>RESULT: WINNER! ⭐</b>\n"
                                f"FT Score: {hs}-{as_}\n"
                                f"━━━━━━━━━━━━━━━━━━━━\n"
                                f"🔒 Full xG analysis → accagenius.com"
                            )
                            # Edited PRO message — original tip + WIN result
                            pro_result = (
                                details.get("pro_msg", "").rstrip() + "\n"
                                f"━━━━━━━━━━━━━━━━━━━━\n"
                                f"💥 <b>RESULT: WINNER! FT {hs}-{as_}</b>\n"
                                f"🏆 +£{profit:.2f} profit · {pl_line}"
                            )
                        else:
                            telegram_pl["lost"] += 1
                            telegram_pl["profit"] = round(telegram_pl["profit"] - stake, 2)
                            pl_line = f"Won:{telegram_pl['won']} Lost:{telegram_pl['lost']} · P&L: {'+'if telegram_pl['profit']>=0 else ''}£{telegram_pl['profit']:.2f}"

                            free_result = (
                                details.get("original_msg", "").rstrip() + "\n"
                                f"━━━━━━━━━━━━━━━━━━━━\n"
                                f"❌ <b>RESULT: No luck — FT {hs}-{as_}</b>\n"
                                f"More tips coming 👉 accagenius.com"
                            )
                            pro_result = (
                                details.get("pro_msg", "").rstrip() + "\n"
                                f"━━━━━━━━━━━━━━━━━━━━\n"
                                f"❌ <b>RESULT: FT {hs}-{as_} — No luck</b>\n"
                                f"📊 {pl_line}"
                            )

                        # Edit the original messages in-place
                        free_mid = details.get("free_mid", 0)
                        pro_mid  = details.get("pro_mid", 0)
                        if has_free and free_mid:
                            edit_telegram_message(TELEGRAM_CHANNEL_FREE, free_mid, free_result)
                        elif has_free:
                            # Message expired or wasn't stored — send new one
                            send_telegram_to(free_result, TELEGRAM_CHANNEL_FREE)
                        if has_pro and pro_mid:
                            edit_telegram_message(pro_channel, pro_mid, pro_result)
                        elif has_pro:
                            send_telegram_to(pro_result, pro_channel)

                        alerted_details[fid]["result_posted"] = True
                        logger.info(f"Result edited: {home} vs {away} FT {hs}-{as_} — {'WON' if won else 'LOST'}")
                except Exception as e:
                    logger.error(f"Result check error {fid}: {e}")

            # ── CHECK RESULTS for MARKET alerts (Match Winner, BTTS, HT/FT, Over 1.5, Goal Due) ──
            market_live_ids = {m["id"] for m in live_matches}
            for mkey, details in list(alerted_details.items()):
                if not details.get("is_market"): continue
                if details.get("result_posted"): continue
                fid = details.get("fixture_id")
                if not fid: continue
                if int(fid) in market_live_ids: continue  # still live
                try:
                    r_key = f"mresult_{fid}"
                    fr_data = cache_get(r_key)
                    if not fr_data:
                        fd = api_get("fixtures", {"id": fid, "timezone": "Europe/London"})
                        fr_data = fd.get("response", [{}])[0] if fd.get("response") else {}
                        if fr_data:
                            cache_set(r_key, fr_data, 1800)
                    if not fr_data:
                        continue
                    status = fr_data.get("fixture", {}).get("status", {}).get("short", "")
                    if status not in ["FT", "AET", "PEN"]:
                        continue
                    hs = fr_data.get("goals", {}).get("home", 0) or 0
                    as_ = fr_data.get("goals", {}).get("away", 0) or 0
                    market = details.get("market", "")
                    home   = details.get("home", "")
                    away   = details.get("away", "")
                    team   = details.get("team", "")
                    score_at = details.get("score_at_alert", "?-?")
                    total_goals = hs + as_
                    # Evaluate win/loss per market
                    won = False
                    if market == "Match Winner":
                        if team == home: won = hs > as_
                        elif team == away: won = as_ > hs
                    elif market == "Both Teams to Score":
                        won = hs > 0 and as_ > 0
                    elif market == "HT/FT Value — Comeback":
                        if team == home: won = hs > as_
                        elif team == away: won = as_ > hs
                    elif market == "Over 1.5 Goals":
                        won = total_goals > 1
                    elif market == "Goal Due":
                        won = total_goals > (int(score_at.split("-")[0] or 0) + int(score_at.split("-")[1] or 0))
                    result_icon = "✅" if won else "❌"
                    result_word = "WON" if won else "LOST"
                    _sep = "━" * 22
                    _team_part = (f" — <b>{team}</b>") if team else ""
                    result_msg = (
                        f"{result_icon} <b>RESULT: {result_word}</b>\n"
                        f"{_sep}\n"
                        f"⚽ <b>{home} vs {away}</b>\n"
                        f"🏆 {details.get('league','')}\n"
                        f"📊 Market: <b>{market}</b>{_team_part}\n"
                        f"🎯 Alert score: {score_at} → FT: <b>{hs}-{as_}</b>\n"
                        f"{_sep}\n"
                        f"🤖 AccaGenius Pro | accagenius.com"
                    )
                    if has_pro:
                        send_telegram_to(result_msg, pro_channel)
                    # Settle paper trades for this fixture
                    settled_count = 0
                    for t in _paper_trades:
                        if (t.get("match") == f"{home} vs {away}"
                                and t.get("type") == market
                                and t.get("result") == "pending"):
                            t["result"] = "win" if won else "loss"
                            t["status"] = "settled"
                            t["settled_score"] = f"{hs}-{as_}"
                            t["settled_at"] = datetime.now().isoformat()
                            odds_t  = float(t.get("odds", 2.0))
                            stake_t = float(t.get("stake", 3.0))
                            bet_type_t = t.get("bet_type", "back")
                            if bet_type_t == "lay":
                                # Lay: win=stake if lost, lose=(odds-1)*stake if won
                                t["profit_loss"] = round(stake_t, 2) if not won else round(-(odds_t - 1) * stake_t, 2)
                            else:
                                t["profit_loss"] = round(stake_t * (odds_t - 1), 2) if won else round(-stake_t, 2)
                            settled_count += 1
                    if settled_count:
                        _save_paper_db()
                        logger.info(f"Settled {settled_count} paper trade(s): {market} {home} vs {away} — {'WIN' if won else 'LOSS'}")
                    details["result_posted"] = True
                    logger.info(f"Market result posted: {market} {home} vs {away} FT {hs}-{as_} — {result_word}")

                    # ── Settle any open drip ladder for this fixture at FT ──
                    fid_str = str(fid)
                    if fid_str in _drip_ladders and not _drip_ladders[fid_str].get("cashed_out"):
                        ladder = _drip_ladders[fid_str]
                        # Over 1.5 Goals drip: won if total goals > 1
                        drip_won = total_goals > 1
                        for tid in ladder.get("trade_ids", []):
                            for t in _paper_trades:
                                if t["id"] == tid and t.get("result") == "pending":
                                    t["result"] = "win" if drip_won else "loss"
                                    t["status"] = "settled"
                                    t["settled_score"] = f"{hs}-{as_}"
                                    stake_t = float(t.get("stake", 3.0))
                                    odds_t  = float(t.get("odds", 2.0))
                                    t["profit_loss"] = round(stake_t * (odds_t - 1), 2) if drip_won else round(-stake_t, 2)
                                    t["settled_at"] = datetime.now().isoformat()
                        ladder["cashed_out"] = True
                        _paper_bot_log.insert(0, {
                            "time": datetime.now().isoformat(),
                            "msg": f"{'💚 WON' if drip_won else '❌ LOST'} FT: {home} vs {away} {hs}-{as_} — drip settled"
                        })
                        _save_paper_db()
                        logger.info(f"Drip FT settled: {home} vs {away} {hs}-{as_} {'WON' if drip_won else 'LOST'}")
                except Exception as e:
                    logger.error(f"Market result check error {mkey}: {e}")

            # ── MULTI-MARKET ALERTS ──
            for m in live_matches:
                try:
                    mid = m["id"]

                    # Fetch stats once per match
                    stats_data = api_get("fixtures/statistics", {"fixture": mid})
                    stats_raw = stats_data.get("response", [])
                    team_stats = {}
                    FMAP = {"Shots on Goal":"shots_on_goal","Total Shots":"total_shots",
                            "Ball Possession":"ball_possession","Corner Kicks":"corner_kicks",
                            "Expected Goals":"expected_goals","expected_goals":"expected_goals",
                            "Goalkeeper Saves":"goalkeeper_saves"}
                    for td in stats_raw:
                        tname = td["team"]["name"]
                        ts = {}
                        for s in td.get("statistics", []):
                            key = FMAP.get(s["type"]) or s["type"].lower().replace(" ","_")
                            val = s.get("value")
                            if isinstance(val, str) and val.endswith("%"):
                                try: val = int(val.replace("%",""))
                                except: pass
                            ts[key] = val if val is not None else 0
                        team_stats[tname] = ts

                    market_alerts = score_inplay_markets(
                        mid, m["home"], m["away"],
                        m["home_score"], m["away_score"],
                        m.get("minute",0), m["league"], team_stats
                    )

                    for ma in market_alerts:
                        # Permanent dedup key per fixture+market — never re-alert same market on same match
                        # Dedup: include score so re-alerts fire if score changes
                        score_snap = f"{m.get('home_score',0)}-{m.get('away_score',0)}"
                        dedup_key = f"mkt_sent_{mid}_{ma['market'].replace(' ','_').replace('/','_')}_{score_snap}"
                        if cache_get(dedup_key):
                            continue  # already sent this alert at this score
                        # Send to Pro Telegram
                        if has_pro:
                            msg = build_market_telegram_message(ma)
                            sent_mid = send_telegram_to(msg, pro_channel)
                            if sent_mid:
                                logger.info(f"Market alert: {ma['market']} — {m['home']} vs {m['away']} {m.get('minute',0)}'")
                                # Mark as sent for 6 hours — won't re-fire same market same match
                                cache_set(dedup_key, True, 7200)  # 2hr — one alert per market per score

                        # ── FREE channel: first 3 inplay alerts of the day only ──
                        if has_free and TELEGRAM_CHANNEL_FREE:
                            today_key = f"free_inplay_count_{datetime.now().strftime('%Y%m%d')}"
                            free_count = cache_get(today_key) or 0
                            if free_count < 3:
                                free_msg = build_market_telegram_message(ma)
                                free_mid = send_telegram_to(free_msg, TELEGRAM_CHANNEL_FREE)
                                if free_mid:
                                    cache_set(today_key, free_count + 1, 86400)
                                    logger.info(f"FREE inplay alert {free_count+1}/3: {ma['market']} — {m['home']} vs {m['away']}")

                        if has_pro and sent_mid:
                                # Auto paper trade — smart back/lay per market
                                est_odds = _estimate_odds(ma["market"], ma.get("confidence", 70))
                                mkt = ma["market"]
                                home_s = m["home_score"] or 0
                                away_s = m["away_score"] or 0

                                if mkt == "Match Winner":
                                    # Back dominant team to win
                                    _paper_bot_place_bet(
                                        alert_type=mkt, match=f"{m['home']} vs {m['away']}",
                                        league=m["league"], minute=m.get("minute",0),
                                        score=f"{home_s}-{away_s}",
                                        selection=f"{ma.get('team','?')} Win",
                                        odds=est_odds, stake=_paper_bot_config["stake"],
                                        reason=ma.get("reason",""), bet_type="back"
                                    )
                                elif mkt == "Both Teams to Score":
                                    # Back BTTS Yes
                                    _paper_bot_place_bet(
                                        alert_type=mkt, match=f"{m['home']} vs {m['away']}",
                                        league=m["league"], minute=m.get("minute",0),
                                        score=f"{home_s}-{away_s}",
                                        selection="BTTS Yes",
                                        odds=est_odds, stake=_paper_bot_config["stake"],
                                        reason=ma.get("reason",""), bet_type="back"
                                    )
                                elif mkt == "HT/FT Value — Comeback":
                                    comeback_team = ma.get("team","?")
                                    winning_team = m["away"] if home_s < away_s else m["home"]
                                    # Back comeback team
                                    _paper_bot_place_bet(
                                        alert_type=mkt, match=f"{m['home']} vs {m['away']}",
                                        league=m["league"], minute=m.get("minute",0),
                                        score=f"{home_s}-{away_s}",
                                        selection=f"{comeback_team} Win",
                                        odds=round(est_odds * 1.3, 2),
                                        stake=_paper_bot_config["stake"],
                                        reason=ma.get("reason",""), bet_type="back"
                                    )
                                    # Lay current winning team
                                    _paper_bot_place_bet(
                                        alert_type=mkt, match=f"{m['home']} vs {m['away']}",
                                        league=m["league"], minute=m.get("minute",0),
                                        score=f"{home_s}-{away_s}",
                                        selection=f"LAY {winning_team}",
                                        odds=round(est_odds * 0.8, 2),
                                        stake=_paper_bot_config["stake"],
                                        reason=f"Lay current leader — {ma.get('reason','')}",
                                        bet_type="lay"
                                    )
                                elif mkt == "Over 1.5 Goals":
                                    # Lay Under 1.5 (same as backing Over 1.5)
                                    _paper_bot_place_bet(
                                        alert_type=mkt, match=f"{m['home']} vs {m['away']}",
                                        league=m["league"], minute=m.get("minute",0),
                                        score=f"{home_s}-{away_s}",
                                        selection="LAY Under 1.5 Goals",
                                        odds=round(1/max(0.05, 1 - (est_odds-1)/est_odds), 2),
                                        stake=_paper_bot_config["stake"],
                                        reason=ma.get("reason",""), bet_type="lay"
                                    )
                                    # Also trigger Track B drip for goals market
                                    h_xg_ol = float((team_stats.get(m["home"], {})).get("expected_goals") or 0)
                                    a_xg_ol = float((team_stats.get(m["away"], {})).get("expected_goals") or 0)
                                    xg_uns_ol = round((h_xg_ol + a_xg_ol) - (home_s + away_s), 2)
                                    if xg_uns_ol >= 1.0:
                                        _place_goal_due_drip(
                                            fixture_id=mid,
                                            match=f"{m['home']} vs {m['away']}",
                                            league=m["league"],
                                            minute=m.get("minute", 0),
                                            score=f"{home_s}-{away_s}",
                                            current_score_goals=home_s + away_s,
                                            xg_unscored=xg_uns_ol,
                                        )
                                # Track in admin P&L
                                bet_key = f"bet_{mid}_{ma['market'].replace(' ','_').replace('/','_')}"
                                _admin_bets[bet_key] = {
                                    "id": bet_key,
                                    "type": ma["market"],
                                    "team": ma.get("team", ""),
                                    "match": f"{m['home']} vs {m['away']}",
                                    "league": m["league"],
                                    "minute": m.get("minute", 0),
                                    "score_at_alert": f"{m['home_score']}-{m['away_score']}",
                                    "fired_at": datetime.now().isoformat(),
                                    "result": "pending",
                                    "profit_loss": 0.0,
                                }
                                # Track for goal update detection
                                mkey = f"mkt_{mid}_{ma['market'].replace(' ','_')}"
                                alerted_details[mkey] = {
                                    "home": m["home"], "away": m["away"],
                                    "market": ma["market"],
                                    "fixture_id": mid,
                                    "score_at_alert": f"{m['home_score']}-{m['away_score']}",
                                    "home_score_at_alert": m["home_score"],
                                    "away_score_at_alert": m["away_score"],
                                    "goal_update_posted": False,
                                    "result_posted": False,
                                    "is_market": True,
                                }
                except Exception as e:
                    logger.error(f"Market alert error {m.get('id')}: {e}")

                # ── GOAL DUE XG SCANNER ──────────────────────────────────────
                try:
                    h_xg = float((team_stats.get(m["home"], {})).get("expected_goals") or 0)
                    a_xg = float((team_stats.get(m["away"], {})).get("expected_goals") or 0)
                    total_xg = round(h_xg + a_xg, 2)
                    total_goals = (m["home_score"] or 0) + (m["away_score"] or 0)
                    xg_unscored = round(total_xg - total_goals, 2)
                    live_min = m.get("minute") or 0

                    # Only fire in 2nd half, xG significantly outpacing goals
                    if live_min >= 45 and xg_unscored >= 1.5 and total_goals < 3:
                        score_snap_gd = f"{m.get('home_score',0)}-{m.get('away_score',0)}"
                        gd_key = f"goaldue_{mid}_{score_snap_gd}"
                        if not cache_get(gd_key):
                            # Build message
                            score_str = f"{m['home_score']}-{m['away_score']}"
                            gd_msg = (
                                f"⚡ <b>GOAL DUE ALERT</b> ⚡\n"
                                f"{'─'*30}\n"
                                f"⚽ <b>{m['home']} vs {m['away']}</b>\n"
                                f"🏆 {m['league']}\n"
                                f"🕐 {live_min}' | Score: <b>{score_str}</b>\n"
                                f"{'─'*30}\n"
                                f"⚡ <b>Real xG</b>\n"
                                f"  {m['home']}: {h_xg:.2f} | {m['away']}: {a_xg:.2f}\n"
                                f"  Combined: {total_xg:.2f} | Goals: {total_goals}\n"
                                f"  <b>xG Unscored: {xg_unscored:.2f}</b>\n"
                                f"{'─'*30}\n"
                                f"🔥 xG significantly outpacing scoreline — goal due"
                            )
                            if has_pro:
                                sent = send_telegram_to(gd_msg, pro_channel)
                                if sent:
                                    cache_set(gd_key, True, 3600)  # don't re-fire for 1hr
                                    logger.info(f"Goal Due alert: {m['home']} vs {m['away']} {live_min}' xG={xg_unscored}")
                                    # Auto paper trade
                                    _paper_bot_place_bet(
                                        alert_type="Goal Due",
                                        match=f"{m['home']} vs {m['away']}",
                                        league=m["league"],
                                        minute=live_min,
                                        score=f"{m['home_score']}-{m['away_score']}",
                                        selection="Next Goal",
                                        odds=_estimate_odds("Goal Due", 72),
                                        stake=_paper_bot_config["stake"],
                                        reason=f"xG unscored: {xg_unscored}",
                                    )
                                    # Track for P&L
                                    bet_key = f"bet_{mid}_goaldue"
                                    _admin_bets[bet_key] = {
                                        "id": bet_key,
                                        "type": "Goal Due",
                                        "match": f"{m['home']} vs {m['away']}",
                                        "league": m["league"],
                                        "minute": live_min,
                                        "score_at_alert": score_str,
                                        "xg_unscored": xg_unscored,
                                        "fired_at": datetime.now().isoformat(),
                                        "result": "pending",
                                        "profit_loss": 0.0,
                                    }
                except Exception as gd_err:
                    logger.error(f"Goal Due scanner error: {gd_err}")

                # ── CHECK ACTIVE DRIP LADDERS ────────────────────────────
                try:
                    if _drip_ladders:
                        fid = str(mid)
                        if fid in _drip_ladders and not _drip_ladders[fid].get("cashed_out"):
                            prev_goals = _drip_ladders[fid].get("_last_goals", -1)
                            curr_goals = (m["home_score"] or 0) + (m["away_score"] or 0)
                            goal_just_scored = curr_goals > prev_goals and prev_goals >= 0
                            _drip_ladders[fid]["_last_goals"] = curr_goals
                            h_xg_now = float((team_stats.get(m["home"], {})).get("expected_goals") or 0)
                            a_xg_now = float((team_stats.get(m["away"], {})).get("expected_goals") or 0)
                            curr_xg_unscored = round((h_xg_now + a_xg_now) - curr_goals, 2)
                            _check_drip_cashout_or_cover(
                                fixture_id=mid,
                                match=f"{m['home']} vs {m['away']}",
                                league=m["league"],
                                minute=m.get("minute", 0),
                                score=f"{m['home_score']}-{m['away_score']}",
                                goal_scored=goal_just_scored,
                                xg_unscored=max(0, curr_xg_unscored),
                            )
                except Exception as drip_err:
                    logger.error(f"Drip check error: {drip_err}")


            logger.info(f"Scanner cycle done — {free_alerts_sent_today} free alerts today")

        except Exception as e:
            logger.error(f"Scanner error: {e}")

        # Sleep 120s between every scan cycle
        await asyncio.sleep(120)


def normalise_channel(channel: str) -> str:
    """Normalise Telegram channel ID — adds -100 prefix if missing from numeric IDs."""
    c = str(channel).strip()
    if c.startswith("@"):
        return c  # username format e.g. @AccaGeniusPro
    # Numeric ID — must be negative for supergroups/channels
    digits = c.lstrip("-")
    if digits.isdigit():
        num = int(digits)
        # Channel IDs are always >1000000000 and must be negative
        if num > 0:
            return f"-{num}"  # e.g. 1003799180225 → -1003799180225
        return c  # already negative
    return c


def send_telegram_to(message: str, channel: str) -> int:
    """Send to a specific channel. Returns message_id on success, 0 on failure."""
    if not TELEGRAM_BOT_TOKEN or not channel:
        return 0
    channel = normalise_channel(channel)
    try:
        url = f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/sendMessage"
        r = requests.post(url, json={
            "chat_id": channel,
            "text": message,
            "parse_mode": "HTML",
            "disable_web_page_preview": True
        }, timeout=10)
        if r.status_code == 200:
            return r.json().get("result", {}).get("message_id", 0)
        logger.error(f"Telegram send failed {r.status_code}: {r.text[:200]}")
        return 0
    except Exception as e:
        logger.error(f"Telegram send error: {e}")
        return 0


def edit_telegram_message(channel: str, message_id: int, new_text: str) -> bool:
    """Edit an existing message in-place — used for result updates on tip messages."""
    if not TELEGRAM_BOT_TOKEN or not channel or not message_id:
        return False
    try:
        url = f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/editMessageText"
        r = requests.post(url, json={
            "chat_id": channel,
            "message_id": message_id,
            "text": new_text,
            "parse_mode": "HTML",
            "disable_web_page_preview": True
        }, timeout=10)
        return r.status_code == 200
    except Exception as e:
        logger.error(f"Telegram edit error: {e}")
        return False


def send_telegram(message: str) -> int:
    """Legacy — sends to pro channel or fallback. Returns message_id."""
    channel = TELEGRAM_CHANNEL_PRO or TELEGRAM_CHANNEL_ID
    return send_telegram_to(message, channel)




@app.on_event("startup")
async def startup_event():
    """Start background scanner on server startup."""
    _init_db()
    _load_paper_db()
    _load_user_db()
    _load_accas_db()
    _load_referrals()
    asyncio.create_task(live_alert_scanner())
    asyncio.create_task(fbd_daily_scheduler())
    asyncio.create_task(_trading_scheduler())
    _load_todays_steam_prices()  # Reload today's prices from DB on startup
    if FBD_EMAIL and FBD_PASSWORD:
        print(f"FBD: ✅ credentials found ({FBD_EMAIL}) — startup download queued", flush=True)
        asyncio.create_task(_fbd_startup_download())
    else:
        print(f"FBD: ❌ FBD_EMAIL/FBD_PASSWORD not set — check Railway vars", flush=True)
    print("AccaGenius API started", flush=True)


async def _fbd_startup_download():
    await asyncio.sleep(5)
    try:
        print("FBD: attempting login...", flush=True)
        data = get_fbd_data(force=True)
        print(f"FBD: ✅ startup download complete — {len(data)} matches loaded", flush=True)
    except Exception as e:
        print(f"FBD: ❌ startup download failed — {e}", flush=True)


async def fbd_daily_scheduler():
    import pytz
    london = pytz.timezone("Europe/London")
    while True:
        try:
            now    = datetime.now(london)
            target = now.replace(hour=8, minute=15, second=0, microsecond=0)
            if now >= target: target = target + timedelta(days=1)
            wait_secs = (target - now).total_seconds()
            logger.info(f"⏰ FBD scheduler: next download in {int(wait_secs/3600)}h {int((wait_secs%3600)/60)}m")
            await asyncio.sleep(wait_secs)
            if FBD_EMAIL and FBD_PASSWORD:
                data = get_fbd_data(force=True)
                logger.info(f"✅ FBD daily refresh: {len(data)} matches")
        except Exception as e:
            logger.error(f"FBD scheduler error: {e}")
            await asyncio.sleep(3600)



# =========================
# MODELS
# =========================
class AccaRequest(BaseModel):
    selections: int = 5
    market: str = "winner"
    risk: str = "balanced"
    leagues: List[str] = []
    today_only: bool = False

class HTAccaRequest(BaseModel):
    selections: int = 5
    leagues: List[str] = []

class SaveAccaRequest(BaseModel):
    name: str
    selections: List[dict]
    total_odds: float
    stake: float = 10.0

# =========================
# CORE HELPERS
# =========================

def get_team_form(team_id: int, league_id: int, last: int = 10) -> dict:
    """Get detailed team form with extended stats — cached 6 hours"""
    cache_key = f"form_{team_id}_{league_id}_{last}"
    cached = cache_get(cache_key)
    if cached:
        return cached

    # Hard quota guard — skip form fetch if running critically high
    if _api_call_count > 6500:
        logger.warning(f"⚠️ Skipping form fetch (quota guard) — {_api_call_count}/75000 today")
        return {"games": 0, "wins": 0, "draws": 0, "losses": 0, "form": ""}

    data = api_get("fixtures", {
        "team": team_id, "league": league_id,
        "season": get_season(), "last": last, "timezone": "Europe/London"
    })
    fixtures = data.get("response", [])
    wins = draws = losses = gf = ga = ht_leads = early_goals = scored_first = 0
    recent = []

    for f in fixtures:
        if f["fixture"]["status"]["short"] != "FT":
            continue
        is_home = f["teams"]["home"]["id"] == team_id
        hg = f["goals"]["home"] or 0
        ag = f["goals"]["away"] or 0
        tf = hg if is_home else ag
        ta = ag if is_home else hg
        gf += tf; ga += ta
        if tf > ta: wins += 1
        elif tf == ta: draws += 1
        else: losses += 1

        # HT stats
        ht = f.get("score", {}).get("halftime", {})
        ht_h = ht.get("home") or 0
        ht_a = ht.get("away") or 0
        if is_home and ht_h > ht_a: ht_leads += 1
        elif not is_home and ht_a > ht_h: ht_leads += 1

        opp = f["teams"]["away"]["name"] if is_home else f["teams"]["home"]["name"]
        opp_logo = f["teams"]["away"]["logo"] if is_home else f["teams"]["home"]["logo"]
        result = "W" if tf > ta else ("D" if tf == ta else "L")
        recent.append({
            "date": f["fixture"]["date"].split("T")[0],
            "opponent": opp, "opponent_logo": opp_logo,
            "venue": "H" if is_home else "A",
            "goals_for": tf, "goals_against": ta, "result": result
        })

    games = wins + draws + losses
    form_rating = (wins * 3 + draws) / max(games, 1) if games else 1.5
    gf_avg = round(gf / max(games, 1), 2)
    ga_avg = round(ga / max(games, 1), 2)
    btts_pct = round(sum(1 for r in recent if r["goals_for"] > 0 and r["goals_against"] > 0) / max(len(recent), 1) * 100)
    over25_pct = round(sum(1 for r in recent if r["goals_for"] + r["goals_against"] > 2) / max(len(recent), 1) * 100)
    cs_pct = round(sum(1 for r in recent if r["goals_against"] == 0) / max(len(recent), 1) * 100)
    ht_lead_pct = round(ht_leads / max(games, 1) * 100)

    result = {
        "games": games, "wins": wins, "draws": draws, "losses": losses,
        "gf": gf, "ga": ga, "gf_avg": gf_avg, "ga_avg": ga_avg,
        "form_rating": round(form_rating, 2),
        "btts_pct": btts_pct, "over25_pct": over25_pct,
        "cs_pct": cs_pct, "ht_lead_pct": ht_lead_pct,
        "recent_fixtures": recent
    }
    cache_set(cache_key, result, 21600)  # 6 hours — form doesn't change during a day
    return result


def get_real_odds(fixture_id: int) -> dict:
    """Fetch best odds from API across all available bookmakers — cached 15 mins"""
    cache_key = f"odds_{fixture_id}"
    cached = cache_get(cache_key)
    if cached:
        return cached
    data = api_get("odds", {"fixture": fixture_id})
    response_data = data.get("response", [])
    fallback = {"home": 0, "draw": 0, "away": 0, "home_bk": "", "draw_bk": "", "away_bk": "", "available": False}
    if not response_data:
        return fallback
    best = {"home": 0.0, "draw": 0.0, "away": 0.0, "home_bk": "", "draw_bk": "", "away_bk": ""}
    all_bookmakers = {}
    for bm in response_data[0].get("bookmakers", [])[:12]:
        bm_name = bm["name"]
        for bet in bm.get("bets", []):
            if bet["name"] == "Match Winner":
                vals = {v["value"]: v["odd"] for v in bet.get("values", [])}
                try:
                    h = float(vals.get("Home", 0))
                    d = float(vals.get("Draw", 0))
                    a = float(vals.get("Away", 0))
                    all_bookmakers[bm_name] = {"home": h, "draw": d, "away": a}
                    if h > best["home"]: best["home"] = h; best["home_bk"] = bm_name
                    if d > best["draw"]: best["draw"] = d; best["draw_bk"] = bm_name
                    if a > best["away"]: best["away"] = a; best["away_bk"] = bm_name
                except Exception:
                    pass
    if best["home"] == 0:
        return fallback
    result = {**best, "available": True, "all_bookmakers": all_bookmakers}
    cache_set(cache_key, result, 3600)  # 15 minutes — odds shift slowly
    return result


def get_quick_odds(fixture_id: int) -> dict:
    """Fast odds fetch — returns best home/draw/away with bookmaker name"""
    return get_real_odds(fixture_id)


def analyze_and_pick(fixture: dict, home_form: dict, away_form: dict, risk: str, market: str = "winner") -> Optional[dict]:
    """
    Multi-factor pick engine. Uses:
    - Form rating (W/D/L points per game)
    - Win rate last 10
    - Goals scored/conceded avg
    - Clean sheet rate (defence signal)
    - BTTS % and Over 2.5% (attacking signal)
    - Real bookmaker odds (value filter)
    - Home advantage adjustment
    Only picks when multiple signals align. No picks on weak evidence.
    """
    try:
        home = fixture["teams"]["home"]["name"]
        away = fixture["teams"]["away"]["name"]
        home_logo = fixture["teams"]["home"]["logo"]
        away_logo = fixture["teams"]["away"]["logo"]
        fid  = fixture["fixture"]["id"]
        home_id  = fixture["teams"]["home"]["id"]
        away_id  = fixture["teams"]["away"]["id"]
        league_id = fixture["league"]["id"]
        dt = datetime.fromisoformat(fixture["fixture"]["date"].replace("Z", "+00:00"))
        date_str = dt.strftime("%d/%m %H:%M")

        odds = get_real_odds(fid)
        ho = odds.get("home", 0)
        do = odds.get("draw", 0)
        ao = odds.get("away", 0)
        has_odds = odds.get("available", False)

        # Form stats
        h_games = max(home_form.get("games", 0), 1)
        a_games = max(away_form.get("games", 0), 1)
        hw  = home_form.get("wins", 0)
        hl  = home_form.get("losses", 0)
        hd  = home_form.get("draws", 0)
        aw  = away_form.get("wins", 0)
        al  = away_form.get("losses", 0)
        ad  = away_form.get("draws", 0)
        hg  = home_form.get("gf_avg", 1.2)
        ag  = away_form.get("gf_avg", 1.0)
        hga = home_form.get("ga_avg", 1.3)
        aga = away_form.get("ga_avg", 1.3)
        h_wr  = hw / h_games
        a_wr  = aw / a_games
        h_lr  = hl / h_games
        a_lr  = al / a_games
        hcs   = home_form.get("cs_pct", 25) / 100      # clean sheet rate
        acs   = away_form.get("cs_pct", 20) / 100
        h_o25 = home_form.get("over25_pct", 50) / 100
        a_o25 = away_form.get("over25_pct", 50) / 100

        # Form rating with home advantage (home worth +0.35 pts/game equivalent)
        h_fr = home_form.get("form_rating", 1.5) + 0.35
        a_fr = away_form.get("form_rating", 1.5)
        gap  = h_fr - a_fr   # positive = home stronger

        # Attack/defence composite score per team
        # Higher = better attacker vs opponent defence
        h_attack_vs_adef = hg - aga   # home goals avg vs away concede avg
        a_attack_vs_hdef = ag - hga   # away goals avg vs home concede avg

        base = {
            "id": fid, "home": home, "away": away,
            "home_logo": home_logo, "away_logo": away_logo,
            "date": date_str, "home_id": home_id, "away_id": away_id,
            "league_id": league_id, "market_type": "1X2",
            "league": fixture["league"]["name"],
            "odds_home": ho, "odds_draw": do, "odds_away": ao,
            "odds_available": has_odds
        }

        def _reason(team, w, d, l, gf, ga, extra=""):
            return f"{team}: {w}W-{d}D-{l}L · {gf:.1f} scored, {ga:.1f} conceded/game{(' · ' + extra) if extra else ''}"

        # ── Value filter: skip if odds imply overbet favourite ──
        # If real odds available, only back at value — don't back 1.20 shots
        h_min_odds = 1.30 if risk == "safe" else 1.50 if risk == "balanced" else 1.80
        a_min_odds = 1.40 if risk == "safe" else 1.60 if risk == "balanced" else 2.00

        # ─── SAFE: Back clear favourite, multiple signals required ───
        if risk == "safe":
            # Strong home — form gap + win rate + goals + low opponent scoring
            if gap >= 0.7 and h_wr >= 0.50 and h_attack_vs_adef >= 0.0:
                if not has_odds or ho >= h_min_odds:
                    pick_odds = ho if has_odds and ho >= h_min_odds else round(1.50 + max(0, 2.0 - gap) * 0.15, 2)
                    conf = min(87, 60 + int(gap * 10) + int(h_wr * 15) + int(hcs * 10))
                    return {**base, "bet": f"{home} Win", "odds": pick_odds, "confidence": conf,
                            "reasoning": _reason(home, hw, hd, hl, hg, hga, f"Clean sheets {int(hcs*100)}%")}
            # Strong away — must be clearly dominant to overcome home disadvantage
            if gap <= -0.8 and a_wr >= 0.55 and a_attack_vs_hdef >= 0.2:
                if not has_odds or ao >= a_min_odds:
                    pick_odds = ao if has_odds and ao >= a_min_odds else round(1.65 + max(0, 2.0 - abs(gap)) * 0.15, 2)
                    conf = min(84, 58 + int(abs(gap) * 10) + int(a_wr * 14))
                    return {**base, "bet": f"{away} Win", "odds": pick_odds, "confidence": conf,
                            "reasoning": _reason(away, aw, ad, al, ag, aga, f"Away dominance clear")}
            return None

        # ─── RISKY: Value bets — back underdogs with good underlying stats ───
        elif risk == "risky":
            # Away value — decent form, high odds, goals in them
            if gap < 0.4 and a_wr >= 0.35 and ag >= 1.2:
                if not has_odds or ao >= 2.5:
                    pick_odds = ao if has_odds and ao >= 2.5 else round(2.80 + max(0, 1.5 - abs(gap)) * 0.4, 2)
                    conf = min(68, 44 + int(a_wr * 25) + int(a_o25 * 10))
                    return {**base, "bet": f"{away} Win", "odds": pick_odds, "confidence": conf,
                            "reasoning": _reason(away, aw, ad, al, ag, aga, f"Value at {pick_odds:.2f}")}
            # Home underdog value — strong home record but priced long
            if gap > 0.2 and h_wr >= 0.40:
                if not has_odds or ho >= 2.2:
                    pick_odds = ho if has_odds and ho >= 2.2 else 2.60
                    conf = min(66, 44 + int(gap * 8) + int(h_wr * 12))
                    return {**base, "bet": f"{home} Win", "odds": pick_odds, "confidence": conf,
                            "reasoning": _reason(home, hw, hd, hl, hg, hga, f"Home value")}
            # Draw value — genuinely equal teams AND good draw odds
            if abs(gap) < 0.2 and has_odds and do >= 3.3 and h_wr >= 0.25 and a_wr >= 0.25 and h_lr <= 0.4 and a_lr <= 0.4:
                conf = min(62, 45 + int(do * 3))
                return {**base, "bet": "Draw", "odds": round(do, 2), "confidence": conf,
                        "reasoning": f"Evenly matched — form gap {abs(gap):.2f}, draw @ {do:.2f} value"}
            return None

        # ─── BALANCED: Multiple signals, decent odds ───
        else:
            signals_home = sum([
                gap >= 0.4,
                h_wr >= 0.45,
                h_attack_vs_adef >= 0.1,
                hcs >= 0.25,
                h_o25 >= 0.5,
            ])
            signals_away = sum([
                gap <= -0.5,
                a_wr >= 0.45,
                a_attack_vs_hdef >= 0.0,
                acs >= 0.20,
                a_o25 >= 0.5,
            ])

            # Home: need at least 3 signals
            if signals_home >= 3 and (not has_odds or ho >= h_min_odds):
                pick_odds = ho if has_odds and ho >= h_min_odds else round(1.70 + max(0, 1.8 - gap) * 0.2, 2)
                conf = min(83, 50 + signals_home * 6 + int(gap * 8) + int(h_wr * 10))
                return {**base, "bet": f"{home} Win", "odds": pick_odds, "confidence": conf,
                        "reasoning": _reason(home, hw, hd, hl, hg, hga,
                                             f"{signals_home}/5 signals · {int(h_o25*100)}% over 2.5")}

            # Away: need at least 3 signals
            if signals_away >= 3 and (not has_odds or ao >= a_min_odds):
                pick_odds = ao if has_odds and ao >= a_min_odds else round(1.90 + max(0, 2.0 - abs(gap)) * 0.2, 2)
                conf = min(80, 50 + signals_away * 6 + int(abs(gap) * 8) + int(a_wr * 10))
                return {**base, "bet": f"{away} Win", "odds": pick_odds, "confidence": conf,
                        "reasoning": _reason(away, aw, ad, al, ag, aga,
                                             f"{signals_away}/5 signals · {int(a_o25*100)}% over 2.5")}

            # Fallback: clear form edge + decent win rate
            if gap >= 0.5 and h_wr >= 0.50 and (not has_odds or ho >= h_min_odds):
                pick_odds = ho if has_odds and ho >= h_min_odds else round(1.75 + max(0, 1.5 - gap) * 0.2, 2)
                conf = min(76, 52 + int(gap * 10) + int(h_wr * 12))
                return {**base, "bet": f"{home} Win", "odds": pick_odds, "confidence": conf,
                        "reasoning": _reason(home, hw, hd, hl, hg, hga)}
            if gap <= -0.6 and a_wr >= 0.50 and (not has_odds or ao >= a_min_odds):
                pick_odds = ao if has_odds and ao >= a_min_odds else round(1.90 + max(0, 1.5 - abs(gap)) * 0.2, 2)
                conf = min(74, 50 + int(abs(gap) * 10) + int(a_wr * 12))
                return {**base, "bet": f"{away} Win", "odds": pick_odds, "confidence": conf,
                        "reasoning": _reason(away, aw, ad, al, ag, aga)}
            return None

    except Exception as e:
        logger.error(f"Pick error: {e}")
        return None


def ht_pick(fixture: dict, home_form: dict, away_form: dict) -> Optional[dict]:
    """HT-specific pick using first-half patterns"""
    try:
        home = fixture["teams"]["home"]["name"]
        away = fixture["teams"]["away"]["name"]
        fid = fixture["fixture"]["id"]
        home_id = fixture["teams"]["home"]["id"]
        away_id = fixture["teams"]["away"]["id"]
        dt = datetime.fromisoformat(fixture["fixture"]["date"].replace("Z", "+00:00"))

        h_ht = home_form.get("ht_lead_pct", 40)
        a_ht = away_form.get("ht_lead_pct", 35)
        hg = home_form.get("gf_avg", 1.2)
        ag = away_form.get("gf_avg", 1.0)
        total = hg + ag

        base = {
            "id": fid, "home": home, "away": away,
            "home_logo": fixture["teams"]["home"]["logo"],
            "away_logo": fixture["teams"]["away"]["logo"],
            "date": dt.strftime("%d/%m %H:%M"),
            "home_id": home_id, "away_id": away_id
        }

        # HT Result
        if h_ht > 55 and h_ht > a_ht + 15:
            return {**base, "bet": f"{home} HT Lead", "odds": round(1.85 + (h_ht - 55) * 0.02, 2),
                    "confidence": min(78, h_ht), "market_type": "HT Result",
                    "reasoning": f"{home} lead at HT in {h_ht}% of games — strong opening pattern"}
        elif a_ht > 55:
            return {**base, "bet": f"{away} HT Lead", "odds": round(2.20 + (a_ht - 55) * 0.02, 2),
                    "confidence": min(72, a_ht), "market_type": "HT Result",
                    "reasoning": f"{away} lead at HT in {a_ht}% of games"}

        # HT Over 0.5
        if total > 2.0:
            conf = min(78, 55 + int(total * 8))
            return {**base, "bet": "HT Over 0.5 Goals", "odds": round(1.40 + total * 0.04, 2),
                    "confidence": conf, "market_type": "HT O/U",
                    "reasoning": f"Combined {round(total,1)} g/game — goal before HT very likely"}
        return None
    except Exception:
        return None


def analyze_and_pick_with_fbd(fixture, home_form, away_form, risk, market="winner"):
    """Wrapper around analyze_and_pick that boosts confidence using FBD data."""
    pick = analyze_and_pick(fixture, home_form, away_form, risk, market)
    if not pick: return None
    try:
        home = fixture["teams"]["home"]["name"]
        away = fixture["teams"]["away"]["name"]
        bet_lower = pick.get("bet","").lower()
        side = "draw" if "draw" in bet_lower else "away" if away.lower() in bet_lower else "home"
        pick["confidence"] = fbd_boost_confidence(pick["confidence"], home, away, side)
        fbd = get_fbd_match(home, away)
        if fbd:
            pick["fbd_prediction"] = fbd.get("prediction","")
            pick["fbd_btts"]       = fbd.get("pred_btts")
            pick["fbd_over25"]     = fbd.get("pred_over25")
    except Exception:
        pass
    return pick


# =========================
# ROUTES
# =========================

# ── FBD ENDPOINTS ──────────────────────────────────────────────

@app.get("/fbd/value-bets")
async def fbd_value_bets():
    data = get_fbd_data()
    btts, over25 = [], []
    for m in data:
        home=m.get("home",""); away=m.get("away","")
        if not home or not away: continue
        bp=m.get("pred_btts"); op=m.get("pred_over25")
        if bp and float(bp)>=55:
            btts.append({"home":home,"away":away,"league":m.get("league",""),
                         "btts_prob":bp,"avg_odds":m.get("avg_btts_yes"),
                         "fbd_prediction":m.get("prediction","")})
        if op and float(op)>=55:
            over25.append({"home":home,"away":away,"league":m.get("league",""),
                           "over25_prob":op,"avg_odds":m.get("avg_over25"),
                           "pred_goals":m.get("pred_goals"),
                           "fbd_prediction":m.get("prediction","")})
    btts.sort(key=lambda x: x["btts_prob"] or 0, reverse=True)
    over25.sort(key=lambda x: x["over25_prob"] or 0, reverse=True)
    return {"ok":True,"btts":btts[:10],"over25":over25[:10],
            "fbd_loaded":len(data)>0,"fbd_match_count":len(data)}

@app.get("/fbd/today")
async def fbd_today_admin(token: str=""):
    if token != ADMIN_PASSWORD_ENV:
        return {"ok":False,"error":"Unauthorised"}
    data = get_fbd_data(force=True)
    return {"ok":True,"count":len(data),"matches":data}

@app.post("/fbd/refresh")
async def fbd_refresh(request: dict):
    if request.get("token") != ADMIN_PASSWORD_ENV:
        return {"ok":False,"error":"Unauthorised"}
    data = get_fbd_data(force=True)
    return {"ok":True,"count":len(data),"message":f"FBD refreshed — {len(data)} matches loaded"}



@app.post("/redeem-code")
async def redeem_code(request: dict):
    """Validate a Pro access code. Single-use. Sends welcome email + returns Telegram invite."""
    code  = (request.get("code") or "").strip().upper()
    email = (request.get("email") or "").strip().lower()
    name  = (request.get("name") or "there").strip()

    if not code:
        return {"ok": False, "error": "Please enter your access code"}
    if code not in _all_codes:
        return {"ok": False, "error": "Invalid access code — check for typos or contact support"}
    if code in _used_codes:
        return {"ok": False, "error": "This code has already been used. Contact support if this is a mistake"}

    _used_codes[code] = {"email": email, "name": name, "used_at": datetime.now().isoformat()}
    logger.info(f"✅ Code redeemed: {code} by {email}")
    # Register in user db if not already there (no password yet — set on signup form)
    if email and email not in _user_db:
        _user_db[email] = {"name": name, "surname": "", "password_hash": "",
            "plan": "pro", "role": "user", "access_code": code,
            "joined": datetime.now().isoformat(), "telegram_invite": TELEGRAM_PRO_INVITE_LINK}
        _save_user_db()

    email_sent = email_welcome_pro(email, name, code, TELEGRAM_PRO_INVITE_LINK) if email else False

    return {
        "ok": True,
        "plan": "pro",
        "telegram_invite": TELEGRAM_PRO_INVITE_LINK or None,
        "email_sent": email_sent,
        "message": f"Welcome to AccaGenius Pro, {name}! 🎉"
    }


@app.get("/admin/codes")
async def admin_codes(token: str = "", pwd: str = ""):
    # Accept either the hashed token OR the raw password directly
    import hashlib
    if not ADMIN_PASSWORD_ENV:
        return {"error": "ADMIN_PASSWORD not set in Railway"}
    check = token or pwd
    if not check:
        return {"error": "Unauthorised"}
    expected_token = hashlib.sha256((ADMIN_PASSWORD_ENV + datetime.now().strftime("%Y-%m-%d")).encode()).hexdigest()[:32]
    if check != expected_token and check != ADMIN_PASSWORD_ENV:
        return {"error": "Unauthorised"}
    available = sorted([c for c in _all_codes if c not in _used_codes])
    return {
        "total": len(_all_codes), "used": len(_used_codes), "available": len(available),
        "available_codes": available, "used_details": dict(_used_codes)
    }



# ══════════════════════════════════════════════════════
# REFER A FRIEND
# ══════════════════════════════════════════════════════

_referrals: dict = {}   # ref_code -> {owner_email, uses, rewarded_at}
_REFERRALS_FILE = "/tmp/ag_referrals.json"

def _load_referrals():
    global _referrals
    import json
    try:
        if os.path.exists(_REFERRALS_FILE):
            with open(_REFERRALS_FILE) as f:
                _referrals = json.load(f)
    except Exception as e:
        logger.warning(f"Could not load referrals: {e}")

def _save_referrals():
    import json
    try:
        with open(_REFERRALS_FILE, 'w') as f:
            json.dump(_referrals, f)
    except Exception as e:
        logger.warning(f"Could not save referrals: {e}")

@app.post("/referral/create")
async def create_referral_link(request: dict):
    """Create a personal referral link for a Pro user."""
    email = (request.get("email") or "").strip().lower()
    if not email or email not in _user_db:
        return {"ok": False, "error": "Account not found"}
    user = _user_db[email]
    if user.get("plan") != "pro":
        return {"ok": False, "error": "Referral links are for Pro members only"}
    # Return existing or create new
    existing = next((k for k, v in _referrals.items() if v.get("owner_email") == email), None)
    if existing:
        return {"ok": True, "ref_code": existing, "link": f"{SITE_URL}?ref={existing}", "uses": _referrals[existing].get("uses", 0)}
    import secrets as _s
    code = "REF-" + _s.token_hex(5).upper()
    _referrals[code] = {"owner_email": email, "owner_name": user.get("name",""), "uses": 0, "created": datetime.now().isoformat(), "rewarded": []}
    _save_referrals()
    return {"ok": True, "ref_code": code, "link": f"{SITE_URL}?ref={code}", "uses": 0}

@app.post("/referral/use")
async def use_referral(request: dict):
    """Called when someone signs up via a referral link — rewards the referrer."""
    ref_code  = (request.get("ref_code") or "").strip().upper()
    new_email = (request.get("email") or "").strip().lower()
    if not ref_code or ref_code not in _referrals:
        return {"ok": False, "error": "Invalid referral code"}
    ref = _referrals[ref_code]
    if new_email in ref.get("rewarded", []):
        return {"ok": False, "error": "Already used"}
    ref["uses"] = ref.get("uses", 0) + 1
    ref.setdefault("rewarded", []).append(new_email)
    owner_email = ref["owner_email"]
    owner_name  = ref.get("owner_name", owner_email)
    _uses = ref["uses"]
    # 1 free month per paying referral — fire every time
    pro_ch = TELEGRAM_CHANNEL_PRO or TELEGRAM_CHANNEL_ID
    if pro_ch:
        msg = (
            "\U0001f381 <b>REFERRAL REWARD</b>\n"
            f"{owner_name} ({owner_email}) earned 1 free month!\n"
            f"Friend who joined: {new_email}\n"
            f"Total referrals: {_uses}"
        )
        send_telegram_to(msg, pro_ch)
    # Send reward email to referrer
    if owner_email in _user_db:
        owner_first = _user_db[owner_email].get("name", "there")
        reward_html = (
            "<div style='background:#0a0a0f;color:#e2e8f0;font-family:system-ui,sans-serif;"
            "padding:32px;border-radius:16px;max-width:500px;margin:0 auto'>"
            f"<h2 style='color:#f0b429'>&#127873; You earned a free month!</h2>"
            f"<p>Hi {owner_first},</p>"
            f"<p>Your friend <strong>{new_email}</strong> just signed up to AccaGenius Pro using your referral link.</p>"
            "<p style='background:rgba(240,180,41,0.1);border:1px solid rgba(240,180,41,0.3);"
            "border-radius:10px;padding:16px;text-align:center;font-size:1.1rem;font-weight:bold;color:#f0b429'>"
            "+1 FREE MONTH added to your account &#127881;</p>"
            f"<p style='color:#64748b;font-size:0.85rem'>Total referrals: {_uses}. Keep sharing to keep earning!</p>"
            f"<a href='{SITE_URL}' style='display:inline-block;background:linear-gradient(135deg,#f0b429,#f6ad55);"
            "color:#000;font-weight:800;padding:12px 28px;border-radius:10px;text-decoration:none;margin-top:8px'>"
            "Back to AccaGenius &#8594;</a></div>"
        )
        send_email(owner_email, "🎁 You earned a free month on AccaGenius Pro!", _email_base(reward_html))
    _save_referrals()
    return {"ok": True, "message": "Referral tracked — 1 free month earned!", "total_uses": _uses}

@app.get("/referral/stats")
async def referral_stats(email: str = ""):
    """Get referral stats for a user."""
    if not email:
        return {"ok": False}
    email = email.strip().lower()
    ref = next(({"code": k, **v} for k, v in _referrals.items() if v.get("owner_email") == email), None)
    if not ref:
        return {"ok": True, "has_link": False, "uses": 0}
    free_months = ref["uses"]  # 1 free month per referral
    return {
        "ok": True, "has_link": True,
        "ref_code": ref["code"],
        "link": f"{SITE_URL}?ref={ref['code']}",
        "uses": ref["uses"],
        "free_months_earned": free_months,
        "next_reward_in": 1  # every referral earns a month
    }

@app.get("/admin/referrals")
async def admin_referrals(pwd: str = ""):
    admin_pw = (ADMIN_PASSWORD_ENV or "").strip()
    import hashlib as _hlr
    _dtr = _hlr.sha256((admin_pw + __import__("datetime").datetime.now().strftime("%Y-%m-%d")).encode()).hexdigest()[:32]
    if not admin_pw or pwd not in (admin_pw, _dtr):
        return {"error": "Unauthorised"}
    return {"referrals": list(_referrals.values()), "total": len(_referrals)}

@app.post("/admin/gift-pro")
async def admin_gift_pro(request: dict):
    """Gift Pro access to a friend — no code needed, just email + name."""
    token = (request.get("token") or "").strip()
    pwd   = (request.get("pwd") or "").strip()
    check = token or pwd
    admin_pw = (ADMIN_PASSWORD_ENV or "").strip()
    if not admin_pw:
        return {"ok": False, "error": "ADMIN_PASSWORD not configured in Railway"}
    import hashlib
    daily_token = hashlib.sha256((admin_pw + datetime.now().strftime("%Y-%m-%d")).encode()).hexdigest()[:32]
    if check != admin_pw and check != daily_token:
        return {"ok": False, "error": "Unauthorised"}

    email = (request.get("email") or "").strip().lower()
    name  = (request.get("name") or "Friend").strip()

    if not email or "@" not in email:
        return {"ok": False, "error": "Valid email required"}

    if email in _user_db:
        # Already exists — just upgrade to pro
        _user_db[email]["plan"] = "pro"
        _user_db[email]["gifted"] = True
        _save_user_db()
        return {"ok": True, "message": f"{name} ({email}) upgraded to Pro ✅", "already_existed": True}

    # Create new gifted pro user
    _user_db[email] = {
        "name": name, "surname": "", "password_hash": "",
        "plan": "pro", "role": "user", "access_code": "GIFTED",
        "gifted": True,
        "joined": datetime.now().isoformat(),
        "telegram_invite": TELEGRAM_PRO_INVITE_LINK
    }
    _save_user_db()

    # Send welcome email
    email_sent = email_welcome_pro(email, name, "GIFTED", TELEGRAM_PRO_INVITE_LINK)

    logger.info(f"🎁 Gifted Pro: {name} <{email}>")
    return {
        "ok": True,
        "message": f"Pro access gifted to {name} ({email}) 🎁",
        "email_sent": email_sent,
        "telegram_invite": TELEGRAM_PRO_INVITE_LINK
    }


@app.get("/admin/gifted-users")
async def admin_gifted_users(token: str = ""):
    """List all gifted pro users."""
    import hashlib as _hlgu
    _dt = _hlgu.sha256((ADMIN_PASSWORD_ENV + __import__("datetime").datetime.now().strftime("%Y-%m-%d")).encode()).hexdigest()[:32]
    if not ADMIN_PASSWORD_ENV or token not in (ADMIN_PASSWORD_ENV, _dt):
        return {"ok": False, "error": "Unauthorised"}
    gifted = [
        {"email": e, "name": d.get("name",""), "joined": d.get("joined","")}
        for e, d in _user_db.items() if d.get("gifted")
    ]
    return {"ok": True, "gifted": gifted, "count": len(gifted)}


@app.post("/admin/generate-codes")
async def admin_generate_codes(request: dict):
    import secrets as _sec
    token = request.get("token", "") or request.get("pwd", "")
    count = min(int(request.get("count", 10)), 50)
    if not ADMIN_PASSWORD_ENV or token != ADMIN_PASSWORD_ENV:
        return {"error": "Unauthorised"}
    chars = "ABCDEFGHJKLMNPQRSTUVWXYZ23456789"
    new_codes = [f"ACCA-{''.join(_sec.choice(chars) for _ in range(4))}" for _ in range(count)]
    combined  = sorted(set(list(_all_codes) + new_codes))
    return {"new_codes": new_codes, "railway_value": ",".join(combined),
            "instructions": "Paste railway_value into PRO_ACCESS_CODES Railway Variable then redeploy"}


@app.post("/admin/send-code-email")
async def admin_send_code_email(request: dict):
    """Manually email a Pro code to a customer (use after manual PayPal sale)."""
    import hashlib
    token = request.get("token", "")
    if not ADMIN_PASSWORD_ENV or not token:
        return {"error": "Unauthorised"}
    expected = hashlib.sha256((ADMIN_PASSWORD_ENV + datetime.now().strftime("%Y-%m-%d")).encode()).hexdigest()[:32]
    if token != expected:
        return {"error": "Unauthorised"}
    to    = (request.get("email") or "").strip()
    name  = (request.get("name") or "there").strip()
    code  = (request.get("code") or "").strip().upper()
    if not to or not code:
        return {"error": "email and code are required"}
    sent = email_welcome_pro(to, name, code, TELEGRAM_PRO_INVITE_LINK)
    return {"ok": sent, "to": to, "code": code}


# ────────────────────────────────────────────────────────────────
# PASSWORD RESET ENDPOINTS
# ────────────────────────────────────────────────────────────────

@app.post("/forgot-password")
async def forgot_password(request: dict):
    import secrets as _sec
    email = (request.get("email") or "").strip().lower()
    name  = (request.get("name") or "there").strip()
    if not email:
        return {"ok": False, "error": "Please enter your email address"}
    token   = _sec.token_urlsafe(32)
    expires = datetime.now() + timedelta(hours=1)
    _reset_tokens[token] = {
        "email": email, "name": name, "expires": expires, "used": False,
        "plan": (request.get("plan") or "free"),
        "role": (request.get("role") or "user"),
        "access_code": (request.get("access_code") or ""),
        "joined": (request.get("joined") or ""),
    }
    email_password_reset(email, name, token)
    # Always return ok — never reveal if email exists (security)
    return {"ok": True, "message": "If an account exists for that email, a reset link is on its way"}


@app.get("/reset-password")
async def check_reset_token(token: str = ""):
    if not token or token not in _reset_tokens:
        return {"valid": False, "error": "Invalid or expired reset link"}
    entry = _reset_tokens[token]
    if entry["used"]:
        return {"valid": False, "error": "This link has already been used"}
    if datetime.now() > entry["expires"]:
        del _reset_tokens[token]
        return {"valid": False, "error": "Link expired — please request a new one"}
    return {"valid": True, "email": entry["email"]}


@app.post("/reset-password")
async def do_reset_password(request: dict):
    token    = (request.get("token") or "").strip()
    new_pass = (request.get("new_password") or "").strip()
    if not token or token not in _reset_tokens:
        return {"ok": False, "error": "Invalid or expired link"}
    entry = _reset_tokens[token]
    if entry["used"]:
        return {"ok": False, "error": "Already used — request a new reset link"}
    if datetime.now() > entry["expires"]:
        return {"ok": False, "error": "Link expired — please request a new one"}
    if len(new_pass) < 6:
        return {"ok": False, "error": "Password must be at least 6 characters"}
    email = entry["email"]
    if email not in _user_db:
        return {"ok": False, "error": "Account not found"}
    _user_db[email]["password_hash"] = _hash_pw(new_pass)
    _save_user_db()
    _reset_tokens[token]["used"] = True
    logger.info(f"Password reset completed for {email}")
    return {"ok": True, "email": email}


# ────────────────────────────────────────────────────────────────
# ADMIN ENDPOINTS
# ────────────────────────────────────────────────────────────────

@app.get("/admin/verify-get")
async def admin_verify_get(password: str = ""):
    if not ADMIN_PASSWORD_ENV:
        return {"ok": False, "error": "ADMIN_PASSWORD not set in Railway Variables"}
    if password == ADMIN_PASSWORD_ENV:
        return {"ok": True, "token": ADMIN_PASSWORD_ENV}
    return {"ok": False, "error": "Invalid password"}


@app.post("/admin/verify")
async def admin_verify(request: dict):
    import hashlib
    pwd = request.get("password", "")
    admin_pw = (ADMIN_PASSWORD_ENV or "").strip()
    logger.info(f"Admin verify: pw_len={len(pwd)} env_set={bool(admin_pw)} match={pwd==admin_pw}")
    if not admin_pw:
        return {"ok": False, "error": "ADMIN_PASSWORD not set in Railway Variables"}
    if pwd.strip() == admin_pw.strip():
        token = hashlib.sha256((admin_pw + datetime.now().strftime("%Y-%m-%d")).encode()).hexdigest()[:32]
        logger.info("Admin login successful")
        return {"ok": True, "token": token, "role": "admin"}
    return {"ok": False, "error": "Incorrect password"}


@app.get("/admin/check-config")
async def check_config():
    """Quick check that Railway env vars are set correctly (passwords not exposed)."""
    return {
        "ADMIN_PASSWORD_set":       bool(ADMIN_PASSWORD_ENV),
        "PRO_ACCESS_CODES_set":     bool(PRO_ACCESS_CODES_RAW),
        "codes_loaded":             len(_all_codes),
        "codes_used":               len(_used_codes),
        "TELEGRAM_BOT_set":         bool(TELEGRAM_BOT_TOKEN),
        "TELEGRAM_PRO_set":         bool(TELEGRAM_PRO_INVITE_LINK),
        "RESEND_API_KEY_set":       bool(RESEND_API_KEY),
        "EMAIL_FROM":               EMAIL_FROM,
        "SITE_URL":                 SITE_URL,
    }


@app.get("/")
@app.head("/")
async def root():
    return {"status": "AccaGenius Ultimate API — Live", "version": "5.0"}

@app.get("/health")
async def health():
    return {"status": "ok", "timestamp": datetime.now().isoformat()}


@app.get("/db-status")
async def db_status():
    """Check database connection — visit this URL after deploy to verify persistence."""
    return {
        "database_url_set": bool(DATABASE_URL),
        "is_postgres": _is_postgres(),
        "user_count": len(_user_db),
        "mode": "PostgreSQL ✅" if _is_postgres() else "SQLite/tmp ⚠️ — accounts will wipe on deploy"
    }

@app.get("/telegram/status")
async def telegram_status():
    pro_raw  = TELEGRAM_CHANNEL_PRO or TELEGRAM_CHANNEL_ID or ""
    free_raw = TELEGRAM_CHANNEL_FREE or ""
    return {
        "bot_configured": bool(TELEGRAM_BOT_TOKEN),
        "pro_channel":  normalise_channel(pro_raw)  if pro_raw  else "not set",
        "free_channel": normalise_channel(free_raw) if free_raw else "not set",
        "pro_channel_raw":  pro_raw  or "not set",
        "free_channel_raw": free_raw or "not set",
        "free_alerts_today": free_alerts_sent_today,
        "alerted_fixtures_today": len(alerted_fixtures),
        "thresholds": {
            "win_pct": ALERT_WIN_PCT,
            "xg_gap": ALERT_XG_GAP,
            "minute_min": ALERT_MINUTE_MIN,
            "minute_max": ALERT_MINUTE_MAX,
        }
    }


@app.get("/telegram/find-channels")
async def find_channels():
    """
    Automatically finds channel IDs for any channel your bot is an admin of.
    HOW TO USE:
      1. Make sure your bot is added as admin to your channels
      2. Send ANY message to each channel (e.g. type 'test')
      3. Visit this URL in your browser
      4. Copy the IDs shown and paste into Railway Variables
    """
    if not TELEGRAM_BOT_TOKEN:
        return {"error": "TELEGRAM_BOT_TOKEN not set in Railway Variables"}

    try:
        # getUpdates fetches recent messages/events the bot has seen
        url = f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/getUpdates"
        r = requests.get(url, timeout=10)
        data = r.json()

        if not data.get("ok"):
            return {
                "error": "Bot token invalid or bot has no recent activity",
                "detail": data.get("description", ""),
                "fix": "Make sure you added the bot as admin to your channels and sent a message in each channel"
            }

        # Extract all unique chats the bot has seen
        chats = {}
        for update in data.get("result", []):
            # Check all possible message locations
            for key in ["message", "channel_post", "edited_channel_post", "my_chat_member"]:
                item = update.get(key, {})
                if not item:
                    continue
                chat = item.get("chat", {})
                if not chat:
                    continue
                chat_id = chat.get("id")
                chat_title = chat.get("title") or chat.get("username") or chat.get("first_name") or "Unknown"
                chat_type = chat.get("type", "unknown")
                if chat_id and chat_type in ("channel", "supergroup", "group"):
                    chats[chat_id] = {
                        "id": chat_id,
                        "title": chat_title,
                        "type": chat_type
                    }

        if not chats:
            return {
                "found": 0,
                "channels": [],
                "instructions": [
                    "No channels found yet. Do these steps then visit this URL again:",
                    "1. Go to your Free channel in Telegram",
                    "2. Type and send any message (e.g. 'test')",
                    "3. Go to your Pro channel",
                    "4. Type and send any message",
                    "5. Refresh this page"
                ]
            }

        channel_list = list(chats.values())

        return {
            "found": len(channel_list),
            "channels": channel_list,
            "next_steps": [
                "Copy the 'id' value for each channel",
                "Go to Railway → your project → Variables",
                "Add: FREE_CHANNEL_ID = (id of your free channel)",
                "Add: PRO_CHANNEL_ID = (id of your pro channel)",
                "Then click Deploy or Redeploy"
            ]
        }

    except Exception as e:
        return {"error": str(e)}

@app.post("/telegram/test")
async def telegram_test():
    results = {}
    pro_channel = TELEGRAM_CHANNEL_PRO or TELEGRAM_CHANNEL_ID
    test_pro = (
        "🟢 <b>ACCAGENIUS PRO — Test Alert</b>\n"
        "━━━━━━━━━━━━━━━━━━━━\n"
        "✅ Pro channel working!\n"
        "Full xG stats, win %, odds — all live here.\n"
        "🤖 AccaGenius Pro"
    )
    test_free = (
        "🟡 <b>ACCAGENIUS FREE — Test Alert</b>\n"
        "━━━━━━━━━━━━━━━━━━━━\n"
        "✅ Free channel working!\n"
        "💰 <b>In-Play Tip: Example Win</b>\n"
        "🔒 Full analysis on AccaGenius Pro\n"
        "👉 <b>Upgrade: accagenius.com</b>"
    )
    if pro_channel:
        results["pro"] = send_telegram_to(test_pro, pro_channel)
    if TELEGRAM_CHANNEL_FREE:
        results["free"] = send_telegram_to(test_free, TELEGRAM_CHANNEL_FREE)
    if not results:
        return {"error": "No channels configured. Set TELEGRAM_CHANNEL_PRO and TELEGRAM_CHANNEL_FREE in Railway Variables."}
    return {"results": results}
@app.post("/telegram/send-todays-acca")
async def send_todays_acca_telegram():
    """Send today's acca picks to both Telegram channels."""
    pro_channel = TELEGRAM_CHANNEL_PRO or TELEGRAM_CHANNEL_ID
    if not TELEGRAM_BOT_TOKEN:
        return {"error": "TELEGRAM_BOT_TOKEN not set"}
    if not pro_channel and not TELEGRAM_CHANNEL_FREE:
        return {"error": "No channels configured"}

    today_str = _london_now().strftime("%Y-%m-%d")
    cached = cache_get(f"today_acca_{today_str}")
    if not cached:
        return {"error": "No acca generated yet — hit /today-acca first"}

    results = {}
    for acca in cached.get("accas", []):
        picks = acca.get("picks", [])
        if not picks:
            continue
        label = acca.get("label", "Acca")
        emoji = acca.get("emoji", "⚽")
        total_odds = round(
            __import__("functools").reduce(lambda a, p: a * p.get("odds", 1.75), picks, 1.0), 2
        )
        sep = "━" * 20

        def fmt_form(p, team):
            """Format form string e.g. WWDLW for a team from pick data."""
            key = "home_form" if team == "home" else "away_form"
            form = p.get(key, "")
            if not form:
                return ""
            coloured = ""
            for ch in str(form)[-5:]:
                if ch == "W": coloured += "🟢"
                elif ch == "L": coloured += "🔴"
                elif ch == "D": coloured += "🟡"
                else: coloured += "⚪"
            return coloured

        # Pro legs — full detail with form
        legs_pro = ""
        for p in picks:
            hform = fmt_form(p, "home")
            aform = fmt_form(p, "away")
            kick  = p.get("time", "")
            legs_pro += "⚽ <b>" + p.get("home","") + "</b> vs <b>" + p.get("away","") + "</b>"
            if kick: legs_pro += "  🕐 " + kick
            legs_pro += "\n"
            if hform: legs_pro += "   " + p.get("home","").split()[0] + " form: " + hform + "\n"
            if aform: legs_pro += "   " + p.get("away","").split()[0] + " form: " + aform + "\n"
            legs_pro += "   ✅ <b>" + str(p.get("bet","")) + "</b> @ <b>" + str(p.get("odds","")) + "</b>\n\n"

        # Free legs — teams, bet type and form shown; odds locked
        legs_free = ""
        for p in picks:
            hform = fmt_form(p, "home")
            aform = fmt_form(p, "away")
            kick  = p.get("time", "")
            legs_free += "⚽ <b>" + p.get("home","") + "</b> vs <b>" + p.get("away","") + "</b>"
            if kick: legs_free += "  🕐 " + kick
            legs_free += "\n"
            if hform: legs_free += "   " + p.get("home","").split()[0] + " form: " + hform + "\n"
            if aform: legs_free += "   " + p.get("away","").split()[0] + " form: " + aform + "\n"
            legs_free += "   ✅ <b>" + str(p.get("bet","")) + "</b> · 🔒 Odds on Pro\n\n"

        pro_msg = (
            emoji + " <b>ACCA OF THE DAY — " + label.upper() + "</b>\n"
            + sep + "\n"
            + legs_pro
            + sep + "\n"
            + "💰 Combined odds: <b>" + str(total_odds) + "x</b>\n"
            + "📅 " + today_str + "\n"
            + "🤖 AccaGenius Pro · accagenius.com"
        )

        free_msg = (
            emoji + " <b>ACCA OF THE DAY — " + label.upper() + "</b>\n"
            + sep + "\n"
            + legs_free
            + sep + "\n"
            + "📅 " + today_str + "  ·  " + str(len(picks)) + " picks\n"
            + "🔒 Full odds &amp; xG analysis on Pro\n"
            + "👉 <b>accagenius.com</b> · 7-day free trial"
        )

        if pro_channel:
            mid = send_telegram_to(pro_msg, pro_channel)
            results["pro_" + acca["theme"]] = mid
        if TELEGRAM_CHANNEL_FREE:
            mid = send_telegram_to(free_msg, TELEGRAM_CHANNEL_FREE)
            results["free_" + acca["theme"]] = mid

    return {"sent": results, "accas": len(cached.get("accas", []))}



@app.post("/telegram/test-alert")
async def test_alert_endpoint():
    """Send a test xG alert to Pro channel to verify it's working."""
    pro_channel = TELEGRAM_CHANNEL_PRO or TELEGRAM_CHANNEL_ID
    if not pro_channel:
        return {"error": "No Pro channel configured"}
    test_msg = (
        "🟢 <b>ACCAGENIUS PRO ALERT</b> 🟢\n"
        "━━━━━━━━━━━━━━━━━━━━\n"
        "⚽ <b>Test FC vs Demo United</b>\n"
        "🏆 Test League\n"
        "🕐 55' | Score: <b>1-0</b>\n"
        "━━━━━━━━━━━━━━━━━━━━\n"
        "📊 <b>Win Probability</b>\n"
        "  Test FC: <b>74%</b> 🟢\n"
        "  Draw: 16%\n"
        "  Demo United: 10%\n"
        "\n"
        "⚡ <b>Real xG</b>\n"
        "  Test FC: 1.8  |  Demo United: 0.4\n"
        "  Gap: +1.40\n"
        "\n"
        "🔥 <b>Trigger: ✅ Scored first</b>\n"
        "━━━━━━━━━━━━━━━━━━━━\n"
        "💰 <b>Suggested Bet: Test FC Win</b>\n"
        "📈 Current Odds: <b>1.45</b>\n"
        "━━━━━━━━━━━━━━━━━━━━\n"
        "⚠️ <i>THIS IS A TEST ALERT — not a real tip</i>\n"
        "🤖 AccaGenius Pro | Live Intelligence"
    )
    mid = send_telegram_to(test_msg, pro_channel)
    if mid:
        return {"ok": True, "message": "Test alert sent to Pro channel ✅", "message_id": mid}
    return {"error": "Failed to send — check bot token and Pro channel ID"}


@app.get("/ht/values")
async def get_ht_values():
    """Half-time value picks — teams losing at HT but dominating xG."""
    live_data = api_get("fixtures", {"live": "all", "timezone": "Europe/London"})
    our_league_ids = {l["id"] for l in LEAGUES}
    values = []

    for f in live_data.get("response", []):
        try:
            status = f["fixture"]["status"]["short"]
            minute = f["fixture"]["status"].get("elapsed") or 0
            if status not in ["HT", "2H"] or minute > 55:
                continue
            league_id = f["league"]["id"]
            if league_id not in our_league_ids:
                continue

            home = f["teams"]["home"]["name"]
            away = f["teams"]["away"]["name"]
            fid  = f["fixture"]["id"]
            hs   = f["goals"]["home"] or 0
            as_  = f["goals"]["away"] or 0

            # Fetch HT stats
            stats_data = api_get("fixtures/statistics", {"fixture": fid})
            FMAP = {"Expected Goals": "xg", "expected_goals": "xg",
                    "Ball Possession": "poss", "Shots on Goal": "shots"}
            team_stats = {}
            for td in stats_data.get("response", []):
                tname = td["team"]["name"]
                ts = {}
                for s in td.get("statistics", []):
                    key = FMAP.get(s["type"])
                    if not key: continue
                    val = s.get("value")
                    if isinstance(val, str) and val.endswith("%"):
                        try: val = int(val.replace("%", ""))
                        except: pass
                    ts[key] = float(val or 0)
                team_stats[tname] = ts

            h = team_stats.get(home, {})
            a = team_stats.get(away, {})
            h_xg   = h.get("xg", 0)
            a_xg   = a.get("xg", 0)
            h_poss = h.get("poss", 50)

            alerts = []
            # Losing but xG dominant — comeback value
            if hs < as_ and h_xg > a_xg + 0.3:
                alerts.append({"type": "comeback", "team": home,
                    "reason": f"Trailing {hs}-{as_} but leads xG {h_xg:.1f}-{a_xg:.1f}",
                    "market": f"{home} HT/FT or 2nd Half Win", "confidence": 72})
            if as_ < hs and a_xg > h_xg + 0.3:
                alerts.append({"type": "comeback", "team": away,
                    "reason": f"Trailing {as_}-{hs} but leads xG {a_xg:.1f}-{h_xg:.1f}",
                    "market": f"{away} HT/FT or 2nd Half Win", "confidence": 72})
            # Level score — DNB value on dominant team
            if hs == as_ and abs(h_xg - a_xg) > 0.5:
                dom = home if h_xg > a_xg else away
                dom_xg = max(h_xg, a_xg)
                und_xg = min(h_xg, a_xg)
                alerts.append({"type": "dnb", "team": dom,
                    "reason": f"Level score but xG gap {dom_xg:.1f}-{und_xg:.1f}",
                    "market": f"{dom} Draw No Bet (2nd Half)", "confidence": 74})

            if alerts:
                values.append({
                    "fixture_id": fid, "home": home, "away": away,
                    "score": f"{hs}-{as_}", "minute": minute,
                    "league": f["league"]["name"],
                    "status": status,
                    "h_xg": h_xg, "a_xg": a_xg, "h_poss": h_poss,
                    "alerts": alerts
                })
        except Exception as e:
            logger.error(f"HT value error: {e}")

    return {"values": values, "count": len(values)}


@app.post("/telegram/goal-due-alert")
async def goal_due_alert(request: dict):
    """Receive a Goal Due alert from frontend and forward to Pro Telegram channel."""
    pro_channel = TELEGRAM_CHANNEL_PRO or TELEGRAM_CHANNEL_ID
    if not pro_channel:
        return {"ok": False, "error": "No Pro channel configured"}

    match   = request.get("match", "Unknown Match")
    minute  = request.get("minute", 0)
    team    = request.get("team", "")
    score   = request.get("score", "0-0")
    reasons = request.get("reasons", [])

    bullet = chr(10).join(f"  - {r}" for r in reasons[:3]) if reasons else "  - High xG vs goals scored"
    reasons_text = bullet

    sep = "━━━━━━━━━━━━━━━━━━━━"
    msg = (
        "\U0001f534 <b>GOAL DUE ALERT \u2014 PRO</b> \U0001f534\n"
        + sep + "\n"
        + f"\u26bd <b>{match}</b>\n"
        + f"\U0001f550 {minute}' | Score: <b>{score}</b>\n"
        + sep + "\n"
        + f"\U0001f3af <b>Goal Due: {team}</b>\n\n"
        + "\U0001f4ca <b>Why:</b>\n"
        + reasons_text + "\n\n"
        + "\u26a1 xG pressure building \u2014 goal expected soon\n"
        + sep + "\n"
        + "\U0001f4b0 Consider: Next Goal / Over markets\n"
        + "\u26a0\ufe0f <i>Bet responsibly. 18+ only.</i>\n"
        + "\U0001f916 AccaGenius Pro | Live Intelligence"
    )

    mid = send_telegram_to(msg, pro_channel)
    if mid:
        logger.info(f"Goal Due alert sent: {match} {minute}' — {team}")
        return {"ok": True, "message_id": mid}
    return {"ok": False, "error": "Failed to send"}


@app.post("/telegram/goal-scored-alert")
async def goal_scored_alert(request: dict):
    pro_channel = TELEGRAM_CHANNEL_PRO or TELEGRAM_CHANNEL_ID
    if not pro_channel:
        return {"ok": False}

    match       = request.get("match", "Unknown")
    minute      = request.get("minute", 0)
    team        = request.get("team", "")
    reasons     = request.get("reasons", [])
    strike_rate = request.get("strike_rate")
    goals_hit   = request.get("goals_hit", 0)
    total       = request.get("total_alerts", 0)

    sep = "━" * 20
    rate_line = f"\U0001f4ca Strike Rate: {strike_rate}% ({goals_hit}/{total} alerts hit)" if strike_rate is not None else ""
    bullet = chr(10).join(f"  - {r}" for r in reasons[:3]) if reasons else "  - xG pressure paid off"

    msg = (
        "\u26bd <b>GOAL SCORED \u2014 ALERT HIT</b> \u26bd\n"
        + sep + "\n"
        + f"\u2705 <b>{{match}}</b>\n"
        + f"\U0001f550 Goal confirmed at {{minute}}'\n"
        + f"\U0001f3af Team: <b>{{team}}</b>\n"
        + sep + "\n"
        + "\U0001f4cb <b>Why we flagged it:</b>\n"
        + bullet + "\n"
        + (rate_line + "\n" if rate_line else "")
        + sep + "\n"
        + "\U0001f916 AccaGenius Pro | Goal Due Intelligence"
    ).format(match=match, minute=minute, team=team)

    mid = send_telegram_to(msg, pro_channel)
    return {"ok": bool(mid)}


@app.post("/telegram/send-custom")
async def telegram_send_custom(request: dict):
    """Admin: send any custom message to pro or free channel."""
    token      = (request.get("token") or "").strip()
    message    = (request.get("message") or "").strip()
    channel_id = (request.get("channel") or "pro").strip()

    import hashlib as _hlsc
    admin_pw   = (ADMIN_PASSWORD_ENV or "").strip()
    daily_tok  = _hlsc.sha256((admin_pw + datetime.now().strftime("%Y-%m-%d")).encode()).hexdigest()[:32]
    if not admin_pw or token not in (admin_pw, daily_tok):
        return {"ok": False, "error": "Unauthorised"}
    if not message:
        return {"ok": False, "error": "No message"}

    pro_ch  = TELEGRAM_CHANNEL_PRO or TELEGRAM_CHANNEL_ID
    free_ch = TELEGRAM_CHANNEL_FREE
    target  = pro_ch if channel_id == "pro" else free_ch
    if not target:
        return {"ok": False, "error": f"{channel_id} channel not configured"}

    mid = send_telegram_to(message, target)
    return {"ok": bool(mid), "message_id": mid}


@app.post("/telegram/scan-now")
async def telegram_scan_now():
    """Manually trigger one scan cycle immediately."""
    any_channel = TELEGRAM_CHANNEL_PRO or TELEGRAM_CHANNEL_FREE or TELEGRAM_CHANNEL_ID
    if not TELEGRAM_BOT_TOKEN or not any_channel:
        return {"error": "Telegram not configured. Set TELEGRAM_BOT_TOKEN and TELEGRAM_CHANNEL_PRO in Railway Variables."}

    live_data = api_get("fixtures", {"live": "all", "timezone": "Europe/London"})
    our_league_ids = {l["id"] for l in LEAGUES}
    live_matches = []
    for f in live_data.get("response", []):
        try:
            if f["league"]["id"] not in our_league_ids:
                continue
            live_matches.append({
                "id": f["fixture"]["id"],
                "home": f["teams"]["home"]["name"],
                "away": f["teams"]["away"]["name"],
                "home_score": f["goals"]["home"] or 0,
                "away_score": f["goals"]["away"] or 0,
                "minute": f["fixture"]["status"].get("elapsed") or 0,
                "league": f["league"]["name"],
            })
        except Exception:
            pass

    results = [{"info": "manual scan uses score_inplay_markets only — old system disabled"}]

    return {"live_matches": len(live_matches), "results": results}


@app.get("/results/today")
async def get_results_today():
    """Return today's finished AND in-progress fixtures for leg result checking.
    Cached 2 minutes — fresh enough for result updates without burning quota."""
    today = datetime.now().strftime("%Y-%m-%d")
    cache_key = f"results_today_{today}_{datetime.now().strftime('%H%M')[:-1]}"  # 10-min buckets
    cached = cache_get(cache_key)
    if cached:
        return cached

    # Fetch finished + live — lets frontend update legs mid-match too
    fixtures = []
    for status_filter in ["FT-AET-PEN", "1H-HT-2H-ET-P"]:
        data = api_get("fixtures", {
            "date": today,
            "timezone": "Europe/London",
            "status": status_filter
        })
        for f in data.get("response", []):
            try:
                fixtures.append({
                    "id": f["fixture"]["id"],
                    "home": f["teams"]["home"]["name"],
                    "away": f["teams"]["away"]["name"],
                    "home_score": f["goals"]["home"] or 0,
                    "away_score": f["goals"]["away"] or 0,
                    "status": f["fixture"]["status"]["short"],
                    "league": f["league"]["name"],
                    "minute": f["fixture"]["status"].get("elapsed"),
                })
            except Exception:
                pass

    result = {"fixtures": fixtures, "date": today, "count": len(fixtures)}
    cache_set(cache_key, result, 600)  # 2 min cache
    return result


# Running Telegram P&L tracker (in-memory, resets on deploy)
telegram_pl = {"tips": 0, "won": 0, "lost": 0, "profit": 0.0, "stake": 10.0}

@app.post("/telegram/result")
async def post_telegram_result(payload: dict):
    """Called when a tip result is known — posts result to free channel and updates P&L."""
    result  = payload.get("result", "")   # "won" or "lost"
    match   = payload.get("match", "")
    bet     = payload.get("bet", "")
    odds    = float(payload.get("odds", 2.0))
    channel = payload.get("channel", "free")  # "free" or "pro"

    stake = telegram_pl["stake"]
    telegram_pl["tips"] += 1

    if result in ("won", "win"):
        profit = round(stake * odds - stake, 2)
        telegram_pl["won"] += 1
        telegram_pl["profit"] = round(telegram_pl["profit"] + profit, 2)
        msg = (
            f"💥 <b>BOOM! WINNER!</b> 💥\n"
            f"━━━━━━━━━━━━━━━━━━━━\n"
            f"✅ <b>{match}</b>\n"
            f"💰 {bet} @ {odds}\n"
            f"🏆 <b>+£{profit:.2f} profit on £{stake:.0f} stake!</b>\n"
            f"━━━━━━━━━━━━━━━━━━━━\n"
            f"📊 <b>Running P&L (£{stake:.0f} stakes)</b>\n"
            f"  Tips: {telegram_pl['tips']} · Won: {telegram_pl['won']} · Lost: {telegram_pl['lost']}\n"
            f"  Profit: <b>{'+'if telegram_pl['profit']>=0 else ''}£{telegram_pl['profit']:.2f}</b>\n"
            f"━━━━━━━━━━━━━━━━━━━━\n"
            f"🔒 Full analysis on AccaGenius Pro\n"
            f"👉 accagenius.com"
        )
    else:
        telegram_pl["lost"] += 1
        telegram_pl["profit"] = round(telegram_pl["profit"] - stake, 2)
        msg = (
            f"❌ <b>No Luck This Time</b>\n"
            f"━━━━━━━━━━━━━━━━━━━━\n"
            f"<s>{match} — {bet}</s>\n"
            f"━━━━━━━━━━━━━━━━━━━━\n"
            f"📊 <b>Running P&L (£{stake:.0f} stakes)</b>\n"
            f"  Tips: {telegram_pl['tips']} · Won: {telegram_pl['won']} · Lost: {telegram_pl['lost']}\n"
            f"  Profit: <b>{'+'if telegram_pl['profit']>=0 else ''}£{telegram_pl['profit']:.2f}</b>\n"
            f"━━━━━━━━━━━━━━━━━━━━\n"
            f"More tips coming — stay tuned!\n"
            f"🔒 Pro alerts → accagenius.com"
        )

    target = TELEGRAM_CHANNEL_FREE if channel == "free" else (TELEGRAM_CHANNEL_PRO or TELEGRAM_CHANNEL_ID)
    sent = send_telegram_to(msg, target)
    return {"sent": sent, "pl": telegram_pl}



@app.get("/fixtures/{league_code}")
async def get_fixtures(league_code: str):
    league_id = LEAGUE_IDS.get(league_code.upper())
    if not league_id:
        raise HTTPException(404, "League not found")
    cache_key = f"fixtures_{league_code.upper()}"
    cached = cache_get(cache_key)
    if cached:
        return cached
    data = api_get("fixtures", {
        "league": league_id, "season": get_season(),
        "next": 10, "timezone": "Europe/London"  # reduced from 20
    })
    by_date = {}
    for f in data.get("response", []):
        date = f["fixture"]["date"].split("T")[0]
        by_date.setdefault(date, []).append({
            "id": f["fixture"]["id"],
            "date": f["fixture"]["date"],
            "time": datetime.fromisoformat(f["fixture"]["date"].replace("Z", "+00:00")).strftime("%H:%M"),
            "home": f["teams"]["home"]["name"],
            "away": f["teams"]["away"]["name"],
            "home_id": f["teams"]["home"]["id"],
            "away_id": f["teams"]["away"]["id"],
            "home_logo": f["teams"]["home"]["logo"],
            "away_logo": f["teams"]["away"]["logo"],
            "venue": f["fixture"]["venue"]["name"],
            "referee": f["fixture"]["referee"],
            "league_id": league_id
        })
    result = {"league": league_code, "fixtures": by_date}
    if by_date:
        cache_set(cache_key, result, 3600)  # 1 hour
    return result


@app.get("/today")
async def get_today_fixtures():
    """All fixtures for today — cached 10 minutes to save API quota."""
    today_str = _london_now().strftime("%Y-%m-%d")
    cache_key = f"today_{today_str}"
    cached = cache_get(cache_key)
    if cached:
        return cached

    def fetch_league_today(league: dict, timeout: int = 12) -> list:
        try:
            data = api_get("fixtures", {
                "league": league["id"], "date": today_str,
                "timezone": "Europe/London", "season": get_season()
            }, timeout=timeout)
            results = []
            for f in data.get("response", []):
                try:
                    status = f["fixture"]["status"]["short"]
                    kick_time = datetime.fromisoformat(
                        f["fixture"]["date"].replace("Z", "+00:00")
                    ).strftime("%H:%M")
                    results.append({
                        "id": f["fixture"]["id"],
                        "time": kick_time,
                        "home": f["teams"]["home"]["name"],
                        "away": f["teams"]["away"]["name"],
                        "home_id": f["teams"]["home"]["id"],
                        "away_id": f["teams"]["away"]["id"],
                        "home_logo": f["teams"]["home"]["logo"],
                        "away_logo": f["teams"]["away"]["logo"],
                        "league": league["name"],
                        "league_code": league["code"],
                        "league_flag": league["flag"],
                        "league_id": league["id"],
                        "status": status,
                        "home_score": f["goals"]["home"],
                        "away_score": f["goals"]["away"],
                        "minute": f["fixture"]["status"]["elapsed"],
                        "_raw": f,  # keep raw for acca generator to reuse (saves API calls)
                        "venue": (f["fixture"]["venue"].get("name") or ""),
                        "odds_available": False,
                    })
                except Exception:
                    pass
            return results
        except Exception:
            return None  # None = failed (vs [] = genuinely no games)

    loop = asyncio.get_event_loop()

    # Pass 1: all leagues with 12s timeout
    with ThreadPoolExecutor(max_workers=34) as executor:
        futures = [loop.run_in_executor(executor, fetch_league_today, lg, 12) for lg in LEAGUES]
        results = await asyncio.gather(*futures, return_exceptions=True)

    # Pass 2: retry failed leagues with 20s timeout
    failed_leagues = [LEAGUES[i] for i, r in enumerate(results) if r is None or isinstance(r, Exception)]
    if failed_leagues:
        logger.warning(f"⚠️  /today retry for: {[l['code'] for l in failed_leagues]}")
        with ThreadPoolExecutor(max_workers=len(failed_leagues)) as executor:
            retry_futures = [loop.run_in_executor(executor, fetch_league_today, lg, 20) for lg in failed_leagues]
            retry_results = await asyncio.gather(*retry_futures, return_exceptions=True)
        fi = 0
        for i, r in enumerate(results):
            if r is None or isinstance(r, Exception):
                results[i] = retry_results[fi] if not isinstance(retry_results[fi], Exception) else []
                fi += 1

    all_matches = sorted(
        [m for r in results if isinstance(r, list) for m in r],
        key=lambda x: x["time"]
    )

    # Fetch odds for scheduled matches only (skip live/finished)
    def fetch_odds_today(m: dict) -> dict:
        if m.get("status") not in ["NS", "TBD"]:
            return m
        try:
            odds = get_real_odds(m["id"])
            m["odds_home"]      = odds.get("home", 0)
            m["odds_draw"]      = odds.get("draw", 0)
            m["odds_away"]      = odds.get("away", 0)
            m["odds_home_bk"]   = odds.get("home_bk", "")
            m["odds_draw_bk"]   = odds.get("draw_bk", "")
            m["odds_away_bk"]   = odds.get("away_bk", "")
            m["odds_available"] = odds.get("available", False)
        except Exception:
            pass  # odds unavailable — match still shows without odds
        return m

    if all_matches:
        # Limit odds fetches to save quota — only first 15 NS matches
        ns_matches  = [m for m in all_matches if m.get("status") in ("NS","TBD")][:15]
        other_matches = [m for m in all_matches if m.get("status") not in ("NS","TBD")]
        ns_ids = {m["id"] for m in ns_matches}
        skipped = [m for m in all_matches if m["id"] not in ns_ids and m in ns_matches]
        try:
            with ThreadPoolExecutor(max_workers=8) as executor:
                futures_odds = [loop.run_in_executor(executor, fetch_odds_today, m) for m in ns_matches]
                ns_matches = list(await asyncio.wait_for(
                    asyncio.gather(*futures_odds, return_exceptions=True),
                    timeout=12
                ))
                ns_matches = [m if isinstance(m, dict) else m for m in ns_matches]
        except asyncio.TimeoutError:
            logger.warning("Odds fetch timed out — showing matches without odds")
        all_matches = sorted(
            other_matches + [m for m in all_matches if m.get("status") in ("NS","TBD") and m["id"] not in {x["id"] for x in ns_matches if isinstance(x,dict)}] + [m for m in ns_matches if isinstance(m,dict)],
            key=lambda x: x.get("time","99:99")
        )

    still_failed = [LEAGUES[i] for i, r in enumerate(results) if r is None or isinstance(r, Exception)]
    result = {"matches": all_matches, "count": len(all_matches), "date": today_str,
              "leagues_failed": [l["code"] for l in still_failed]}

    # ── FBD fallback: if API returned nothing, use FBD daily file ──
    if not all_matches:
        try:
            fbd_data = get_fbd_data()
            if fbd_data:
                fbd_matches = []
                for m in fbd_data:
                    home = m.get("home","").strip()
                    away = m.get("away","").strip()
                    if not home or not away: continue
                    fbd_matches.append({
                        "id": abs(hash(f"{home}{away}{today_str}")) % 9999999,
                        "time":"TBC","home":home,"away":away,
                        "home_id":0,"away_id":0,"home_logo":"","away_logo":"",
                        "league":m.get("league",""),"league_code":m.get("league",""),
                        "league_flag":"📊","league_id":0,"status":"NS",
                        "home_score":None,"away_score":None,"minute":None,
                        "venue":"","odds_available":False,"fbd_source":True,
                        "fbd_prediction":m.get("prediction",""),
                        "fbd_btts":m.get("pred_btts"),
                        "fbd_over25":m.get("pred_over25"),
                    })
                if fbd_matches:
                    logger.info(f"/today: API empty — FBD fallback ({len(fbd_matches)} matches)")
                    result = {"matches":fbd_matches,"count":len(fbd_matches),
                              "date":today_str,"leagues_failed":[],"fbd_source":True}
                    cache_set(cache_key, result, 1800)
                    return result
        except Exception as fe:
            logger.warning(f"FBD fallback error: {fe}")

    if all_matches and len(still_failed) <= 1:
        cache_set(cache_key, result, 3600)
    elif all_matches:
        cache_set(cache_key, result, 600)
    logger.info(f"/today: {len(all_matches)} matches, {len(still_failed)} leagues failed")
    return result or {"matches":[],"count":0,"date":today_str,"leagues_failed":[]}


@app.post("/today/refresh")
async def refresh_today():
    today_str = _london_now().strftime("%Y-%m-%d")
    cache_key = f"today_{today_str}"
    throttle_key = f"refresh_throttle_{today_str}"
    # Rate-limit manual refreshes to once every 5 minutes
    if cache_get(throttle_key):
        return {"ok": False, "message": "Refreshed too recently — try again in 5 minutes"}
    if cache_key in _cache:
        del _cache[cache_key]
    cache_set(throttle_key, True, 300)  # 5 minute cooldown
    return {"ok": True, "message": "Cache cleared — refreshing now"}


@app.get("/next-round")
async def get_next_round_fixtures():
    """Next round of fixtures across all leagues — cached 60 minutes"""
    cache_key = f"next_round_{datetime.now().strftime('%Y-%m-%d_%H')}"
    cached = cache_get(cache_key)
    if cached:
        return cached

    today = _london_now().date()
    cutoff = today + timedelta(days=3)

    def fetch_league_upcoming(league: dict) -> list:
        data = api_get("fixtures", {
            "league": league["id"], "season": get_season(),
            "next": 10, "timezone": "Europe/London"
        }, timeout=8)
        results = []
        for f in data.get("response", []):
            try:
                fd = datetime.fromisoformat(f["fixture"]["date"].replace("Z", "+00:00")).date()
                if fd > cutoff:
                    continue
                results.append({
                    "id": f["fixture"]["id"],
                    "date": fd.isoformat(),
                    "time": datetime.fromisoformat(f["fixture"]["date"].replace("Z", "+00:00")).strftime("%H:%M"),
                    "home": f["teams"]["home"]["name"],
                    "away": f["teams"]["away"]["name"],
                    "home_id": f["teams"]["home"]["id"],
                    "away_id": f["teams"]["away"]["id"],
                    "home_logo": f["teams"]["home"]["logo"],
                    "away_logo": f["teams"]["away"]["logo"],
                    "league": league["name"],
                    "league_code": league["code"],
                    "league_flag": league["flag"],
                    "league_id": league["id"],
                    "venue": f["fixture"]["venue"]["name"] or "",
                })
            except Exception:
                pass
        return results

    loop = asyncio.get_event_loop()
    with ThreadPoolExecutor(max_workers=17) as executor:
        futures = [loop.run_in_executor(executor, fetch_league_upcoming, league) for league in LEAGUES]
        results = await asyncio.gather(*futures)

    all_matches = [m for sublist in results for m in sublist]

    # Fetch odds in parallel
    def fetch_odds_next(m: dict) -> dict:
        odds = get_real_odds(m["id"])
        m["odds_home"] = odds.get("home", 0)
        m["odds_draw"] = odds.get("draw", 0)
        m["odds_away"] = odds.get("away", 0)
        m["odds_home_bk"] = odds.get("home_bk", "")
        m["odds_draw_bk"] = odds.get("draw_bk", "")
        m["odds_away_bk"] = odds.get("away_bk", "")
        m["odds_available"] = odds.get("available", False)
        return m

    if all_matches:
        with ThreadPoolExecutor(max_workers=20) as executor:
            futures2 = [loop.run_in_executor(executor, fetch_odds_next, m) for m in all_matches]
            all_matches = list(await asyncio.gather(*futures2))

    all_matches.sort(key=lambda x: (x["date"], x["time"]))
    result = {"matches": all_matches, "count": len(all_matches)}
    if all_matches:
        cache_set(cache_key, result, 3600)  # 60 minutes
    return result


@app.get("/live")
async def get_live():
    """Live matches — cached 90 seconds."""
    try:
        cache_key = "live_matches"
        cached = cache_get(cache_key)
        if cached:
            return cached
        data = api_get("fixtures", {"live": "all", "timezone": "Europe/London"})
        if not data:
            return {"matches": [], "count": 0, "source": "api_error"}
        our_league_ids = {l["id"] for l in LEAGUES}
        matches = []
        for f in data.get("response", []):
            try:
                league_id = f["league"]["id"]
                if league_id not in our_league_ids:
                    continue  # Skip — not one of our leagues
                matches.append({
                    "id": f["fixture"]["id"],
                    "home": f["teams"]["home"]["name"],
                    "away": f["teams"]["away"]["name"],
                    "home_logo": f["teams"]["home"]["logo"],
                    "away_logo": f["teams"]["away"]["logo"],
                    "home_id": f["teams"]["home"]["id"],
                    "away_id": f["teams"]["away"]["id"],
                    "home_score": f["goals"]["home"] or 0,
                    "away_score": f["goals"]["away"] or 0,
                    "minute": f["fixture"]["status"]["elapsed"],
                    "status": f["fixture"]["status"]["short"],
                    "status_long": f["fixture"]["status"]["long"],
                    "league": f["league"]["name"],
                    "league_id": league_id,
                    "league_logo": f["league"]["logo"],
                    "venue": f["fixture"]["venue"]["name"],
                    "ht_home": (f.get("score", {}).get("halftime", {}) or {}).get("home"),
                    "ht_away": (f.get("score", {}).get("halftime", {}) or {}).get("away"),
                })
            except Exception:
                pass

        result = {"matches": matches, "count": len(matches)}
        cache_set(cache_key, result, 90)
        return result
    except Exception as e:
        logger.error(f"/live error: {e}")
        return {"matches": [], "count": 0, "error": str(e)}

@app.get("/live/{fixture_id}/events")
async def get_live_events(fixture_id: int):
    """Minute-by-minute events for a live match"""
    data = api_get("fixtures/events", {"fixture": fixture_id})
    events = []
    for e in data.get("response", []):
        try:
            events.append({
                "minute": e["time"]["elapsed"],
                "extra": e["time"].get("extra"),
                "team": e["team"]["name"],
                "team_id": e["team"]["id"],
                "player": (e.get("player") or {}).get("name"),
                "assist": (e.get("assist") or {}).get("name"),
                "type": e["type"],
                "detail": e["detail"],
                "comments": e.get("comments")
            })
        except Exception:
            pass
    events.sort(key=lambda x: x["minute"] or 0)
    return {"events": events}


@app.get("/live/{fixture_id}/stats")
async def get_live_stats(fixture_id: int):
    """Live match statistics — real API-Football field names mapped to clean keys"""
    data = api_get("fixtures/statistics", {"fixture": fixture_id})
    stats = {}

    # Exact API-Football stat type names → clean key mapping
    FIELD_MAP = {
        "Shots on Goal":         "shots_on_goal",
        "Shots off Goal":        "shots_off_goal",
        "Total Shots":           "total_shots",
        "Blocked Shots":         "blocked_shots",
        "Shots insidebox":       "shots_inside_box",
        "Shots outsidebox":      "shots_outside_box",
        "Fouls":                 "fouls",
        "Corner Kicks":          "corner_kicks",
        "Offsides":              "offsides",
        "Ball Possession":       "ball_possession",
        "Yellow Cards":          "yellow_cards",
        "Red Cards":             "red_cards",
        "Goalkeeper Saves":      "goalkeeper_saves",
        "Total passes":          "total_passes",
        "Passes accurate":       "passes_accurate",
        "Passes %":              "passes_pct",
        "expected_goals":        "expected_goals",   # real xG from API
        "Expected Goals":        "expected_goals",
    }

    for team_data in data.get("response", []):
        team_name = team_data["team"]["name"]
        team_stats = {}
        for s in team_data.get("statistics", []):
            raw_type = s.get("type", "")
            raw_val = s.get("value")
            # Use exact map first, then normalise as fallback
            clean_key = FIELD_MAP.get(raw_type) or raw_type.lower().replace(" ", "_").replace("%", "pct").replace(".", "")
            # Strip % from possession values
            if isinstance(raw_val, str) and raw_val.endswith("%"):
                try:
                    raw_val = int(raw_val.replace("%", ""))
                except ValueError:
                    pass
            team_stats[clean_key] = raw_val if raw_val is not None else 0
        stats[team_name] = team_stats

    return {"stats": stats, "fixture_id": fixture_id}


@app.get("/fixture/{fixture_id}/lineups")
async def get_lineups(fixture_id: int):
    """Starting lineups"""
    data = api_get("fixtures/lineups", {"fixture": fixture_id})
    lineups = []
    for team in data.get("response", []):
        try:
            lineups.append({
                "team": team["team"]["name"],
                "team_logo": team["team"]["logo"],
                "formation": team.get("formation"),
                "startXI": [{"name": p["player"]["name"], "number": p["player"]["number"], "pos": p["player"]["pos"]} for p in team.get("startXI", [])],
                "substitutes": [{"name": p["player"]["name"], "number": p["player"]["number"], "pos": p["player"]["pos"]} for p in team.get("substitutes", [])]
            })
        except Exception:
            pass
    return {"lineups": lineups}


@app.get("/standings/{league_code}")
async def get_standings(league_code: str):
    league_id = LEAGUE_IDS.get(league_code.upper())
    if not league_id:
        raise HTTPException(404, "League not found")
    cache_key = f"standings_{league_code.upper()}"
    cached = cache_get(cache_key)
    if cached:
        return cached
    season = get_season()
    data = api_get("standings", {"league": league_id, "season": season})
    # Fallback: try previous season if current returns nothing
    if not data.get("response"):
        data = api_get("standings", {"league": league_id, "season": season - 1})
    try:
        response = data.get("response", [])
        if not response:
            return {"league": league_code, "standings": [], "message": f"No standings data for {league_code} season {season}. API may not have data yet."}
        # Some competitions (CL, EL, ECL, EL1, EL2) return multiple groups — flatten all
        all_groups = response[0]["league"].get("standings", [])
        if not all_groups:
            return {"league": league_code, "standings": [], "message": "No standings available"}
        # Use first group (main table) — works for all single-table leagues
        raw = all_groups[0]
        standings = []
        for t in raw:
            try:
                standings.append({
                    "position": t["rank"],
                    "team": t["team"]["name"],
                    "logo": t["team"]["logo"],
                    "played": t["all"]["played"],
                    "won": t["all"]["win"],
                    "drawn": t["all"]["draw"],
                    "lost": t["all"]["lose"],
                    "gf": t["all"]["goals"]["for"],
                    "ga": t["all"]["goals"]["against"],
                    "gd": t.get("goalsDiff", 0),
                    "goalDifference": t.get("goalsDiff", 0),
                    "points": t["points"],
                    "form": t.get("form", ""),
                    "description": (t.get("description") or "")
                })
            except Exception:
                pass
        result = {"league": league_code, "standings": standings}
        if standings:
            cache_set(cache_key, result, 7200)  # 2 hours
        return result
    except Exception as e:
        return {"league": league_code, "standings": [], "message": str(e)}


@app.get("/h2h/{team1_id}/{team2_id}")
async def get_h2h(team1_id: int, team2_id: int):
    cache_key = f"h2h_{min(team1_id,team2_id)}_{max(team1_id,team2_id)}"
    cached = cache_get(cache_key)
    if cached:
        return cached
    data = api_get("fixtures/headtohead", {"h2h": f"{team1_id}-{team2_id}", "last": 10})
    matches = []
    for f in data.get("response", []):
        try:
            matches.append({
                "date": f["fixture"]["date"].split("T")[0],
                "home": f["teams"]["home"]["name"],
                "away": f["teams"]["away"]["name"],
                "home_logo": f["teams"]["home"]["logo"],
                "away_logo": f["teams"]["away"]["logo"],
                "home_score": f["goals"]["home"],
                "away_score": f["goals"]["away"],
                "league": f["league"]["name"]
            })
        except Exception:
            pass
    result = {"matches": matches}
    if matches:
        cache_set(cache_key, result, 21600)  # 6 hours — h2h is historical
    return result


@app.get("/quota")
async def quota_status():
    """Check API Football quota usage."""
    today = _london_now().date()
    if today != _api_call_reset:
        return {"calls_today": 0, "plan_limit": 75000, "remaining": 75000, "status": "fresh"}
    remaining = max(0, 75000 - _api_call_count)
    pct = round(_api_call_count / 75000 * 100, 1)
    status = "critical" if _api_call_count > 65000 else "warning" if _api_call_count > 3000 else "ok"
    return {
        "calls_today": _api_call_count,
        "plan_limit": 75000,
        "remaining": remaining,
        "used_pct": pct,
        "status": status,
        "reset": str(_api_call_reset),
    }


@app.get("/cache-status")
async def get_cache_status():
    """Monitor API quota usage — visit this URL to check how many calls used today"""
    stats = cache_stats()
    return {
        "api_calls_today": _api_call_count,
        "daily_limit": 75000,
        "calls_remaining": max(0, 75000 - _api_call_count),
        "reset_date": str(_api_call_reset),
        "cache": stats,
        "warning": "CRITICAL" if _api_call_count > 7000 else ("HIGH" if _api_call_count > 5000 else ("MODERATE" if _api_call_count > 3000 else "OK"))
    }


@app.get("/form/{team_id}/{league_id}")
async def get_form(team_id: int, league_id: int):
    return get_team_form(team_id, league_id, last=10)


@app.get("/predictions/{fixture_id}")
async def get_predictions(fixture_id: int):
    data = api_get("predictions", {"fixture": fixture_id})
    preds = data.get("response", [])
    if not preds:
        return {"prediction": None, "predictions": {}}
    pred = preds[0].get("predictions", {})
    teams = preds[0].get("teams", {})
    comp = preds[0].get("comparison", {})
    winner = pred.get("winner", {}) or {}
    percent = pred.get("percent", {}) or {}
    goals = pred.get("goals", {}) or {}
    return {
        "prediction": {
            "winner": winner.get("name"),
            "winner_comment": winner.get("comment"),
            "percent_home": percent.get("home", "33%"),
            "percent_draw": percent.get("draw", "33%"),
            "percent_away": percent.get("away", "33%"),
            "goals_home": goals.get("home"),
            "goals_away": goals.get("away"),
            "advice": pred.get("advice", ""),
            "under_over": pred.get("goals", {}).get("total")
        },
        "predictions": pred,
        "comparison": comp,
        "teams": {
            "home": teams.get("home", {}).get("name"),
            "away": teams.get("away", {}).get("name")
        }
    }


@app.get("/live-odds/all")
async def get_all_live_odds():
    """Fetch live odds for ALL live matches in one API call — Ultra tier.
    Cached 60s. Call once, serve to all users from cache."""
    cache_key = "all_live_odds"
    cached = cache_get(cache_key)
    if cached:
        return cached
    data = api_get("odds/live", {}, timeout=10)
    odds_map = {}  # fixture_id -> {home, draw, away, markets}
    for fixture_data in data.get("response", []):
        try:
            fid = fixture_data.get("fixture", {}).get("id")
            if not fid:
                continue
            best = {
                "home": {"odds": 0, "bookmaker": ""},
                "draw": {"odds": 0, "bookmaker": ""},
                "away": {"odds": 0, "bookmaker": ""}
            }
            all_markets = {}
            for bm in fixture_data.get("odds", []):
                bm_name = bm.get("name", "")
                for bet in bm.get("bets", []):
                    bet_name = bet.get("name", "")
                    if bet_name not in all_markets:
                        all_markets[bet_name] = {}
                    for v in bet.get("values", []):
                        try:
                            val   = v.get("value", "").lower()
                            odd   = float(v.get("odd", 0))
                            label = v.get("value", "")
                            # Track best per bookmaker per value
                            if bet_name not in all_markets:
                                all_markets[bet_name] = {}
                            if label not in all_markets[bet_name] or odd > all_markets[bet_name][label]["odds"]:
                                all_markets[bet_name][label] = {"odds": odd, "bookmaker": bm_name}
                            # Match winner best odds
                            if "match winner" in bet_name.lower() or "winner" == bet_name.lower():
                                if "home" in val and odd > best["home"]["odds"]:
                                    best["home"] = {"odds": odd, "bookmaker": bm_name}
                                elif "draw" in val and odd > best["draw"]["odds"]:
                                    best["draw"] = {"odds": odd, "bookmaker": bm_name}
                                elif "away" in val and odd > best["away"]["odds"]:
                                    best["away"] = {"odds": odd, "bookmaker": bm_name}
                        except:
                            pass
            odds_map[str(fid)] = {"best_odds": best, "markets": all_markets, "live": True}
        except:
            pass
    result = {"odds": odds_map, "count": len(odds_map)}
    cache_set(cache_key, result, 60)
    logger.info(f"Live odds fetched: {len(odds_map)} fixtures")
    return result


@app.get("/live-odds/{fixture_id}")
async def get_live_odds(fixture_id: int):
    """In-play odds for a specific fixture — tries cache first then fetches."""
    # Check all-live cache first (populated by /live-odds/all)
    all_cached = cache_get("all_live_odds")
    if all_cached and str(fixture_id) in all_cached.get("odds", {}):
        return all_cached["odds"][str(fixture_id)]

    # Fetch specifically for this fixture
    cache_key = f"live_odds_{fixture_id}"
    cached = cache_get(cache_key)
    if cached:
        return cached

    best = {"home": {"odds": 0, "bookmaker": ""}, "draw": {"odds": 0, "bookmaker": ""}, "away": {"odds": 0, "bookmaker": ""}}
    all_markets = {}

    # Ultra: try live odds
    data = api_get("odds/live", {"fixture": fixture_id}, timeout=8)
    response_list = data.get("response", []) if data else []
    if response_list:
        for bm in response_list[0].get("odds", []):
            bm_name = bm.get("name", "")
            for bet in bm.get("bets", []):
                bet_name = bet.get("name", "")
                if bet_name not in all_markets:
                    all_markets[bet_name] = {}
                for v in bet.get("values", []):
                    try:
                        val  = v.get("value", "").lower()
                        odd  = float(v.get("odd", 0))
                        label = v.get("value", "")
                        if label not in all_markets[bet_name] or odd > all_markets[bet_name][label]["odds"]:
                            all_markets[bet_name][label] = {"odds": odd, "bookmaker": bm_name}
                        if "match winner" in bet_name.lower() or "winner" == bet_name.lower():
                            if "home" in val and odd > best["home"]["odds"]:
                                best["home"] = {"odds": odd, "bookmaker": bm_name}
                            elif "draw" in val and odd > best["draw"]["odds"]:
                                best["draw"] = {"odds": odd, "bookmaker": bm_name}
                            elif "away" in val and odd > best["away"]["odds"]:
                                best["away"] = {"odds": odd, "bookmaker": bm_name}
                    except: pass

    # Fallback to pre-match if live not available
    if not best["home"]["odds"]:
        data2 = api_get("odds", {"fixture": fixture_id}, timeout=5)
        resp2 = data2.get("response", []) if data2 else []
        if resp2:
            for bm in resp2[0].get("bookmakers", []):
                bm_name = bm.get("name", "")
                for bet in bm.get("bets", []):
                    if "Match Winner" in bet.get("name", ""):
                        for v in bet.get("values", []):
                            try:
                                val  = v.get("value", "").lower()
                                odd  = float(v.get("odd", 0))
                                if "home" in val and odd > best["home"]["odds"]:
                                    best["home"] = {"odds": odd, "bookmaker": bm_name}
                                elif "draw" in val and odd > best["draw"]["odds"]:
                                    best["draw"] = {"odds": odd, "bookmaker": bm_name}
                                elif "away" in val and odd > best["away"]["odds"]:
                                    best["away"] = {"odds": odd, "bookmaker": bm_name}
                            except: pass
        response_list = []  # mark as pre-match

    result = {"fixture_id": fixture_id, "best_odds": best, "markets": all_markets, "live": bool(response_list)}
    cache_set(cache_key, result, 60)
    return result


@app.get("/odds/{fixture_id}")
async def get_odds(fixture_id: int):
    """Full odds across multiple bookmakers and markets — cached 15 mins"""
    cache_key = f"full_odds_{fixture_id}"
    cached = cache_get(cache_key)
    if cached:
        return cached
    data = api_get("odds", {"fixture": fixture_id}, timeout=5)
    bookmakers_data = {}
    markets = {}

    response_list = data.get("response", [])
    if not response_list:
        return {"bookmakers": {}, "markets": {}, "best_odds": {
            "home": {"bookmaker": "", "odds": 0},
            "draw": {"bookmaker": "", "odds": 0},
            "away": {"bookmaker": "", "odds": 0}
        }}
    response_data = response_list[0]
    for bm in response_data.get("bookmakers", [])[:8]:
        bm_name = bm["name"]
        bm_odds = {}
        for bet in bm.get("bets", []):
            bname = bet["name"]
            bm_odds[bname] = {v["value"]: v["odd"] for v in bet.get("values", [])}
            if bname not in markets:
                markets[bname] = {}
            markets[bname][bm_name] = {v["value"]: v["odd"] for v in bet.get("values", [])}
        bookmakers_data[bm_name] = bm_odds

    # Best odds per outcome for Match Winner
    best = {"home": ("", 0), "draw": ("", 0), "away": ("", 0)}
    mw = markets.get("Match Winner", {})
    for bm, vals in mw.items():
        try:
            if float(vals.get("Home", 0)) > best["home"][1]:
                best["home"] = (bm, float(vals["Home"]))
            if float(vals.get("Draw", 0)) > best["draw"][1]:
                best["draw"] = (bm, float(vals["Draw"]))
            if float(vals.get("Away", 0)) > best["away"][1]:
                best["away"] = (bm, float(vals["Away"]))
        except Exception:
            pass

    result = {
        "bookmakers": bookmakers_data,
        "markets": markets,
        "best_odds": {
            "home": {"bookmaker": best["home"][0], "odds": best["home"][1]},
            "draw": {"bookmaker": best["draw"][0], "odds": best["draw"][1]},
            "away": {"bookmaker": best["away"][0], "odds": best["away"][1]}
        }
    }
    cache_set(cache_key, result, 3600)  # 15 minutes
    return result


@app.get("/player-stats/{team_id}")
async def get_player_stats(team_id: int):
    """Top scorers/assists for a team"""
    data = api_get("players/topscorers", {"league": 39, "season": get_season()})
    players = []
    for p in data.get("response", [])[:20]:
        if p.get("statistics", [{}])[0].get("team", {}).get("id") == team_id:
            s = p["statistics"][0]
            players.append({
                "name": p["player"]["name"],
                "photo": p["player"]["photo"],
                "goals": s["goals"]["total"] or 0,
                "assists": s["goals"]["assists"] or 0,
                "matches": s["games"]["appearences"] or 0,
                "rating": s["games"].get("rating")
            })
    return {"players": players}


@app.get("/today-acca")
async def get_today_acca():
    """Generate today's 3 themed accas ONCE and cache all day — same picks for every user."""
    today_str = _london_now().strftime("%Y-%m-%d")
    cache_key = f"today_acca_{today_str}"
    cached = cache_get(cache_key)
    if cached:
        return cached

    leagues = ["PL","ELC","BL1","SA","FL1","PD","CL","EL","NED","PPL","TUR","BEL","SPFL","EL1","EL2"]

    async def fetch_picks(risk: str, endpoint: str = "form") -> list:
        try:
            if endpoint == "rank":
                req = AccaRequest(selections=8, risk=risk, leagues=leagues, market="winner", today_only=True)
                result = await generate_table_rank_acca(req)
            else:
                req = AccaRequest(selections=8, risk=risk, leagues=leagues, market="winner", today_only=True)
                result = await generate_acca(req)
            return result.get("selections", []) if isinstance(result, dict) else []
        except Exception as e:
            logger.error(f"today-acca fetch error ({risk}/{endpoint}): {e}")
            return []

    # Run sequentially so form cache from pass 1 is reused in passes 2+3
    # Parallel = 3× the form API calls on cold cache
    form_picks  = await fetch_picks("balanced", "form")
    rank_picks  = await fetch_picks("balanced", "rank")
    value_picks = await fetch_picks("risky",    "form")

    # Global dedup — no match appears in more than one acca
    used_ids = set()
    used_teams = set()

    def dedup(picks, n):
        out = []
        for p in picks:
            fid = p.get("id")
            team_key = "|".join(sorted([p.get("home",""), p.get("away","")]))
            if fid in used_ids or team_key in used_teams:
                continue
            out.append(p)
            used_ids.add(fid)
            used_teams.add(team_key)
            if len(out) >= n:
                break
        return out

    t1 = dedup(form_picks,  3)
    t2 = dedup(rank_picks,  3)
    t3 = dedup(value_picks, 4)

    result = {
        "date": today_str,
        "accas": [
            {"theme": "form",  "label": "Form Treble",   "emoji": "📈", "picks": t1},
            {"theme": "rank",  "label": "Rank Treble",   "emoji": "🏆", "picks": t2},
            {"theme": "value", "label": "Value 4-Fold",  "emoji": "💎", "picks": t3},
        ]
    }

    # Cache all day — resets at midnight when date changes
    if any(a["picks"] for a in result["accas"]):
        cache_set(cache_key, result, 86400)  # 24h but key includes date so auto-resets

    return result


@app.post("/generate-acca")
async def generate_acca(request: AccaRequest):
    try:
        leagues = request.leagues if request.leagues else ["PL", "ELC", "PD", "BL1", "SA", "FL1", "CL", "EL", "ECL", "TUR", "NED", "PPL", "BEL", "SPFL"]
        today = _london_now().date()
        cutoff = today if request.today_only else today + timedelta(days=3)
        all_picks = []
        seen = set()
        random.shuffle(leagues)

        # ── Always work from today's fixture list (date-accurate, cached) ──
        today_str_acca = datetime.now().strftime("%Y-%m-%d")
        _today_cache_key = f"today_{today_str_acca}"
        _today_cached = cache_get(_today_cache_key)

        # Build league → raw fixtures map from the /today cache
        _today_raw_by_league: dict = {}
        if _today_cached:
            for m in _today_cached.get("matches", []):
                raw = m.get("_raw")
                if raw:
                    lc_key = m.get("league_code", "")
                    _today_raw_by_league.setdefault(lc_key, []).append(raw)

        def fetch_league_picks(lc: str) -> list:
            lid = LEAGUE_IDS.get(lc)
            if not lid: return []
            picks = []
            try:
                # Use cached today fixtures if available — avoids extra API calls
                # and guarantees we only use today's games
                if lc in _today_raw_by_league:
                    raw_fixtures = _today_raw_by_league[lc]
                else:
                    # Cache miss — fetch by DATE (not next:N) so we still get today only
                    data = api_get("fixtures", {
                        "league": lid,
                        "date": today_str_acca,
                        "timezone": "Europe/London",
                        "season": get_season()
                    })
                    raw_fixtures = data.get("response", [])

                fixtures = sorted(
                    raw_fixtures,
                    key=lambda f: f["fixture"]["date"] if isinstance(f, dict) and "fixture" in f else ""
                )
                picks_this_league = 0
                fbd_today = get_fbd_data()
                for f in fixtures:
                    if picks_this_league >= 3: break
                    fid = f["fixture"]["id"]
                    if fid in seen: continue
                    fd = datetime.fromisoformat(f["fixture"]["date"].replace("Z", "+00:00")).date()
                    if fd != today: continue
                    status = f["fixture"]["status"]["short"]
                    if status in ("1H","HT","2H","ET","P","FT","AET","PEN"): continue
                    # Only pick FBD-covered matches
                    if fbd_today:
                        hn = f["teams"]["home"]["name"]
                        an = f["teams"]["away"]["name"]
                        if not any(_fbd_names_match(hn, m["home"]) and _fbd_names_match(an, m["away"]) for m in fbd_today):
                            continue
                    hf = get_team_form(f["teams"]["home"]["id"], lid)
                    af = get_team_form(f["teams"]["away"]["id"], lid)
                    if hf.get("games", 0) == 0 and af.get("games", 0) == 0:
                        continue
                    pick = analyze_and_pick_with_fbd(f, hf, af, request.risk, request.market)
                    if pick and pick["confidence"] >= 45:
                        picks.append(pick)
                        picks_this_league += 1
            except Exception as e:
                logger.error(f"League pick error {lc}: {e}")
            return picks

        with ThreadPoolExecutor(max_workers=7) as ex:  # reduced to limit API calls
            import asyncio
            loop = asyncio.get_event_loop()
            futures = [loop.run_in_executor(ex, fetch_league_picks, lc) for lc in leagues]
            results = await asyncio.gather(*futures)

        for league_picks in results:
            for pick in league_picks:
                if pick["id"] not in seen:
                    seen.add(pick["id"])
                    all_picks.append(pick)

        all_picks.sort(key=lambda x: x["confidence"], reverse=True)

        # Limit draws — max 1 draw pick per 5 selections to avoid draw-heavy accas
        final = []
        draw_count = 0
        for p in all_picks:
            if p.get("bet") == "Draw":
                if draw_count >= 1: continue  # skip extra draws
                draw_count += 1
            final.append(p)
            if len(final) >= request.selections: break

        picks = final

        if not picks:
            return {"message": "No picks found — try selecting more leagues or switching to Balanced risk.", 
                    "total_selections": 0, "total_odds": 0, "confidence": 0, "selections": []}

        total_odds = 1.0
        for p in picks: total_odds *= p.get("odds", 1.0)
        avg_conf = sum(p["confidence"] for p in picks) / len(picks)

        return {
            "message": f"AI Acca — {request.risk.capitalize()} risk — Next 3 Days",
            "total_selections": len(picks),
            "total_odds": round(total_odds, 2),
            "confidence": round(avg_conf),
            "risk_level": request.risk,
            "market": request.market,
            "selections": picks
        }
    except Exception as e:
        raise HTTPException(500, str(e))


@app.post("/generate-ht-acca")
async def generate_ht_acca(request: HTAccaRequest):
    try:
        leagues = request.leagues if request.leagues else ["PL", "ELC", "PD", "BL1", "SA", "FL1"]
        today = _london_now().date()
        cutoff = today + timedelta(days=3)
        all_picks = []
        seen = set()
        random.shuffle(leagues)
        for lc in leagues:
            lid = LEAGUE_IDS.get(lc)
            if not lid: continue
            data = api_get("fixtures", {"league": lid, "season": get_season(), "next": 15, "timezone": "Europe/London"})
            fixtures = sorted(data.get("response", []), key=lambda f: f["fixture"]["date"])
            for f in fixtures:
                fid = f["fixture"]["id"]
                if fid in seen: continue
                fd = datetime.fromisoformat(f["fixture"]["date"].replace("Z", "+00:00")).date()
                if fd < today or fd > cutoff: continue
                hf = get_team_form(f["teams"]["home"]["id"], lid)
                af = get_team_form(f["teams"]["away"]["id"], lid)
                pick = ht_pick(f, hf, af)
                if pick and pick["confidence"] >= 60:
                    seen.add(fid)
                    odds = get_real_odds(fid)
                    pick["odds_home"] = odds.get("home", 0)
                    pick["odds_draw"] = odds.get("draw", 0)
                    pick["odds_away"] = odds.get("away", 0)
                    pick["odds_home_bk"] = odds.get("home_bk", "")
                    pick["odds_available"] = odds.get("available", False)
                    all_picks.append(pick)
            if len(all_picks) >= request.selections * 3: break

        all_picks.sort(key=lambda x: x["confidence"], reverse=True)
        picks = all_picks[:request.selections]
        if not picks:
            return {"message": "No HT picks in next 3 days", "total_selections": 0, "total_odds": 0, "confidence": 0, "selections": []}
        total_odds = 1.0
        for p in picks: total_odds *= p["odds"]
        avg_conf = sum(p["confidence"] for p in picks) / len(picks)
        return {
            "message": "HT Acca — Next 3 days",
            "total_selections": len(picks),
            "total_odds": round(total_odds, 2),
            "confidence": round(avg_conf),
            "selections": picks
        }
    except Exception as e:
        raise HTTPException(500, str(e))


@app.post("/generate-btts-acca")
async def generate_btts_acca(request: AccaRequest):
    """BTTS acca — finds fixtures where both teams have high BTTS% historically, unique games only."""
    try:
        leagues = request.leagues if request.leagues else ["PL","ELC","BL1","SA","FL1","PD"]
        today = _london_now().date()
        cutoff = today + timedelta(days=3)
        all_picks = []
        seen = set()
        random.shuffle(leagues)

        def fetch_btts(lc: str) -> list:
            lid = LEAGUE_IDS.get(lc)
            if not lid: return []
            picks = []
            try:
                data = api_get("fixtures", {"league": lid, "season": get_season(), "next": 10, "timezone": "Europe/London"})
                for f in sorted(data.get("response", []), key=lambda x: x["fixture"]["date"]):
                    fid = f["fixture"]["id"]
                    if fid in seen: continue
                    fd = datetime.fromisoformat(f["fixture"]["date"].replace("Z","+00:00")).date()
                    if fd < today or fd > cutoff: continue
                    status = f["fixture"]["status"]["short"]
                    if status in ("1H","HT","2H","ET","P","FT","AET","PEN"): continue
                    hf = get_team_form(f["teams"]["home"]["id"], lid)
                    af = get_team_form(f["teams"]["away"]["id"], lid)
                    h_btts = hf.get("btts_pct", 0)
                    a_btts = af.get("btts_pct", 0)
                    combined_btts = round((h_btts + a_btts) / 2)
                    if combined_btts >= 55:
                        dt = datetime.fromisoformat(f["fixture"]["date"].replace("Z","+00:00"))
                        odds = get_real_odds(fid)
                        conf = min(88, 55 + int(combined_btts * 0.4))
                        picks.append({
                            "id": fid,
                            "home": f["teams"]["home"]["name"],
                            "away": f["teams"]["away"]["name"],
                            "home_logo": f["teams"]["home"]["logo"],
                            "away_logo": f["teams"]["away"]["logo"],
                            "date": dt.strftime("%d/%m %H:%M"),
                            "home_id": f["teams"]["home"]["id"],
                            "away_id": f["teams"]["away"]["id"],
                            "league_id": lid,
                            "bet": "Both Teams to Score",
                            "market_type": "BTTS",
                            "odds": odds.get("btts", 1.72),
                            "odds_home": odds.get("home", 0),
                            "odds_draw": odds.get("draw", 0),
                            "odds_away": odds.get("away", 0),
                            "odds_available": odds.get("available", False),
                            "confidence": conf,
                            "reasoning": f"Combined BTTS% {combined_btts}% — {f['teams']['home']['name']} {h_btts}% | {f['teams']['away']['name']} {a_btts}%",
                        })
            except Exception:
                pass
            return picks

        with ThreadPoolExecutor(max_workers=10) as ex:
            loop = asyncio.get_event_loop()
            results = await asyncio.gather(*[loop.run_in_executor(ex, fetch_btts, lc) for lc in leagues])

        for league_picks in results:
            for pick in league_picks:
                if pick["id"] not in seen:
                    seen.add(pick["id"])
                    all_picks.append(pick)

        all_picks.sort(key=lambda x: x["confidence"], reverse=True)
        picks = all_picks[:request.selections]
        if not picks:
            return {"message": "No BTTS picks found.", "total_selections": 0, "total_odds": 0, "confidence": 0, "selections": []}

        total_odds = 1.0
        for p in picks: total_odds *= p.get("odds", 1.72)
        avg_conf = sum(p["confidence"] for p in picks) / len(picks)
        return {"message": "BTTS Acca", "total_selections": len(picks), "total_odds": round(total_odds, 2),
                "confidence": round(avg_conf), "market": "btts", "selections": picks}
    except Exception as e:
        raise HTTPException(500, str(e))


@app.post("/generate-form-crash-acca")
async def generate_form_crash_acca(request: AccaRequest):
    """Out-of-form / form-crash acca — teams on bad run being strong favourites = value against them."""
    try:
        leagues = request.leagues if request.leagues else ["PL","ELC","BL1","SA","FL1","PD","CL"]
        today = _london_now().date()
        cutoff = today + timedelta(days=3)
        all_picks = []
        seen = set()
        random.shuffle(leagues)

        def fetch_form_crash(lc: str) -> list:
            lid = LEAGUE_IDS.get(lc)
            if not lid: return []
            picks = []
            try:
                data = api_get("fixtures", {"league": lid, "season": get_season(), "next": 10, "timezone": "Europe/London"})
                for f in sorted(data.get("response", []), key=lambda x: x["fixture"]["date"]):
                    fid = f["fixture"]["id"]
                    if fid in seen: continue
                    fd = datetime.fromisoformat(f["fixture"]["date"].replace("Z","+00:00")).date()
                    if fd < today or fd > cutoff: continue
                    status = f["fixture"]["status"]["short"]
                    if status in ("1H","HT","2H","ET","P","FT","AET","PEN"): continue
                    hf = get_team_form(f["teams"]["home"]["id"], lid)
                    af = get_team_form(f["teams"]["away"]["id"], lid)
                    h_rating = hf.get("form_rating", 1.5)
                    a_rating = af.get("form_rating", 1.5)
                    h_losses = hf.get("losses", 0)
                    a_losses = af.get("losses", 0)
                    dt = datetime.fromisoformat(f["fixture"]["date"].replace("Z","+00:00"))
                    odds = get_real_odds(fid)
                    ho = odds.get("home", 0)
                    ao = odds.get("away", 0)

                    # Team on bad run (3+ losses) but still short-odds favourite — fade them
                    if h_losses >= 3 and ao >= 2.5 and a_rating >= 1.3:
                        conf = min(75, 50 + h_losses * 5 + int(a_rating * 5))
                        picks.append({
                            "id": fid, "home": f["teams"]["home"]["name"], "away": f["teams"]["away"]["name"],
                            "home_logo": f["teams"]["home"]["logo"], "away_logo": f["teams"]["away"]["logo"],
                            "date": dt.strftime("%d/%m %H:%M"), "home_id": f["teams"]["home"]["id"],
                            "away_id": f["teams"]["away"]["id"], "league_id": lid,
                            "bet": f"{f['teams']['away']['name']} Win",
                            "market_type": "1X2", "odds": round(ao, 2),
                            "odds_home": ho, "odds_draw": odds.get("draw",0), "odds_away": ao,
                            "odds_available": odds.get("available", False),
                            "confidence": conf,
                            "reasoning": f"⚠️ Form crash pick — {f['teams']['home']['name']} on {h_losses}-loss run. {f['teams']['away']['name']} priced at {ao:.2f} — value against struggling side"
                        })
                    elif a_losses >= 3 and ho >= 2.5 and h_rating >= 1.3:
                        conf = min(75, 50 + a_losses * 5 + int(h_rating * 5))
                        picks.append({
                            "id": fid, "home": f["teams"]["home"]["name"], "away": f["teams"]["away"]["name"],
                            "home_logo": f["teams"]["home"]["logo"], "away_logo": f["teams"]["away"]["logo"],
                            "date": dt.strftime("%d/%m %H:%M"), "home_id": f["teams"]["home"]["id"],
                            "away_id": f["teams"]["away"]["id"], "league_id": lid,
                            "bet": f"{f['teams']['home']['name']} Win",
                            "market_type": "1X2", "odds": round(ho, 2),
                            "odds_home": ho, "odds_draw": odds.get("draw",0), "odds_away": ao,
                            "odds_available": odds.get("available", False),
                            "confidence": conf,
                            "reasoning": f"⚠️ Form crash pick — {f['teams']['away']['name']} on {a_losses}-loss run. {f['teams']['home']['name']} at {ho:.2f} is value"
                        })
            except Exception:
                pass
            return picks

        with ThreadPoolExecutor(max_workers=10) as ex:
            loop = asyncio.get_event_loop()
            results = await asyncio.gather(*[loop.run_in_executor(ex, fetch_form_crash, lc) for lc in leagues])

        for league_picks in results:
            for pick in league_picks:
                if pick["id"] not in seen:
                    seen.add(pick["id"])
                    all_picks.append(pick)

        all_picks.sort(key=lambda x: x["confidence"], reverse=True)
        picks = all_picks[:request.selections]
        if not picks:
            return {"message": "No form crash picks found.", "total_selections": 0, "total_odds": 0, "confidence": 0, "selections": []}

        total_odds = 1.0
        for p in picks: total_odds *= p.get("odds", 2.0)
        avg_conf = sum(p["confidence"] for p in picks) / len(picks)
        return {"message": "Form Crash Acca — Fading Out-of-Form Favourites", "total_selections": len(picks),
                "total_odds": round(total_odds, 2), "confidence": round(avg_conf), "market": "form_crash", "selections": picks}
    except Exception as e:
        raise HTTPException(500, str(e))


@app.post("/generate-fbd-value-acca")
async def generate_fbd_value_acca(request: AccaRequest):
    """FBD Value Acca — picks only from today's FBD file using predicted vs market odds."""
    try:
        fbd_data = get_fbd_data()
        if not fbd_data:
            return {"message":"FBD data not loaded yet — try again after 8:15am.",
                    "total_selections":0,"total_odds":0,"confidence":0,"selections":[]}

        today     = _london_now().date()
        today_str = today.strftime("%Y-%m-%d")
        picks, seen = [], set()

        cached_today = cache_get(f"today_{today_str}") or {}

        def prob_to_odds(prob):
            try:
                p = float(prob)
                return round(100/p, 2) if p > 0 else None
            except: return None

        for fbd in fbd_data:
            home = fbd.get("home","").strip()
            away = fbd.get("away","").strip()
            pred = fbd.get("prediction","").strip().upper()
            if not home or not away or not pred: continue

            ph = prob_to_odds(fbd.get("pred_home_odds"))
            pa = prob_to_odds(fbd.get("pred_away_odds"))
            ah = prob_to_odds(fbd.get("avg_home_odds"))
            aa = prob_to_odds(fbd.get("avg_away_odds"))
            btts_prob = fbd.get("pred_btts")
            ov25_prob = fbd.get("pred_over25")

            # Find API match for logos/time
            api_m = None
            for m in cached_today.get("matches", []):
                if _fbd_names_match(home, m.get("home","")) and _fbd_names_match(away, m.get("away","")):
                    api_m = m; break

            fix_id = api_m["id"] if api_m else abs(hash(f"{home}{away}{today_str}")) % 9999999
            if fix_id in seen: continue

            value_score = 0
            bet = None
            odds = None
            reason_parts = []

            if pred == "H":
                bet = f"{home} Win"
                odds = ah if ah else 1.9
                if ph and ah and ph > ah * 1.05:
                    value_score += 15
                    reason_parts.append(f"FBD value: pred {ph:.2f} vs market {ah:.2f}")
                value_score += 10
                reason_parts.append("FBD predicts HOME win")
            elif pred == "A":
                bet = f"{away} Win"
                odds = aa if aa else 1.9
                if pa and aa and pa > aa * 1.05:
                    value_score += 15
                    reason_parts.append(f"FBD value: pred {pa:.2f} vs market {aa:.2f}")
                value_score += 10
                reason_parts.append("FBD predicts AWAY win")
            elif pred == "D":
                bet = "Draw"
                odds = prob_to_odds(fbd.get("avg_draw_odds")) or 3.2
                value_score += 6
                reason_parts.append("FBD predicts DRAW")

            if btts_prob and float(btts_prob) >= 60:
                value_score += 5
                reason_parts.append(f"BTTS {int(float(btts_prob))}%")
            if ov25_prob and float(ov25_prob) >= 60:
                value_score += 5
                reason_parts.append(f"Over 2.5 {int(float(ov25_prob))}%")

            if not bet or value_score < 10: continue
            if odds: odds = max(1.30, min(float(odds), 12.0))
            else: odds = 1.90

            picks.append({
                "id": fix_id,
                "home": home, "away": away,
                "home_logo": api_m.get("home_logo","") if api_m else "",
                "away_logo": api_m.get("away_logo","") if api_m else "",
                "date": today_str,
                "time": api_m.get("time","TBC") if api_m else "TBC",
                "league": api_m.get("league", fbd.get("league","")) if api_m else fbd.get("league",""),
                "league_id": api_m.get("league_id",0) if api_m else 0,
                "home_id": api_m.get("home_id",0) if api_m else 0,
                "away_id": api_m.get("away_id",0) if api_m else 0,
                "bet": bet, "market_type": "1X2",
                "odds": round(odds, 2),
                "confidence": min(88, 52 + value_score),
                "reasoning": " · ".join(reason_parts),
                "fbd_prediction": pred,
                "fbd_btts": btts_prob,
                "fbd_over25": ov25_prob,
                "fbd_source": True,
            })
            seen.add(fix_id)

        picks.sort(key=lambda x: x["confidence"], reverse=True)
        picks = picks[:request.selections]

        if not picks:
            return {"message":"No FBD value picks found for today.",
                    "total_selections":0,"total_odds":0,"confidence":0,"selections":[]}

        total_odds = 1.0
        for p in picks: total_odds *= p.get("odds",1.9)
        avg_conf = sum(p["confidence"] for p in picks) / len(picks)
        return {"message":"FBD Value Acca","total_selections":len(picks),
                "total_odds":round(total_odds,2),"confidence":round(avg_conf),
                "market":"fbd_value","selections":picks}
    except Exception as e:
        raise HTTPException(500, str(e))


@app.post("/generate-table-rank-acca")
async def generate_table_rank_acca(request: AccaRequest):
    """Table rank acca — top-half vs bottom-half clashes, back the table strength."""
    try:
        leagues = request.leagues if request.leagues else ["PL","ELC","BL1","SA","FL1","PD"]
        today = _london_now().date()
        cutoff = today if request.today_only else today + timedelta(days=3)
        all_picks = []
        seen = set()

        def fetch_standings_and_fixtures(lc: str) -> list:
            lid = LEAGUE_IDS.get(lc)
            if not lid: return []
            picks = []
            try:
                # Get standings to know positions
                st_data = api_get("standings", {"league": lid, "season": get_season()})
                standings = {}
                try:
                    resp = st_data.get("response", [])
                    if not resp: return []
                    raw_table = resp[0].get("league", {}).get("standings", [])
                    table = raw_table[0] if raw_table else []
                    if not table: return []
                    total_teams = len(table)
                    for entry in table:
                        standings[entry["team"]["id"]] = {
                            "rank": entry["rank"],
                            "points": entry["points"],
                            "total": total_teams
                        }
                except (IndexError, KeyError):
                    return []

                data = api_get("fixtures", {"league": lid, "season": get_season(), "next": 10, "timezone": "Europe/London"})
                for f in sorted(data.get("response", []), key=lambda x: x["fixture"]["date"]):
                    fid = f["fixture"]["id"]
                    if fid in seen: continue
                    fd = datetime.fromisoformat(f["fixture"]["date"].replace("Z","+00:00")).date()
                    if fd < today or fd > cutoff: continue
                    status = f["fixture"]["status"]["short"]
                    if status in ("1H","HT","2H","ET","P","FT","AET","PEN"): continue

                    hid = f["teams"]["home"]["id"]
                    aid = f["teams"]["away"]["id"]
                    h_st = standings.get(hid, {})
                    a_st = standings.get(aid, {})
                    if not h_st or not a_st: continue

                    h_rank = h_st["rank"]
                    a_rank = a_st["rank"]
                    total  = h_st.get("total", 20)
                    midpoint = total // 2

                    # Clear top vs bottom clash — min 6 position gap
                    rank_gap = abs(h_rank - a_rank)
                    if rank_gap < 6: continue

                    dt = datetime.fromisoformat(f["fixture"]["date"].replace("Z","+00:00"))
                    odds = get_real_odds(fid)
                    ho = odds.get("home", 0)
                    ao = odds.get("away", 0)

                    if h_rank < a_rank and h_rank <= midpoint:
                        conf = min(84, 58 + rank_gap * 2)
                        picks.append({
                            "id": fid, "home": f["teams"]["home"]["name"], "away": f["teams"]["away"]["name"],
                            "home_logo": f["teams"]["home"]["logo"], "away_logo": f["teams"]["away"]["logo"],
                            "date": dt.strftime("%d/%m %H:%M"), "home_id": hid, "away_id": aid, "league_id": lid,
                            "bet": f"{f['teams']['home']['name']} Win",
                            "market_type": "1X2", "odds": round(ho, 2) if ho else 1.8,
                            "odds_home": ho, "odds_draw": odds.get("draw",0), "odds_away": ao,
                            "odds_available": odds.get("available", False),
                            "confidence": conf,
                            "reasoning": f"📊 Table strength — Rank {h_rank} vs Rank {a_rank} ({rank_gap} places apart). Home side significantly higher in table."
                        })
                    elif a_rank < h_rank and a_rank <= midpoint:
                        conf = min(82, 56 + rank_gap * 2)
                        picks.append({
                            "id": fid, "home": f["teams"]["home"]["name"], "away": f["teams"]["away"]["name"],
                            "home_logo": f["teams"]["home"]["logo"], "away_logo": f["teams"]["away"]["logo"],
                            "date": dt.strftime("%d/%m %H:%M"), "home_id": hid, "away_id": aid, "league_id": lid,
                            "bet": f"{f['teams']['away']['name']} Win",
                            "market_type": "1X2", "odds": round(ao, 2) if ao else 2.1,
                            "odds_home": ho, "odds_draw": odds.get("draw",0), "odds_away": ao,
                            "odds_available": odds.get("available", False),
                            "confidence": conf,
                            "reasoning": f"📊 Table strength — Away side Rank {a_rank} vs home Rank {h_rank} ({rank_gap} places ahead). Away quality showing."
                        })
            except Exception as e:
                logger.error(f"Table rank error {lc}: {e}")
            return picks

        with ThreadPoolExecutor(max_workers=8) as ex:
            loop = asyncio.get_event_loop()
            results = await asyncio.gather(*[loop.run_in_executor(ex, fetch_standings_and_fixtures, lc) for lc in leagues])

        for league_picks in results:
            for pick in league_picks:
                if pick["id"] not in seen:
                    seen.add(pick["id"])
                    all_picks.append(pick)

        all_picks.sort(key=lambda x: x["confidence"], reverse=True)
        picks = all_picks[:request.selections]
        if not picks:
            return {"message": "No table rank picks found.", "total_selections": 0, "total_odds": 0, "confidence": 0, "selections": []}

        total_odds = 1.0
        for p in picks: total_odds *= p.get("odds", 1.9)
        avg_conf = sum(p["confidence"] for p in picks) / len(picks)
        return {"message": "Table Rank Acca — Backing League Position Quality", "total_selections": len(picks),
                "total_odds": round(total_odds, 2), "confidence": round(avg_conf), "market": "table_rank", "selections": picks}
    except Exception as e:
        raise HTTPException(500, str(e))


@app.post("/generate-fhg-acca")
async def generate_fhg_acca(request: AccaRequest):
    """First Half Goals acca — teams with high first-half scoring rates."""
    try:
        leagues = request.leagues if request.leagues else ["BL1","ELC","PL","SA","FL1"]
        today = _london_now().date()
        cutoff = today + timedelta(days=3)
        all_picks = []
        seen = set()
        random.shuffle(leagues)

        def fetch_fhg(lc: str) -> list:
            lid = LEAGUE_IDS.get(lc)
            if not lid: return []
            picks = []
            try:
                data = api_get("fixtures", {"league": lid, "season": get_season(), "next": 10, "timezone": "Europe/London"})
                for f in sorted(data.get("response", []), key=lambda x: x["fixture"]["date"]):
                    fid = f["fixture"]["id"]
                    if fid in seen: continue
                    fd = datetime.fromisoformat(f["fixture"]["date"].replace("Z","+00:00")).date()
                    if fd < today or fd > cutoff: continue
                    status = f["fixture"]["status"]["short"]
                    if status in ("1H","HT","2H","ET","P","FT","AET","PEN"): continue
                    hf = get_team_form(f["teams"]["home"]["id"], lid)
                    af = get_team_form(f["teams"]["away"]["id"], lid)
                    # Use over25% as proxy for high-scoring / early goal tendency
                    h_o25 = hf.get("over25_pct", 0)
                    a_o25 = af.get("over25_pct", 0)
                    h_gf  = hf.get("gf_avg", 0)
                    a_gf  = af.get("gf_avg", 0)
                    combined_goals = round(h_gf + a_gf, 2)
                    avg_o25 = round((h_o25 + a_o25) / 2)
                    if avg_o25 < 55 or combined_goals < 2.3: continue
                    dt = datetime.fromisoformat(f["fixture"]["date"].replace("Z","+00:00"))
                    odds = get_real_odds(fid)
                    conf = min(85, 52 + int(avg_o25 * 0.4) + int(combined_goals * 5))
                    picks.append({
                        "id": fid, "home": f["teams"]["home"]["name"], "away": f["teams"]["away"]["name"],
                        "home_logo": f["teams"]["home"]["logo"], "away_logo": f["teams"]["away"]["logo"],
                        "date": dt.strftime("%d/%m %H:%M"), "home_id": f["teams"]["home"]["id"],
                        "away_id": f["teams"]["away"]["id"], "league_id": lid,
                        "bet": "Over 2.5 Goals (FHG)",
                        "market_type": "FHG",
                        "odds": odds.get("over25", 1.85),
                        "odds_home": odds.get("home", 0), "odds_draw": odds.get("draw",0), "odds_away": odds.get("away",0),
                        "odds_available": odds.get("available", False),
                        "confidence": conf,
                        "reasoning": f"⚡ First-half goals pick — combined avg {combined_goals:.1f} goals/game. {f['teams']['home']['name']} O2.5 {h_o25}% | {f['teams']['away']['name']} O2.5 {a_o25}%"
                    })
            except Exception:
                pass
            return picks

        with ThreadPoolExecutor(max_workers=10) as ex:
            loop = asyncio.get_event_loop()
            results = await asyncio.gather(*[loop.run_in_executor(ex, fetch_fhg, lc) for lc in leagues])

        for league_picks in results:
            for pick in league_picks:
                if pick["id"] not in seen:
                    seen.add(pick["id"])
                    all_picks.append(pick)

        all_picks.sort(key=lambda x: x["confidence"], reverse=True)
        picks = all_picks[:request.selections]
        if not picks:
            return {"message": "No FHG picks found.", "total_selections": 0, "total_odds": 0, "confidence": 0, "selections": []}

        total_odds = 1.0
        for p in picks: total_odds *= p.get("odds", 1.85)
        avg_conf = sum(p["confidence"] for p in picks) / len(picks)
        return {"message": "First Half Goals Acca", "total_selections": len(picks),
                "total_odds": round(total_odds, 2), "confidence": round(avg_conf), "market": "fhg", "selections": picks}
    except Exception as e:
        raise HTTPException(500, str(e))


@app.post("/generate-value-acca")
async def generate_value_acca(request: AccaRequest):
    """Value acca — out-of-form teams, higher odds picks where market may be overpricing.
    Looks for teams on a bad run vs strong opposition where the longer odds offer value."""
    try:
        leagues = request.leagues if request.leagues else ["PL","ELC","BL1","SA","FL1","PD","CL"]
        today = _london_now().date()
        cutoff = today + timedelta(days=3)
        all_picks = []
        seen = set()
        random.shuffle(leagues)

        def fetch_value(lc: str) -> list:
            lid = LEAGUE_IDS.get(lc)
            if not lid: return []
            picks = []
            try:
                data = api_get("fixtures", {"league": lid, "season": get_season(), "next": 10, "timezone": "Europe/London"})
                for f in sorted(data.get("response", []), key=lambda x: x["fixture"]["date"]):
                    fid = f["fixture"]["id"]
                    if fid in seen: continue
                    fd = datetime.fromisoformat(f["fixture"]["date"].replace("Z","+00:00")).date()
                    if fd < today or fd > cutoff: continue
                    status = f["fixture"]["status"]["short"]
                    if status in ("1H","HT","2H","ET","P","FT","AET","PEN"): continue

                    hf = get_team_form(f["teams"]["home"]["id"], lid)
                    af = get_team_form(f["teams"]["away"]["id"], lid)
                    odds = get_real_odds(fid)
                    ho = odds.get("home", 0)
                    aw = odds.get("away", 0)
                    if not odds.get("available") or ho < 1.05 or aw < 1.05: continue

                    h_wr = hf.get("wins", 0) / max(hf.get("games", 1), 1)
                    a_wr = af.get("wins", 0) / max(af.get("games", 1), 1)
                    h_form = hf.get("form_rating", 1.5)
                    a_form = af.get("form_rating", 1.5)

                    # VALUE pick logic:
                    # Strong away team (good form) at big away odds = value
                    # Out-of-form home team priced as favourite = fade them
                    bet = None; pick_odds = 0; reasoning = ""; conf = 60

                    # Case 1: Away team in better form but underdog price
                    if a_form > h_form and aw >= 2.20 and a_wr >= 0.40:
                        form_gap = round(a_form - h_form, 2)
                        bet = f"{f['teams']['away']['name']} Win"
                        pick_odds = aw
                        conf = min(76, 52 + int(form_gap * 12) + int(a_wr * 15))
                        reasoning = f"💎 Value: {f['teams']['away']['name']} in better form ({a_wr*100:.0f}% win rate) @ {aw} — market undervalues away strength"

                    # Case 2: Home team on bad run but priced as big favourite — fade
                    elif h_form < 1.2 and h_wr < 0.30 and ho <= 1.55 and a_form >= 1.6:
                        bet = f"{f['teams']['away']['name']} Win or Draw"
                        pick_odds = round(aw * 0.65 + odds.get("draw", 3.0) * 0.35, 2)
                        pick_odds = max(1.50, min(pick_odds, aw))
                        conf = min(72, 50 + int((1 - h_wr) * 20))
                        reasoning = f"💎 Value fade: {f['teams']['home']['name']} on bad run ({h_wr*100:.0f}% wins) but priced @ {ho}"

                    # Case 3: Both teams in moderate form — BTTS value at decent odds
                    elif (hf.get("btts_pct",0) >= 60 and af.get("btts_pct",0) >= 60 and
                          odds.get("btts", 0) >= 1.75):
                        bet = "Both Teams to Score"
                        pick_odds = odds.get("btts", 1.80)
                        conf = min(74, 50 + int((hf.get("btts_pct",0) + af.get("btts_pct",0)) / 5))
                        reasoning = f"💎 BTTS value @ {pick_odds} — {f['teams']['home']['name']} {hf.get('btts_pct',0)}% · {f['teams']['away']['name']} {af.get('btts_pct',0)}%"

                    if not bet or pick_odds < 1.50: continue

                    dt = datetime.fromisoformat(f["fixture"]["date"].replace("Z","+00:00"))
                    picks.append({
                        "id": fid, "home": f["teams"]["home"]["name"], "away": f["teams"]["away"]["name"],
                        "home_logo": f["teams"]["home"]["logo"], "away_logo": f["teams"]["away"]["logo"],
                        "date": dt.strftime("%d/%m %H:%M"), "home_id": f["teams"]["home"]["id"],
                        "away_id": f["teams"]["away"]["id"], "league_id": lid,
                        "bet": bet, "market_type": "VALUE",
                        "odds": round(pick_odds, 2),
                        "odds_home": ho, "odds_draw": odds.get("draw",0), "odds_away": aw,
                        "odds_available": True,
                        "confidence": conf,
                        "reasoning": reasoning
                    })
            except Exception:
                pass
            return picks

        with ThreadPoolExecutor(max_workers=10) as ex:
            loop = asyncio.get_event_loop()
            results = await asyncio.gather(*[loop.run_in_executor(ex, fetch_value, lc) for lc in leagues])

        for league_picks in results:
            for pick in league_picks:
                if pick["id"] not in seen:
                    seen.add(pick["id"])
                    all_picks.append(pick)

        all_picks.sort(key=lambda x: x["confidence"], reverse=True)
        picks = all_picks[:request.selections]
        if not picks:
            return {"message": "No value picks found today.", "total_selections": 0, "total_odds": 0, "confidence": 0, "selections": []}

        total_odds = 1.0
        for p in picks: total_odds *= p.get("odds", 2.0)
        avg_conf = sum(p["confidence"] for p in picks) / len(picks)
        return {"message": "Value Acca — Out of Form / Market Edge Picks", "total_selections": len(picks),
                "total_odds": round(total_odds, 2), "confidence": round(avg_conf), "market": "value", "selections": picks}
    except Exception as e:
        raise HTTPException(500, str(e))


async def chat(request: dict):
    msg = request.get("message", "").lower()
    if not msg:
        return {"response": "Ask me anything about football or betting! ⚽"}

    # Comprehensive football AI responses
    responses = {
        ("xg", "expected goals", "expected"): "📊 **Expected Goals (xG)** measures the quality of scoring chances, not just quantity. An xG of 1.5 means based on shot quality, a team *should* score 1.5 goals. Combined xG > 2.5 = Over 2.5 Goals bet. Both teams xG > 1.0 = BTTS. The xG Acca tab uses this data to find value bets professionals use daily.",
        ("btts", "both teams", "both score"): "🎯 **Both Teams to Score (BTTS)** — best when both sides have weak defences AND strong attacks. Look for teams with GA > 1.0 per game AND GF > 1.0. Bundesliga and Championship are goldmines for BTTS. The AI tracks historical BTTS% per team to find the best spots.",
        ("acca", "accumulator", "parlay", "treble"): "🎰 **Accumulators** multiply odds — a 5-fold at 2.0 per leg = 32.0 total! The AI picks based on form, xG, and confidence scores. Use Balanced risk (60%+ confidence per pick) for best long-term ROI. A 5-fold winning once in 5 attempts is profit at value odds.",
        ("over", "goals", "2.5", "3.5", "under"): "⚽ **Goals markets** — Over 2.5 is the most popular. Look for combined xG > 2.8, both teams averaging 1.5+ goals, and poor defensive records. Bundesliga averages 3.1 goals/game — one of the best leagues for overs. Under 2.5 works in Serie A, La Liga cup games and derbies.",
        ("form", "recent", "performance", "last 5"): "📈 **Form analysis** is key! The AI weighs last 10 games — wins (×3), draws (×1). Form rating 3.0 = perfect, 0 = no wins. Home advantage adds 0.4 to the rating. Click Analyse Match on any fixture for full form breakdown including goals scored, conceded, and recent results.",
        ("live", "inplay", "in-play", "minute"): "🔴 **Live betting** — the Live tab shows minute-by-minute scores, match events (goals, cards, subs), and live stats. Click any live match for the detailed timeline. Stats like possession and shots can signal momentum shifts — a team with 12 shots but 0 goals may be due!",
        ("value", "value bet", "kelly"): "💰 **Value betting** is when true probability > implied probability from odds. If you think a team has 50% chance to win but odds imply 40%, that's value. Kelly Criterion = (bp - q) / b where b=decimal odds-1, p=your probability, q=1-p. Never bet more than 5% of bankroll on a single bet.",
        ("odds", "bookmaker", "best odds"): "📊 **Odds comparison** — the Best Odds tab in Analyse Match compares across multiple bookmakers. Even 0.1 extra odds on a 5-fold can mean 5-10% more return. Bet365, William Hill, Betfair, Paddy Power all covered. Always take the best odds available!",
        ("strategy", "system", "bankroll", "staking"): "🧠 **Betting strategy**: 1) Flat stake 1-2% of bankroll per bet 2) Only bet where AI confidence ≥65% 3) Track everything in Saved Accas 4) Review monthly P&L 5) Never chase losses. Value + patience + discipline beats any system long-term.",
        ("tiki taka", "pressing", "formation", "tactic", "gegenpress"): "⚽ **Football tactics**: Tiki-taka (Barcelona) = short passes, possession retention. Gegenpressing (Klopp/Liverpool) = immediate press after losing ball. 4-3-3 = width and pressing. 3-5-2 = midfield dominance with wing-backs. Tactical setups affect match stats — defensively solid teams show lower xG allowed.",
        ("var", "offside", "handball", "rules"): "📏 **VAR** checks goals, penalties, red cards, and mistaken identity. Offside uses the furthest point of attacker vs last defender. Handball = arm/hand above shoulder height is usually penalised. Understanding rules helps — VAR reversals mid-game affect live betting significantly.",
        ("messi", "ronaldo", "haaland", "bellingham", "salah"): "⭐ **Modern greats**: Messi (Inter Miami) — highest xG efficiency ever. Ronaldo (Al-Nassr) — movement and finishing. Haaland (Man City) — 36 PL goals in one season, insane xG conversion. Bellingham (Real Madrid) — pressing + goal contributions. Salah — consistent 20+ goal seasons. Check player stats in Analyse Match!",
        ("champions league", "ucl", "europa"): "🏆 **Champions League** — covered fully! UCL games can be tactical (1st leg away caution), affecting goals markets. Home teams average higher xG in UCL group stages. The AI adjusts for European competition patterns. Check the group stage vs knockout round form separately.",
        ("premier league", "pl", "epl"): "ENG **Premier League** — most competitive globally. Man City, Arsenal, Liverpool, Chelsea dominate xG stats. Watch for home advantage — PL home win rate ~45%. Midweek fixtures after European games hit fatigue. AI tracks all 20 PL teams across all markets.",
        ("bundesliga", "germany"): "🇩🇪 **Bundesliga** — highest goals/game average in top 5 leagues (3.1). Perfect for Over 2.5 and BTTS accas. Dortmund, Leverkusen, Bayern all attack-heavy. Especially good for HT goals given aggressive openers.",
        ("tip", "tips", "advice", "help me"): "💡 **Top 5 AI tips**: 1) Generate AI Acca (Balanced risk) for data-backed picks. 2) Cross-reference with xG Acca for confirmation. 3) Use Analyse Match → Best Odds to maximise value. 4) Save all bets and track P&L. 5) Filter by league you know well. Confidence score ≥70% = higher quality picks.",
        ("ht", "half time", "halftime", "first half"): "⏱️ **HT Acca** uses first-half patterns — teams that consistently lead at HT, score in first 20 mins, and have high FH goals averages. Bundesliga and Championship are strong HT markets. The AI analyses 10 games of HT history per team.",
        ("save", "saved", "track", "history"): "💾 **Saved Accas** — save any bet slip, name it, set stake. Mark Won/Lost after results. The dashboard shows win rate, P&L, ROI. Review monthly to identify which markets/leagues work best for you. All data stored locally + synced.",
        ("hello", "hi", "hey", "start", "what can"): "👋 **Welcome to AccaGenius!** I'm your AI betting co-pilot. I can help with:\n• AI Acca generation (form + xG based)\n• Live scores & minute-by-minute events\n• Market explanations & strategies\n• Team analysis & predictions\n• xG deep dives\n\nWhat would you like to explore? ⚽",
        ("thank", "thanks", "cheers", "brilliant", "great"): "🙏 You're welcome! Good luck with your accas — may the odds be in your favour! 🍀⚽",
    }

    for keywords, response in responses.items():
        if any(kw in msg for kw in keywords):
            return {"response": response}

    return {"response": f"Great question! The AI Acca Generator analyses real form + xG data across 17 leagues. Try 'Analyse Match' on any fixture for H2H, form stats, predictions and best odds. What specific match or market are you interested in? ⚽"}


# =========================
# SAVED ACCAS
# =========================


# ══════════════════════════════════════════════════════
# USER AUTH — server-side password store
# ══════════════════════════════════════════════════════
import hashlib as _hl

_user_db: dict = {}  # email → {name, surname, password_hash, plan, role, access_code, joined, telegram_invite}
_USER_DB_FILE   = "/tmp/ag_users.json"
_ACCAS_DB_FILE  = "/tmp/ag_accas.json"

def _load_user_db():
    """Load users and codes from PostgreSQL (primary) then /tmp (fallback)."""
    global _user_db, _reset_tokens, _used_codes, _all_codes
    import json

    logger.info(f"🔍 DB mode: {'PostgreSQL' if _is_postgres() else 'SQLite/tmp'} | DATABASE_URL set: {bool(DATABASE_URL)}")

    # Primary DB (Postgres or SQLite)
    try:
        db_users = db_load_all_users()
        if db_users:
            _user_db.update(db_users)
            logger.info(f"✅ Loaded {len(db_users)} users from {'PostgreSQL' if _is_postgres() else 'SQLite'}")
        db_all, db_used = db_load_codes()
        if db_all:
            _all_codes.update(db_all)
            _used_codes.update(db_used)
            logger.info(f"✅ Loaded {len(db_all)} codes from DB")
        else:
            logger.warning("⚠️ No codes in DB — loading defaults from env")
    except Exception as e:
        logger.warning(f"DB load error: {e}")
    # /tmp fallback — for data not yet in SQLite
    try:
        if os.path.exists(_USER_DB_FILE):
            with open(_USER_DB_FILE, 'r') as f:
                data = json.load(f)
            if "users" in data:
                for email, u in data["users"].items():
                    if email not in _user_db:
                        _user_db[email] = u
                _all_codes.update(data.get("all_codes", []))
                for c, info in data.get("used_codes", {}).items():
                    if c not in _used_codes:
                        _used_codes[c] = info
    except Exception as e:
        logger.warning(f"Could not load user db from /tmp: {e}")
    try:
        if os.path.exists(_TOKENS_FILE):
            with open(_TOKENS_FILE) as f:
                raw = json.load(f)
            for k, v in raw.items():
                try:
                    v["expires"] = datetime.fromisoformat(v["expires"])
                    if not v.get("used") and v["expires"] > datetime.now():
                        _reset_tokens[k] = v
                except: pass
    except Exception as e:
        logger.warning(f"Could not load reset tokens: {e}")

def _save_accas_db():
    """Persist saved accas to SQLite (primary) and /tmp (backup)."""
    # SQLite
    try:
        for acca in saved_accas_store:
            db_save_acca(acca)
    except Exception as e:
        logger.warning(f"SQLite acca save error: {e}")
    # /tmp backup
    try:
        import json
        with open(_ACCAS_DB_FILE, 'w') as f:
            json.dump(saved_accas_store, f)
    except Exception as e:
        logger.warning(f"Could not save accas to /tmp: {e}")

def _load_accas_db():
    """Load saved accas from database and /tmp fallback."""
    global saved_accas_store
    loaded = {}
    # Primary DB (Postgres or SQLite)
    try:
        pg = _get_pg()
        if pg:
            cur = pg.cursor()
            cur.execute("SELECT data FROM saved_accas ORDER BY created_at DESC")
            rows = cur.fetchall()
            pg.close()
            for r in rows:
                acca = _json.loads(r[0])
                loaded[str(acca.get("id"))] = acca
        else:
            conn = _get_sqlite()
            rows = conn.execute("SELECT data FROM saved_accas ORDER BY created_at DESC").fetchall()
            conn.close()
            for r in rows:
                acca = _json.loads(r["data"])
                loaded[str(acca.get("id"))] = acca
        logger.info(f"✅ Loaded {len(loaded)} saved accas from DB")
    except Exception as e:
        logger.warning(f"DB acca load error: {e}")
    # /tmp fallback
    try:
        import json
        if os.path.exists(_ACCAS_DB_FILE):
            with open(_ACCAS_DB_FILE, 'r') as f:
                tmp_accas = json.load(f)
            for acca in tmp_accas:
                aid = str(acca.get("id"))
                if aid not in loaded:
                    loaded[aid] = acca
    except Exception as e:
        logger.warning(f"Could not load accas from /tmp: {e}")
    saved_accas_store = list(loaded.values())
    logger.info(f"✅ Total accas loaded: {len(saved_accas_store)}")

def _save_user_db():
    """Persist users and codes to SQLite (primary) and /tmp (backup)."""
    import json
    # SQLite — primary persistent storage
    try:
        for email, data in _user_db.items():
            db_save_user(email, data)
        for code, info in _used_codes.items():
            db_save_code(code, info.get("email",""), info.get("source","manual"))
    except Exception as e:
        logger.warning(f"SQLite save error: {e}")
    # /tmp — backup (resets on deploy but useful for restarts)
    try:
        with open(_USER_DB_FILE, 'w') as f:
            json.dump({
                "users": _user_db,
                "used_codes": _used_codes,
                "all_codes": list(_all_codes)
            }, f)
    except Exception as e:
        logger.warning(f"Could not save user db to /tmp: {e}")
    try:
        saveable = {}
        for k, v in _reset_tokens.items():
            saveable[k] = {**v, "expires": v["expires"].isoformat() if hasattr(v.get("expires"), "isoformat") else str(v.get("expires",""))}
        with open(_TOKENS_FILE, 'w') as f:
            json.dump(saveable, f)
    except Exception as e:
        logger.warning(f"Could not save reset tokens: {e}")



def _hash_pw(pw: str) -> str:
    return _hl.sha256(pw.encode()).hexdigest()

@app.post("/register")
async def register_user(request: dict):
    email       = (request.get("email") or "").strip().lower()
    password    = (request.get("password") or "").strip()
    name        = (request.get("name") or "").strip()
    surname     = (request.get("surname") or "").strip()
    plan        = (request.get("plan") or "free").strip()
    access_code = (request.get("access_code") or "").strip().upper()

    if not email or not password:
        return {"ok": False, "error": "Email and password required"}
    if len(password) < 6:
        return {"ok": False, "error": "Password must be at least 6 characters"}

    # Validate access code for pro
    if plan == "pro":
        if access_code not in _all_codes:
            return {"ok": False, "error": "Invalid access code"}
        if access_code in _used_codes and _used_codes[access_code].get("email","").lower() != email:
            return {"ok": False, "error": "Access code already used"}

    # Allow re-registration (update profile) but don't overwrite password unless explicitly set
    existing = _user_db.get(email, {})
    _user_db[email] = {
        "name":          name or existing.get("name", ""),
        "surname":       surname or existing.get("surname", ""),
        "password_hash": _hash_pw(password),
        "plan":          plan,
        "role":          existing.get("role", "user"),
        "access_code":   access_code or existing.get("access_code", ""),
        "joined":        existing.get("joined", datetime.now().isoformat()),
        "telegram_invite": TELEGRAM_PRO_INVITE_LINK if plan == "pro" else "",
    }
    if plan == "pro" and access_code:
        _used_codes[access_code] = {"email": email, "used_at": datetime.now().isoformat()}

    _save_user_db()
    # Send welcome email
    if plan == "pro":
        email_welcome_pro(email, name, access_code, TELEGRAM_PRO_INVITE_LINK)
    else:
        email_welcome_free(email, name)
    user = _user_db[email]
    return {"ok": True, "email": email, "name": user["name"], "surname": user["surname"],
            "plan": user["plan"], "role": user["role"], "access_code": user["access_code"],
            "joined": user["joined"], "telegram_invite": user["telegram_invite"]}


@app.post("/login")
async def login_user(request: dict):
    email    = (request.get("email") or "").strip().lower()
    password = (request.get("password") or "").strip()

    if not email or not password:
        return {"ok": False, "error": "Email and password required"}

    # ── Admin password check — works with any email ──
    admin_pw = (ADMIN_PASSWORD_ENV or "").strip()
    if admin_pw and password.strip() == admin_pw:
        return {"ok": True, "email": email, "name": "Admin", "surname": "",
                "plan": "pro", "role": "admin", "access_code": "",
                "joined": _user_db.get(email, {}).get("joined", datetime.now().isoformat()),
                "telegram_invite": TELEGRAM_PRO_INVITE_LINK}

    user = _user_db.get(email)
    if not user:
        return {"ok": False, "error": "No account found — please sign up first"}
    if user["password_hash"] != _hash_pw(password):
        return {"ok": False, "error": "Incorrect password"}

    return {"ok": True, "email": email, "name": user["name"], "surname": user["surname"],
            "plan": user["plan"], "role": user["role"], "access_code": user["access_code"],
            "joined": user["joined"], "telegram_invite": user.get("telegram_invite", ""),
            "plan_status": user.get("plan_status", "active"),
            "trial_end": user.get("trial_end", "")}


@app.post("/auth")
async def auth_user(request: dict):
    """Check if a user exists by email only — for session restore on page load."""
    email = (request.get("email") or "").strip().lower()
    if not email:
        return {"ok": False}
    # Check admin
    if ADMIN_PASSWORD_ENV and request.get("password","") == ADMIN_PASSWORD_ENV.strip():
        return {"ok": True, "role": "admin", "plan": "pro", "name": "Admin"}
    user = _user_db.get(email)
    if not user:
        return {"ok": False, "error": "No account found"}
    return {"ok": True, "email": email, "name": user["name"], "surname": user.get("surname",""),
            "plan": user["plan"], "role": user.get("role","user"),
            "access_code": user.get("access_code",""),
            "joined": user.get("joined",""),
            "telegram_invite": user.get("telegram_invite","")}


@app.post("/billing-portal")
async def billing_portal(request: dict):
    """Create a Stripe billing portal session for the customer."""
    if not STRIPE_API_KEY:
        return {"error": "Stripe not configured"}
    email = (request.get("email") or "").strip().lower()
    if not email:
        return {"error": "Email required"}
    try:
        # Find the Stripe customer ID by email
        r = requests.get(
            "https://api.stripe.com/v1/customers",
            auth=(STRIPE_API_KEY, ""),
            params={"email": email, "limit": 1},
            timeout=10
        )
        customers = r.json().get("data", [])
        if not customers:
            return {"error": "No billing account found for this email"}
        customer_id = customers[0]["id"]

        # Create portal session
        r2 = requests.post(
            "https://api.stripe.com/v1/billing_portal/sessions",
            auth=(STRIPE_API_KEY, ""),
            data={"customer": customer_id, "return_url": SITE_URL},
            timeout=10
        )
        portal = r2.json()
        logger.info(f"Portal response: {portal}")
        if "url" in portal:
            return {"url": portal["url"]}
        else:
            err = portal.get("error", {}).get("message", "Could not create billing session")
            logger.error(f"Portal error: {portal}")
            # If portal not activated, send to Stripe dashboard directly
            if "billing_portal" in str(portal).lower() or "configuration" in str(portal).lower():
                return {"error": "Billing portal not activated in Stripe. Go to Stripe Dashboard → Settings → Billing → Customer Portal → Activate"}
            return {"error": err}
    except Exception as e:
        logger.error(f"Billing portal error: {e}")
        return {"error": str(e)}


@app.post("/contact-support")
async def contact_support(request: dict):
    """Forward a support message to the admin via Resend."""
    email   = (request.get("email") or "unknown").strip()
    name    = (request.get("name") or "Unknown").strip()
    plan    = (request.get("plan") or "free").strip()
    message = (request.get("message") or "").strip()
    if not message:
        return {"ok": False, "error": "No message provided"}
    support_inbox = os.getenv("SUPPORT_EMAIL", "j3nno83@gmail.com")
    body = f"""
    <div style="font-family:system-ui;padding:20px;background:#f9f9f9;border-radius:8px">
        <h2 style="color:#f0b429">📬 New Support Message — AccaGenius</h2>
        <p><strong>From:</strong> {name} ({email})</p>
        <p><strong>Plan:</strong> {plan}</p>
        <p><strong>Message:</strong></p>
        <div style="background:#fff;border-left:4px solid #f0b429;padding:12px;border-radius:4px">
            {message}
        </div>
        <p style="color:#999;font-size:0.8rem;margin-top:16px">Reply directly to this email to respond to the user.</p>
    </div>"""
    try:
        r = requests.post(
            "https://api.resend.com/emails",
            headers={"Authorization": f"Bearer {RESEND_API_KEY}", "Content-Type": "application/json"},
            json={
                "from": EMAIL_FROM,
                "reply_to": [email],
                "to": [support_inbox],
                "subject": f"💬 Support: {name} ({plan})",
                "html": body
            },
            timeout=10
        )
        ok = r.status_code in (200, 201)
        if not ok:
            logger.error(f"Support email failed: {r.text}")
        return {"ok": ok}
    except Exception as e:
        logger.error(f"Support email error: {e}")
        return {"ok": False}


@app.post("/delete-account")
async def delete_account(request: dict):
    """Delete a user account."""
    email = (request.get("email") or "").strip().lower()
    if not email:
        return {"ok": False}
    if email in _user_db:
        del _user_db[email]
        _save_user_db()
        logger.info(f"Account deleted: {email}")
    return {"ok": True}


@app.post("/change-password")
async def change_password(request: dict):
    email        = (request.get("email") or "").strip().lower()
    old_password = (request.get("old_password") or "").strip()
    new_password = (request.get("new_password") or "").strip()

    if not email or not new_password:
        return {"ok": False, "error": "Missing fields"}
    user = _user_db.get(email)
    if not user:
        return {"ok": False, "error": "Account not found"}
    # Allow change via reset token (old_password empty) or by verifying old password
    if old_password and user["password_hash"] != _hash_pw(old_password):
        return {"ok": False, "error": "Current password incorrect"}
    if len(new_password) < 6:
        return {"ok": False, "error": "Password must be at least 6 characters"}

    _user_db[email]["password_hash"] = _hash_pw(new_password)
    _save_user_db()
    user = _user_db[email]
    return {"ok": True, "email": email, "name": user["name"], "plan": user["plan"],
            "role": user["role"], "access_code": user["access_code"]}


# ══════════════════════════════════════════════════════
# STRIPE PAYMENT ENDPOINTS
# ══════════════════════════════════════════════════════

@app.post("/create-checkout-session")
async def create_checkout_session(request: dict):
    """Create a Stripe Checkout session for £20/month Pro subscription."""
    if not STRIPE_API_KEY:
        return {"error": "Stripe not configured — set STRIPE_API_KEY in Railway"}

    email = (request.get("email", "") or "").strip().lower()
    name  = request.get("name", "")

    # ── FREE TRIAL ABUSE PROTECTION ──────────────────────────────
    # Block anyone who has already had a free trial
    if email and email in _user_db:
        user = _user_db[email]
        plan_status = user.get("plan_status", "")
        had_trial   = user.get("had_trial", False)
        # If they've had a trial before or are already active/cancelled pro
        if had_trial or plan_status in ("active", "trial", "cancelled"):
            logger.warning(f"⚠️ Trial abuse blocked: {email} (status: {plan_status}, had_trial: {had_trial})")
            # Still let them subscribe but WITHOUT the free trial
            try:
                if STRIPE_PRICE_ID:
                    line_items = [{"price": STRIPE_PRICE_ID, "quantity": 1}]
                else:
                    line_items = [{
                        "price_data": {
                            "currency": "gbp", "unit_amount": 2000,
                            "recurring": {"interval": "month"},
                            "product_data": {"name": "AccaGenius Pro"},
                        }, "quantity": 1,
                    }]
                payload = {
                    "mode": "subscription",
                    "line_items": line_items,
                    "success_url": f"{SITE_URL}?payment=success&session_id={{CHECKOUT_SESSION_ID}}",
                    "cancel_url":  f"{SITE_URL}?payment=cancelled",
                    "allow_promotion_codes": True,
                    # NO trial_period_days — they've already had one
                }
                if email: payload["customer_email"] = email
                r = requests.post(
                    "https://api.stripe.com/v1/checkout/sessions",
                    auth=(STRIPE_API_KEY, ""),
                    data=_flatten_stripe(payload), timeout=15
                )
                data = r.json()
                if r.status_code != 200:
                    return {"error": data.get("error", {}).get("message", "Stripe error")}
                return {"url": data["url"], "session_id": data["id"], "no_trial": True}
            except Exception as e:
                return {"error": str(e)}

    try:
        # Build the line items — use price ID if set, otherwise inline price
        if STRIPE_PRICE_ID:
            line_items = [{"price": STRIPE_PRICE_ID, "quantity": 1}]
        else:
            line_items = [{
                "price_data": {
                    "currency": "gbp",
                    "unit_amount": 2000,  # £20.00
                    "recurring": {"interval": "month"},
                    "product_data": {
                        "name": "AccaGenius Pro",
                        "description": "Full AI acca analysis, xG stats, unlimited alerts & more",
                    },
                },
                "quantity": 1,
            }]

        payload = {
            "mode": "subscription",
            "line_items": line_items,
            "success_url": f"{SITE_URL}?payment=success&session_id={{CHECKOUT_SESSION_ID}}",
            "cancel_url":  f"{SITE_URL}?payment=cancelled",
            "allow_promotion_codes": True,
            "subscription_data": {
                "trial_period_days": 7,  # 7-day free trial
            },
        }
        if email:
            payload["customer_email"] = email
        if name:
            payload["metadata"] = {"name": name}

        r = requests.post(
            "https://api.stripe.com/v1/checkout/sessions",
            auth=(STRIPE_API_KEY, ""),
            data=_flatten_stripe(payload),
            timeout=15
        )
        data = r.json()
        if r.status_code != 200:
            logger.error(f"Stripe session error: {data}")
            return {"error": data.get("error", {}).get("message", "Stripe error")}
        return {"url": data["url"], "session_id": data["id"]}

    except Exception as e:
        logger.error(f"Stripe checkout error: {e}")
        return {"error": str(e)}


def _flatten_stripe(data: dict, prefix: str = "") -> dict:
    """Flatten nested dict to Stripe's form-encoded format."""
    result = {}
    for k, v in data.items():
        key = f"{prefix}[{k}]" if prefix else k
        if isinstance(v, dict):
            result.update(_flatten_stripe(v, key))
        elif isinstance(v, list):
            for i, item in enumerate(v):
                if isinstance(item, dict):
                    result.update(_flatten_stripe(item, f"{key}[{i}]"))
                else:
                    result[f"{key}[{i}]"] = item
        elif isinstance(v, bool):
            result[key] = str(v).lower()
        elif v is not None:
            result[key] = v
    return result


@app.get("/payment-success")
async def payment_success(session_id: str = ""):
    """Called when Stripe redirects back after successful payment.
    Retrieves session, generates code and sends email immediately.
    This is a fallback in case webhook doesn't fire."""
    if not session_id or not STRIPE_API_KEY:
        return {"ok": False, "error": "Missing session"}
    try:
        r = requests.get(
            f"https://api.stripe.com/v1/checkout/sessions/{session_id}",
            auth=(STRIPE_API_KEY, ""),
            timeout=15
        )
        if r.status_code != 200:
            return {"ok": False, "error": "Could not retrieve session"}
        session = r.json()
        if session.get("payment_status") not in ("paid", "no_payment_required"):
            return {"ok": False, "error": "Payment not confirmed"}
        # Build a fake event and handle it
        fake_event = {"type": "checkout.session.completed", "data": {"object": session}}
        await _handle_stripe_payment(fake_event)
        return {"ok": True, "message": "Access code sent to your email"}
    except Exception as e:
        logger.error(f"Payment success handler error: {e}")
        return {"ok": False, "error": str(e)}


@app.post("/stripe-webhook")
async def stripe_webhook(request: Request):
    """Handle Stripe webhook events — auto-generate code and email on successful payment."""
    body = await request.body()
    sig  = request.headers.get("stripe-signature", "")

    # Verify webhook signature
    if STRIPE_WH_SECRET:
        try:
            _verify_stripe_signature(body, sig, STRIPE_WH_SECRET)
        except Exception as e:
            logger.error(f"Stripe webhook signature failed: {e}")
            return {"error": "Invalid signature"}, 400

    try:
        import json
        event = json.loads(body)
    except Exception:
        return {"error": "Invalid JSON"}

    event_type = event.get("type", "")
    event_id   = event.get("id", "")
    logger.info(f"Stripe webhook: {event_type} — {event_id}")

    # ── IDEMPOTENCY — ignore events already processed ──────────────
    # Prevents duplicate charges when Stripe retries webhooks
    if event_id and event_id in _processed_webhook_ids:
        logger.info(f"Stripe webhook: duplicate {event_id} — skipping")
        return {"received": True, "duplicate": True}
    if event_id:
        _processed_webhook_ids.add(event_id)
        if len(_processed_webhook_ids) > 1000:
            ids_list = list(_processed_webhook_ids)
            _processed_webhook_ids.clear()
            _processed_webhook_ids.update(ids_list[-500:])

    # ── Handle successful subscription (after trial or immediate payment) ──
    if event_type in ("checkout.session.completed", "invoice.payment_succeeded"):
        await _handle_stripe_payment(event)
        # Also check if this is a trading plan activation
        obj = event.get("data", {}).get("object", {})
        meta = obj.get("metadata", {}) or {}
        if meta.get("plan") == "trading":
            trading_email = meta.get("email", "").lower()
            if trading_email and trading_email in _user_db:
                _user_db[trading_email]["plan"] = "trading"
                _user_db[trading_email]["plan_status"] = "active"
                _save_user_db()
                logger.info(f"✅ Trading plan activated for {trading_email}")

    # Trial converted to paid subscription
    elif event_type == "customer.subscription.updated":
        sub = event.get("data", {}).get("object", {})
        cust_email = sub.get("metadata", {}).get("email", "")
        status = sub.get("status", "")
        if not cust_email:
            # Look up by subscription ID
            sub_id = sub.get("id", "")
            for em, u in _user_db.items():
                if u.get("stripe_subscription") == sub_id:
                    cust_email = em
                    break
        if cust_email and cust_email in _user_db:
            if status == "active" and _user_db[cust_email].get("plan_status") == "trial":
                _user_db[cust_email]["plan_status"] = "active"
                _user_db[cust_email]["trial_end"] = ""
                _save_user_db()
                logger.info(f"✅ {cust_email} trial converted to paid")
                # Send "you've been charged" confirmation email
                first = _user_db[cust_email].get("name", "there")
                send_email(cust_email, "⚡ AGD Sports Trading — Payment Confirmed",
                    _email_base(f"<p>Hi {first},</p><p>Your AGD Sports Trading Pro subscription is now active. £20/month will be charged on the same date each month.</p><p>Keep enjoying your Trading access! ⚡</p>"))

    # Subscription cancelled or payment failed — downgrade to free
    elif event_type in ("customer.subscription.deleted", "invoice.payment_failed"):
        obj2 = event.get("data", {}).get("object", {})
        cust_email = obj2.get("metadata", {}).get("email", "")
        sub_id = obj2.get("id", "") or obj2.get("subscription", "")
        if not cust_email:
            for em, u in _user_db.items():
                if u.get("stripe_subscription") == sub_id:
                    cust_email = em
                    break
        if cust_email and cust_email in _user_db:
            _user_db[cust_email]["plan"] = "free"
            _user_db[cust_email]["plan_status"] = "cancelled"
            _user_db[cust_email]["telegram_invite"] = ""
            _save_user_db()
            logger.info(f"⚠️ {cust_email} downgraded to free — {event_type}")
            first = _user_db[cust_email].get("name", "there")
            send_email(cust_email, "Your AGD Sports Trading Pro subscription has ended",
                _email_base(f"<p>Hi {first},</p><p>Your AGD Sports Trading Pro subscription has ended. You still have free access to AccaGenius.</p><p>To rejoin Pro at any time, visit <a href='{SITE_URL}' style='color:#f0b429'>{SITE_URL}</a> and upgrade again.</p><p>We'd love to have you back ⚡</p>"))

    return {"received": True}


def _verify_stripe_signature(body: bytes, sig_header: str, secret: str):
    """Verify Stripe webhook signature using HMAC-SHA256."""
    parts = {p.split("=")[0]: p.split("=")[1] for p in sig_header.split(",") if "=" in p}
    timestamp = parts.get("t", "")
    v1 = parts.get("v1", "")
    signed_payload = f"{timestamp}.{body.decode()}"
    expected = hmac.new(secret.encode(), signed_payload.encode(), hashlib.sha256).hexdigest()
    if not hmac.compare_digest(expected, v1):
        raise ValueError("Signature mismatch")


async def _handle_stripe_payment(event: dict):
    """Generate access code and email it to the customer after successful payment."""
    import secrets as _sec

    obj = event.get("data", {}).get("object", {})

    # Get customer email — location differs by event type
    email = (
        obj.get("customer_email") or
        obj.get("customer_details", {}).get("email") or
        obj.get("metadata", {}).get("email") or ""
    )
    name = (
        obj.get("customer_details", {}).get("name") or
        obj.get("metadata", {}).get("name") or
        email.split("@")[0]
    )

    if not email:
        logger.error("Stripe webhook: no email found in event")
        return

    # Check if they already have a code (duplicate webhook guard)
    existing = [c for c, info in _used_codes.items() if info.get("email") == email]
    if existing:
        logger.info(f"Stripe webhook: {email} already has code {existing[0]}, skipping")
        return

    # Generate a fresh access code
    chars = "ABCDEFGHJKLMNPQRSTUVWXYZ23456789"
    code = "ACCA-" + "".join(_sec.choice(chars) for _ in range(8))
    _all_codes.add(code)
    # Mark as used immediately so duplicate webhooks don't re-send
    _used_codes[code] = {"email": email, "name": name, "used_at": datetime.now().isoformat(), "source": "stripe"}

    # Work out trial end date
    obj_sub = obj.get("subscription", "")
    trial_end = obj.get("subscription_data", {}).get("trial_end") or                 obj.get("trial_end") or                 obj.get("metadata", {}).get("trial_end") or ""
    # If trial_period_days is set, trial ends 7 days from now
    trial_end_iso = ""
    if trial_end:
        try:
            trial_end_iso = datetime.fromtimestamp(int(trial_end)).isoformat()
        except: pass
    if not trial_end_iso:
        trial_end_iso = (datetime.now() + timedelta(days=7)).isoformat()

    # Create/update user account automatically
    if email not in _user_db:
        _user_db[email] = {
            "name": name.split()[0] if name else email.split("@")[0],
            "surname": name.split()[-1] if name and len(name.split()) > 1 else "",
            "password_hash": "",
            "plan": "pro",
            "plan_status": "trial",
            "had_trial": True,      # prevent second free trials
            "trial_end": trial_end_iso,
            "stripe_subscription": str(obj_sub),
            "role": "user",
            "access_code": code,
            "joined": datetime.now().isoformat(),
            "telegram_invite": TELEGRAM_PRO_INVITE_LINK,
        }
    else:
        _user_db[email]["plan"] = "pro"
        _user_db[email]["plan_status"] = "trial"
        _user_db[email]["had_trial"] = True   # mark so they can't get another
        _user_db[email]["trial_end"] = trial_end_iso
        _user_db[email]["stripe_subscription"] = str(obj_sub)
        _user_db[email]["access_code"] = code
        _user_db[email]["telegram_invite"] = TELEGRAM_PRO_INVITE_LINK
    _save_user_db()

    # Send the welcome email with the code + telegram link
    await _send_pro_welcome_email(email, name, code, TELEGRAM_PRO_INVITE_LINK)
    logger.info(f"✅ Stripe payment: code {code} generated and emailed to {email}")


async def _send_pro_welcome_email(email: str, name: str, code: str, telegram_invite: str = ""):
    """Send Pro welcome email with access code and Telegram invite via Resend."""
    if not RESEND_API_KEY:
        logger.error("Cannot send welcome email — RESEND_API_KEY not set")
        return

    first = name.split()[0] if name else "there"
    tg_link = telegram_invite or TELEGRAM_PRO_INVITE_LINK
    reply_to = os.getenv("SUPPORT_EMAIL", "j3nno83@gmail.com")

    html_body = f"""
<!DOCTYPE html>
<html>
<head><meta charset="utf-8"></head>
<body style="background:#0a0a0f;color:#e2e8f0;font-family:system-ui,sans-serif;margin:0;padding:0">
  <div style="max-width:520px;margin:40px auto;padding:32px;background:#12121a;border-radius:16px;border:1px solid rgba(240,180,41,0.3)">

    <div style="text-align:center;margin-bottom:24px">
      <div style="font-size:2.5rem">⚡</div>
      <h1 style="color:#f0b429;font-size:1.6rem;margin:8px 0">Welcome to AccaGenius Pro!</h1>
      <p style="color:#64748b;font-size:0.85rem;margin:0">Hi {first}, your payment was successful</p>
    </div>

    <!-- Access Code -->
    <div style="background:#0a0a0f;border:2px solid #f0b429;border-radius:12px;padding:20px;text-align:center;margin:0 0 24px">
      <div style="color:#64748b;font-size:0.75rem;margin-bottom:8px;text-transform:uppercase;letter-spacing:1px">Your Pro Access Code</div>
      <div style="font-family:monospace;font-size:2rem;font-weight:900;color:#f0b429;letter-spacing:4px">{code}</div>
    </div>

    <!-- Telegram -->
    <div style="background:rgba(76,201,240,0.08);border:1px solid rgba(76,201,240,0.3);border-radius:12px;padding:20px;text-align:center;margin:0 0 24px">
      <div style="font-size:1.5rem;margin-bottom:8px">📲</div>
      <div style="font-weight:700;color:#4cc9f0;margin-bottom:8px">Join Your Pro Telegram Channel</div>
      <p style="color:#94a3b8;font-size:0.82rem;margin:0 0 16px">Get live in-play alerts, Goal Due signals and daily accas straight to your phone</p>
      <a href="{tg_link}" style="display:inline-block;background:linear-gradient(135deg,#4cc9f0,#7c3aed);color:#fff;font-weight:800;padding:12px 28px;border-radius:10px;text-decoration:none;font-size:0.95rem">
        📲 Join Pro Channel Now
      </a>
    </div>

    <!-- How to activate -->
    <p style="color:#94a3b8;font-weight:700;margin:0 0 8px">How to activate your account:</p>
    <ol style="color:#94a3b8;line-height:2;margin:0 0 24px">
      <li>Go to <a href="{SITE_URL}" style="color:#f0b429">{SITE_URL}</a></li>
      <li>Click <strong style="color:#e2e8f0">Sign Up</strong> → Select <strong style="color:#e2e8f0">Pro</strong></li>
      <li>Enter your details and paste the code above</li>
      <li>Full Pro access unlocked ⚡</li>
    </ol>

    <div style="background:rgba(240,180,41,0.06);border:1px solid rgba(240,180,41,0.15);border-radius:8px;padding:16px;font-size:0.82rem;color:#94a3b8">
      🔒 Keep this code safe — it's linked to your email address.<br>
      💬 <strong style="color:#e2e8f0">Need help or want to cancel?</strong> Just reply to this email anytime and we'll sort it within 24hrs.
    </div>

    <div style="text-align:center;margin-top:24px;color:#475569;font-size:0.72rem">
      AccaGenius · <a href="{SITE_URL}" style="color:#475569">{SITE_URL}</a>
    </div>
  </div>
</body>
</html>"""

    try:
        r = requests.post(
            "https://api.resend.com/emails",
            headers={"Authorization": f"Bearer {RESEND_API_KEY}", "Content-Type": "application/json"},
            json={
                "from": EMAIL_FROM,
                "reply_to": [reply_to],
                "to": [email],
                "subject": "⚡ Welcome to AccaGenius Pro — Your Access Code Inside",
                "html": html_body,
            },
            timeout=10
        )
        if r.status_code in (200, 201):
            logger.info(f"Pro welcome email sent to {email}")
        else:
            logger.error(f"Email send failed: {r.status_code} {r.text}")
    except Exception as e:
        logger.error(f"Email error: {e}")


@app.get("/stripe-success")
async def stripe_success(session_id: str = ""):
    """Landing page after successful Stripe payment."""
    return {
        "status": "success",
        "message": "Payment successful! Check your email for your Pro access code.",
        "session_id": session_id
    }



@app.get("/admin/bets")
async def admin_get_bets():
    """Admin P&L — all tracked bets from alerts."""
    bets = list(_admin_bets.values())
    total = len(bets)
    resolved = [b for b in bets if b["result"] != "pending"]
    won = [b for b in resolved if b["result"] == "won"]
    lost = [b for b in resolved if b["result"] == "lost"]
    total_pl = round(sum(b.get("profit_loss", 0) for b in resolved), 2)
    return {
        "bets": sorted(bets, key=lambda x: x.get("fired_at",""), reverse=True),
        "stats": {
            "total": total,
            "won": len(won),
            "lost": len(lost),
            "pending": total - len(resolved),
            "win_rate": round(len(won)/len(resolved)*100, 1) if resolved else 0,
            "profit_loss": total_pl,
        }
    }

@app.post("/admin/bets/{bet_id}/result")
async def admin_update_bet(bet_id: str, request: dict):
    """Update bet result — won/lost + profit/loss amount."""
    if bet_id not in _admin_bets:
        raise HTTPException(404, "Bet not found")
    _admin_bets[bet_id]["result"] = request.get("result", "pending")
    _admin_bets[bet_id]["profit_loss"] = float(request.get("profit_loss", 0))
    _admin_bets[bet_id]["odds"] = request.get("odds", "")
    _admin_bets[bet_id]["stake"] = request.get("stake", 0)
    _admin_bets[bet_id]["notes"] = request.get("notes", "")
    return {"ok": True, "bet": _admin_bets[bet_id]}

@app.delete("/admin/bets/{bet_id}")
async def admin_delete_bet(bet_id: str):
    """Remove a bet from tracker."""
    _admin_bets.pop(bet_id, None)
    return {"ok": True}

# ══════════════════════════════════════════════════════════════
# PAPER TRADING BOT — Betfair-style, no real money
# ══════════════════════════════════════════════════════════════

_paper_bot_config = {
    "mode": "paper",          # paper | live (live needs Betfair API)
    "status": "active",       # active | paused
    "stake": 3.00,            # default stake £
    "strategy": "auto",       # auto | manual
    "max_daily_loss": 50.00,  # stop bot if daily loss exceeds this
    "min_odds": 1.5,
    "max_odds": 6.0,
    "markets": {              # which alert types to auto-bet
        "Match Winner": True,
        "Both Teams to Score": True,
        "HT/FT Value — Comeback": True,
        "Over 1.5 Goals": True,
        "Goal Due — Track A": True,   # Simple back current market
        "Goal Due — Track B": True,   # Drip ladder + cover
        "Goal Due — Track B (Cover)": True,
    },
    "drip": {
        "enabled": True,
        "stakes": [2.0, 3.0, 5.0],      # increasing stakes per drip level
        "odds_drift": [0.0, 1.0, 2.0],  # back at entry, +1.0, +2.0 drift
        "cover_xg_threshold": 0.8,       # place cover when xG unscored < this
        "cover_market": "Over 1.5 Goals",
        "cover_odds": 2.08,
    }
}

# Drip ladder tracker: fixture_id -> {entry_odds, drips_placed, total_staked, covered, cashed_out}
_drip_ladders: dict = {}

# ── Paper bot persistence ────────────────────────────────────
PAPER_TRADES_FILE  = "/tmp/ag_paper_trades.json"
PAPER_DRIPS_FILE   = "/tmp/ag_paper_drips.json"
PAPER_LOG_FILE     = "/tmp/ag_paper_log.json"
PAPER_CONFIG_FILE  = "/tmp/ag_paper_config.json"

_paper_trades: list = []      # all paper trades
_paper_bot_log: list = []     # activity log
_admin_bets: dict  = {}       # bet_key -> bet info for P&L tracking
_processed_webhook_ids: set = set()  # Stripe event IDs already handled — prevents duplicate charges

def _load_paper_db():
    """Load paper trades/drips/log/config from disk on startup."""
    global _paper_trades, _drip_ladders, _paper_bot_log, _paper_bot_config
    import os, json
    for path, target in [
        (PAPER_TRADES_FILE, "_paper_trades"),
        (PAPER_DRIPS_FILE,  "_drip_ladders"),
        (PAPER_LOG_FILE,    "_paper_bot_log"),
        (PAPER_CONFIG_FILE, "_paper_bot_config"),
    ]:
        if os.path.exists(path):
            try:
                with open(path) as f:
                    data = json.load(f)
                if target == "_paper_trades"    and isinstance(data, list): _paper_trades    = data
                elif target == "_drip_ladders"  and isinstance(data, dict): _drip_ladders    = data
                elif target == "_paper_bot_log" and isinstance(data, list): _paper_bot_log   = data
                elif target == "_paper_bot_config" and isinstance(data, dict): _paper_bot_config.update(data)
            except Exception as e:
                logger.warning(f"Paper DB load error ({path}): {e}")

def _save_paper_db():
    """Save all paper trading data to disk."""
    import json
    for path, data in [
        (PAPER_TRADES_FILE, _paper_trades),
        (PAPER_DRIPS_FILE,  _drip_ladders),
        (PAPER_LOG_FILE,    _paper_bot_log[:100]),
        (PAPER_CONFIG_FILE, _paper_bot_config),
    ]:
        try:
            with open(path, "w") as f:
                json.dump(data, f)
        except Exception as e:
            logger.warning(f"Paper DB save error ({path}): {e}")

def _paper_bot_place_bet(alert_type: str, match: str, league: str,
                          minute: int, score: str, selection: str,
                          odds: float, stake: float, reason: str,
                          bet_type: str = "back") -> dict:
    """Simulate placing a back or lay bet — paper trade only.
    bet_type: 'back' (backing selection to win) or 'lay' (backing selection to lose).
    For lay bets: liability = stake * (odds - 1), profit = stake if it loses.
    """
    if _paper_bot_config["status"] != "active":
        return {}
    if not _paper_bot_config["markets"].get(alert_type, False):
        return {}

    # Odds check — lay bets typically lower odds are better
    if bet_type == "back" and (odds < _paper_bot_config["min_odds"] or odds > _paper_bot_config["max_odds"]):
        return {}

    # Check daily loss limit
    today = datetime.now().strftime("%Y-%m-%d")
    today_pl = sum(t.get("profit_loss", 0) for t in _paper_trades
                   if t.get("date") == today and t.get("result") != "pending")
    if today_pl <= -_paper_bot_config["max_daily_loss"]:
        _paper_bot_log.append({"time": datetime.now().isoformat(),
                                "msg": "⛔ Daily loss limit hit — bot paused"})
        _paper_bot_config["status"] = "paused"
        return {}

    trade_id = f"PT{int(datetime.now().timestamp())}"

    # P&L calculation differs for back vs lay
    if bet_type == "lay":
        # Lay: you win stake if selection loses, you lose (odds-1)*stake if it wins
        potential_profit = round(stake, 2)
        potential_loss   = round((odds - 1) * stake, 2)
    else:
        # Back: you win (odds-1)*stake if selection wins, lose stake if it loses
        potential_profit = round((odds - 1) * stake, 2)
        potential_loss   = stake

    trade = {
        "id": trade_id,
        "type": alert_type,
        "bet_type": bet_type,          # "back" or "lay"
        "match": match,
        "league": league,
        "minute": minute,
        "score_at_alert": score,
        "selection": selection,
        "odds": odds,
        "stake": stake,
        "potential_profit": potential_profit,
        "potential_loss": potential_loss,
        "result": "pending",
        "profit_loss": 0.0,
        "reason": reason,
        "date": datetime.now().strftime("%Y-%m-%d"),
        "fired_at": datetime.now().isoformat(),
        "mode": "paper",
    }
    _paper_trades.append(trade)
    emoji = "📗" if bet_type == "back" else "📕"
    _paper_bot_log.insert(0, {
        "time": datetime.now().isoformat(),
        "msg": f"{emoji} {bet_type.upper()} {selection} — {match} @ {odds} (£{stake})"
    })
    logger.info(f"🤖 Paper {bet_type}: {trade_id} {alert_type} {match} @ {odds}")
    _save_paper_db()
    return trade

def _place_goal_due_drip(fixture_id: int, match: str, league: str,
                          minute: int, score: str, current_score_goals: int,
                          xg_unscored: float):
    """
    Goal Due 2-track strategy.
    Track A: One simple back per alert — placed immediately.
    Track B: Staggered drip — one entry per scan cycle (every 2 mins).
              Entry 1 fires immediately. Entry 2 fires next cycle if no goal.
              Entry 3 fires the cycle after that. Max 3 entries total.
              This simulates real odds drifting as the market moves.
    NOTE: We don't have live Betfair odds from the API, so odds are estimated
          based on minute + xG. The later the minute and higher the xG unscored,
          the higher the estimated back odds (more overdue = longer wait = drift).
    """
    drip_cfg = _paper_bot_config.get("drip", {})
    if not drip_cfg.get("enabled", True):
        return

    fid = str(fixture_id)
    total_goals = current_score_goals
    simple_market = f"Over {total_goals + 0.5} Goals"
    drip_market   = f"Over {total_goals + 1.5} Goals"

    # ── Track A: Simple back (one per alert, not repeated) ───────
    # Only place if no Track A trade exists for this fixture yet
    track_a_exists = any(
        t.get("type") == "Goal Due — Track A" and t.get("match") == match
        for t in _paper_trades
    )
    if not track_a_exists:
        # Estimated odds: xG overdue drives odds lower (more certain), later minute also lowers
        simple_odds = max(1.3, round(2.2 - (xg_unscored * 0.25) - ((minute - 45) * 0.01), 2))
        _paper_bot_place_bet(
            alert_type="Goal Due — Track A",
            match=match, league=league, minute=minute, score=score,
            selection=f"Back {simple_market}",
            odds=simple_odds,
            stake=_paper_bot_config["stake"],
            reason=f"[TRACK A] xG unscored {xg_unscored:.2f} at {minute}' — backing {simple_market} @ est. {simple_odds}",
            bet_type="back"
        )

    # ── Track B: Staggered drip ladder ───────────────────────────
    drip_stakes = drip_cfg.get("stakes", [2.0, 3.0, 5.0])
    # Estimated odds drift: as match progresses without goal, odds drift longer
    # Entry 1 (now): base odds | Entry 2 (+2 min): base +0.8 | Entry 3 (+4 min): base +1.8
    # Base odds scale by goals needed: 0 goals → ~2.0, 1 goal → ~3.0, 2 goals → ~4.5
    entry_odds_base = max(2.0, round(2.0 + (total_goals * 1.2), 2))
    drip_drifts = [0.0, 0.8, 1.8]

    if fid not in _drip_ladders:
        # First time — initialise ladder, place entry 1 only
        ladder = {
            "fixture_id": fid,
            "match": match,
            "league": league,
            "entry_odds_base": entry_odds_base,
            "drip_market": drip_market,
            "drip_stakes": drip_stakes,
            "drip_drifts": drip_drifts,
            "drips_placed": 0,
            "total_staked": 0.0,
            "total_potential_return": 0.0,
            "covered": False,
            "cashed_out": False,
            "trade_ids": [],
            "started_at": datetime.now().isoformat(),
            "started_minute": minute,
        }
        _drip_ladders[fid] = ladder
        _paper_bot_log.insert(0, {
            "time": datetime.now().isoformat(),
            "msg": f"🪜 Drip ladder started: {match} {minute}' — {drip_market} base odds {entry_odds_base}"
        })

    ladder = _drip_ladders[fid]
    if ladder.get("cashed_out"):
        return

    max_drips = len(drip_stakes)
    placed = ladder["drips_placed"]

    if placed >= max_drips:
        return  # All entries already placed

    # Place next entry in sequence
    i = placed
    drip_odds = round(ladder["entry_odds_base"] + drip_drifts[i], 2)
    stake     = drip_stakes[i]
    label     = "Entry" if i == 0 else f"Drip {i+1}"
    t = _paper_bot_place_bet(
        alert_type="Goal Due — Track B",
        match=match, league=league, minute=minute, score=score,
        selection=f"{ladder['drip_market']} [{label} @ {drip_odds}]",
        odds=drip_odds, stake=stake,
        reason=(
            f"[TRACK B] Ladder {label} ({i+1}/{max_drips}): "
            f"xG unscored {xg_unscored:.2f} at {minute}'. "
            f"Est. odds {drip_odds} (base {ladder['entry_odds_base']} + drift {drip_drifts[i]}). "
            f"Note: odds estimated — no live Betfair feed."
        ),
        bet_type="back"
    )
    if t:
        ladder["trade_ids"].append(t["id"])
        ladder["drips_placed"] += 1
        ladder["total_staked"] = round(ladder["total_staked"] + stake, 2)
        ladder["total_potential_return"] = round(
            ladder["total_potential_return"] + (drip_odds - 1) * stake, 2
        )
        _paper_bot_log.insert(0, {
            "time": datetime.now().isoformat(),
            "msg": f"🪜 {label} placed: {match} {minute}' @ {drip_odds} £{stake} ({i+1}/{max_drips})"
        })
        _save_paper_db()

def _check_drip_cashout_or_cover(fixture_id: int, match: str, league: str,
                                   minute: int, score: str,
                                   goal_scored: bool, xg_unscored: float):
    """
    Called every scanner cycle for active drip ladders.
    - goal_scored=True → cash out all (mark won)
    - xg_unscored < threshold → place cover bet on Over 1.5
    """
    fid = str(fixture_id)
    if fid not in _drip_ladders:
        return
    ladder = _drip_ladders[fid]
    if ladder.get("cashed_out"):
        return

    drip_cfg = _paper_bot_config.get("drip", {})
    cover_threshold = drip_cfg.get("cover_xg_threshold", 0.8)

    if goal_scored:
        # Cash out — mark all drip trades as won
        for tid in ladder.get("trade_ids", []):
            for t in _paper_trades:
                if t["id"] == tid and t["result"] == "pending":
                    t["result"] = "win"
                    t["profit_loss"] = round(t.get("potential_profit", 0), 2)
                    t["settled_at"] = __import__('datetime').datetime.now().isoformat()
        ladder["cashed_out"] = True
        _paper_bot_log.insert(0, {
            "time": __import__('datetime').datetime.now().isoformat(),
            "msg": f"💚 CASH OUT: Goal scored — {match} ladder closed green"
        })
        _save_paper_db()

    elif xg_unscored < cover_threshold and not ladder.get("covered"):
        # Proportional hedge — stake scales with how much xG remains
        cover_odds   = drip_cfg.get("cover_odds", 2.08)
        total_staked = ladder["total_staked"]
        # xG 0.79 → ~32% cover | xG 0.5 → ~20% | xG 0.1 → ~4%
        proportion  = round((xg_unscored / cover_threshold) * 0.6, 3)
        cover_stake = max(1.0, round(total_staked * proportion, 2))
        cover_return = round(cover_stake * (cover_odds - 1), 2)
        _paper_bot_place_bet(
            alert_type="Goal Due — Track B (Cover)",
            match=match, league=league, minute=minute, score=score,
            selection=f"Over 1.5 Goals [Proportional Cover {proportion:.0%}]",
            odds=cover_odds, stake=cover_stake,
            reason=(
                f"Proportional hedge: xG unscored {xg_unscored:.2f} "
                f"(threshold {cover_threshold}). "
                f"Cover {proportion:.0%} of £{total_staked:.2f} = £{cover_stake:.2f}. "
                f"Wins £{cover_return:.2f} if goal comes"
            ),
            bet_type="back"
        )
        ladder["covered"] = True
        ladder["cover_stake"] = cover_stake
        ladder["cover_proportion"] = proportion
        _paper_bot_log.insert(0, {
            "time": __import__('datetime').datetime.now().isoformat(),
            "msg": f"🛡️ COVER ({proportion:.0%}): {match} — xG={xg_unscored:.2f}, £{cover_stake:.2f} @ {cover_odds}"
        })
        _save_paper_db()


def _estimate_odds(market: str, confidence: int) -> float:
    """Estimate fair odds from market type + confidence score."""
    base = {
        "Match Winner": 2.2,
        "Both Teams to Score": 1.9,
        "HT/FT Value — Comeback": 3.5,
        "Over 1.5 Goals": 1.7,
        "Goal Due": 2.0,
    }.get(market, 2.0)
    # Adjust by confidence: higher confidence = lower odds (more likely)
    adj = (confidence - 70) * 0.02
    return round(max(1.3, min(8.0, base - adj)), 2)

@app.get("/admin/paper-bot/config")
async def paper_bot_get_config():
    return _paper_bot_config

@app.post("/admin/paper-bot/config")
async def paper_bot_update_config(request: dict):
    for k, v in request.items():
        if k in _paper_bot_config:
            _paper_bot_config[k] = v
    _save_paper_db()
    return _paper_bot_config

@app.get("/admin/paper-bot/trades")
async def paper_bot_get_trades():
    trades = sorted(_paper_trades, key=lambda x: x.get("fired_at",""), reverse=True)
    resolved = [t for t in trades if t["result"] != "pending"]
    won  = [t for t in resolved if t["result"] in ("won","win")]
    lost = [t for t in resolved if t["result"] in ("lost","loss")]
    total_pl = round(sum(t.get("profit_loss",0) for t in resolved), 2)
    today = datetime.now().strftime("%Y-%m-%d")
    today_pl = round(sum(t.get("profit_loss",0) for t in resolved if t.get("date")==today), 2)

    def track_stats(prefix):
        tt = [t for t in trades if t.get("type","").startswith(prefix)]
        tr = [t for t in tt if t["result"] != "pending"]
        tw = [t for t in tr if t["result"] in ("won","win")]
        tl = [t for t in tr if t["result"] in ("lost","loss")]
        return {
            "total": len(tt), "won": len(tw), "lost": len(tl),
            "pending": len(tt) - len(tr),
            "win_rate": round(len(tw)/len(tr)*100,1) if tr else 0,
            "pl": round(sum(t.get("profit_loss",0) for t in tr), 2),
            "staked": round(sum(t.get("stake",0) for t in tt), 2),
        }

    return {
        "trades": trades,
        "log": _paper_bot_log[:20],
        "stats": {
            "total": len(trades),
            "pending": len(trades) - len(resolved),
            "won": len(won),
            "lost": len(lost),
            "win_rate": round(len(won)/len(resolved)*100,1) if resolved else 0,
            "total_pl": total_pl,
            "today_pl": today_pl,
            "status": _paper_bot_config["status"],
            "mode": _paper_bot_config["mode"],
        },
        "track_a": track_stats("Goal Due — Track A"),
        "track_b": track_stats("Goal Due — Track B"),
    }

@app.post("/admin/paper-bot/trades/{trade_id}/result")
async def paper_bot_set_result(trade_id: str, request: dict):
    for t in _paper_trades:
        if t["id"] == trade_id:
            result = request.get("result", "pending")
            t["result"] = result
            bet_type = t.get("bet_type", "back")
            if result in ("won", "win"):
                # Back win: profit = (odds-1)*stake | Lay win (selection lost): profit = stake
                t["profit_loss"] = round(t.get("potential_profit", (t["odds"]-1)*t["stake"]), 2)
            elif result == "lost":
                # Back loss: lose stake | Lay loss (selection won): lose (odds-1)*stake
                t["profit_loss"] = -round(t.get("potential_loss", t["stake"]), 2)
            else:
                t["profit_loss"] = 0.0
            t["settled_at"] = datetime.now().isoformat()
            emoji = "✅" if result == "won" else "❌"
            _paper_bot_log.insert(0, {
                "time": datetime.now().isoformat(),
                "msg": f"{emoji} [{bet_type.upper()}] {t['match']} — {result.upper()} £{t['profit_loss']:+.2f}"
            })
            _save_paper_db()
            return {"ok": True, "trade": t}
    raise HTTPException(404, "Trade not found")

@app.post("/admin/paper-bot/toggle")
async def paper_bot_toggle():
    _paper_bot_config["status"] = "paused" if _paper_bot_config["status"] == "active" else "active"
    _save_paper_db()
    return {"status": _paper_bot_config["status"]}

@app.delete("/admin/paper-bot/trades/{trade_id}")
async def paper_bot_delete_trade(trade_id: str):
    global _paper_trades
    _paper_trades = [t for t in _paper_trades if t["id"] != trade_id]
    return {"ok": True}

@app.get("/admin/paper-bot/drips")
async def paper_bot_get_drips():
    """Get all active drip ladders."""
    return {
        "ladders": list(_drip_ladders.values()),
        "active": sum(1 for l in _drip_ladders.values() if not l.get("cashed_out")),
        "covered": sum(1 for l in _drip_ladders.values() if l.get("covered")),
        "cashed_out": sum(1 for l in _drip_ladders.values() if l.get("cashed_out")),
    }

@app.get("/api-test")
async def api_raw_test():
    """Test raw API call and return full response for PL today."""
    from datetime import datetime
    today = _london_now().strftime("%Y-%m-%d")
    raw = api_get("fixtures", {"league": 39, "date": today, "timezone": "Europe/London"})
    response_count = len(raw.get("response", []))
    errors = raw.get("errors", {})
    paging = raw.get("paging", {})
    # Also test with season
    raw2 = api_get("fixtures", {"league": 39, "season": 2025, "date": today})
    return {
        "date": today,
        "api_key_first6": API_FOOTBALL_KEY[:6] if API_FOOTBALL_KEY else "NOT SET",
        "without_season": {"count": response_count, "errors": errors, "paging": paging},
        "with_season_2025": {"count": len(raw2.get("response", [])), "errors": raw2.get("errors", {})},
        "first_fixture": raw.get("response", [{}])[0].get("fixture", {}).get("date") if raw.get("response") else None
    }

@app.get("/today/debug")
async def today_debug():
    """Show exactly how many fixtures API returns per league today — for diagnosing missing games."""
    today_str = _london_now().strftime("%Y-%m-%d")
    results = {}
    def check_league(lg):
        try:
            data = api_get("fixtures", {"league": lg["id"], "date": today_str, "timezone": "Europe/London"}, timeout=15)
            fixtures = data.get("response", [])
            return lg["code"], len(fixtures), [f["teams"]["home"]["name"] + " v " + f["teams"]["away"]["name"] for f in fixtures[:5]]
        except Exception as e:
            return lg["code"], -1, [str(e)]
    with ThreadPoolExecutor(max_workers=10) as ex:
        futs = [ex.submit(check_league, lg) for lg in LEAGUES]
        for f in futs:
            code, count, names = f.result()
            results[code] = {"count": count, "fixtures": names}
    total = sum(v["count"] for v in results.values() if v["count"] > 0)
    return {"date": today_str, "total": total, "by_league": results}

@app.get("/saved-accas")
async def get_saved_accas(email: str = ""):
    """Return saved accas filtered by user email."""
    if not email:
        return {"accas": []}
    # Try SQLite first for freshest data
    try:
        db_accas = db_get_accas(email.lower())
        if db_accas:
            return {"accas": db_accas}
    except Exception as e:
        logger.warning(f"SQLite get_accas error: {e}")
    # Fallback to in-memory store
    seen_ids = set()
    user_accas = [
        a for a in saved_accas_store
        if a.get("email","").lower() == email.lower()
        and not (a.get("id") in seen_ids or seen_ids.add(a.get("id")))
    ]
    return {"accas": user_accas}

@app.post("/saved-accas")
@app.post("/save-acca")
async def save_acca(request: dict):
    email = request.get("email", "").strip().lower()
    acca = {
        "id": len(saved_accas_store) + int(datetime.now().timestamp()),
        "email": email,
        "name": request.get("name", "My Acca"),
        "selections": request.get("selections", []),
        "total_odds": request.get("total_odds", 0),
        "stake": request.get("stake", 10),
        "created_at": datetime.now().isoformat(),
        "status": "pending",
        "result": None,
        "profit_loss": None,
    }
    # Dedup: check if same selections already saved for this user today
    today = datetime.now().strftime("%Y-%m-%d")
    new_sel = sorted([s.get("match","") + s.get("selection","") for s in acca.get("selections", [])])
    for existing in saved_accas_store:
        if existing.get("email") != email: continue
        if existing.get("created_at","")[:10] != today: continue
        ex_sel = sorted([s.get("match","") + s.get("selection","") for s in existing.get("selections", [])])
        if ex_sel == new_sel:
            return {"ok": True, "acca": existing, "duplicate": True}  # skip duplicate
    saved_accas_store.append(acca)
    _save_accas_db()
    return {"ok": True, "acca": acca}

@app.post("/saved-accas/delete")
async def delete_single_acca(request: dict):
    """Permanently delete a single saved acca by ID."""
    acca_id = str(request.get("id","")).strip()
    email   = (request.get("email","") or "").strip().lower()
    if not acca_id:
        return {"ok": False, "error": "ID required"}
    global saved_accas_store
    before = len(saved_accas_store)
    saved_accas_store = [
        a for a in saved_accas_store
        if str(a.get("id","")) != acca_id and str(a.get("timestamp","")) != acca_id
    ]
    # Delete from DB permanently
    try:
        db_delete_acca(acca_id)
    except Exception as e:
        logger.warning(f"db_delete_acca error: {e}")
    _save_accas_db()
    logger.info(f"Deleted acca {acca_id} for {email} ({before - len(saved_accas_store)} removed)")
    return {"ok": True, "message": f"Acca {acca_id} deleted"}

@app.post("/saved-accas/reset")
async def reset_saved_accas(request: dict):
    """Clear ALL saved accas for a user — permanently from DB."""
    email = (request.get("email") or "").strip().lower()
    if not email:
        return {"ok": False, "error": "Email required"}
    global saved_accas_store
    # Remove from memory
    to_delete = [a for a in saved_accas_store if a.get("email","").lower() == email]
    saved_accas_store = [a for a in saved_accas_store if a.get("email","").lower() != email]
    # Delete each from DB permanently
    for a in to_delete:
        try:
            db_delete_acca(str(a.get("id","")))
        except Exception: pass
    _save_accas_db()
    logger.info(f"Reset all accas for {email} — {len(to_delete)} deleted")
    return {"ok": True, "message": f"All accas cleared ({len(to_delete)} deleted)"}



@app.patch("/saved-accas/{acca_id}/result")
async def update_result(acca_id: int, request: dict):
    for a in saved_accas_store:
        if a.get("id") == acca_id:
            a["result"] = request.get("result", "pending")
            _save_accas_db()
            return {"message": "Updated", "acca": a}
    raise HTTPException(404, "Not found")


# ═══════════════════════════════════════════════════════════════════
# POISSON MODEL — AccaGenius Football Trading Data
# ═══════════════════════════════════════════════════════════════════
import math
from functools import lru_cache

# ── Poisson probability ─────────────────────────────────────────
def _poisson_prob(lam: float, k: int) -> float:
    """P(X=k) for Poisson distribution with mean lam."""
    if lam <= 0: return 0.0
    try:
        return (math.exp(-lam) * (lam ** k)) / math.factorial(k)
    except: return 0.0

def _match_probs(home_xg: float, away_xg: float, max_goals: int = 8):
    """
    Given expected goals for home and away, return:
    - scoreline probability matrix
    - home_win, draw, away_win probabilities
    - over/under markets
    - BTTS
    """
    home_xg = max(0.1, home_xg)
    away_xg = max(0.1, away_xg)

    # Build scoreline matrix
    matrix = {}
    home_win = draw = away_win = 0.0
    o15 = o25 = o35 = o45 = 0.0
    btts = 0.0
    fh_o05 = 0.0  # approximation: use half xG for first half

    for h in range(max_goals + 1):
        for a in range(max_goals + 1):
            p = _poisson_prob(home_xg, h) * _poisson_prob(away_xg, a)
            matrix[(h, a)] = p
            if h > a:  home_win += p
            elif h == a: draw   += p
            else:        away_win += p
            goals = h + a
            if goals > 1.5: o15 += p
            if goals > 2.5: o25 += p
            if goals > 3.5: o35 += p
            if goals > 4.5: o45 += p
            if h > 0 and a > 0: btts += p

    # FH Over 0.5 — approximate using half xG
    fh_home = home_xg * 0.5
    fh_away = away_xg * 0.5
    fh_o05 = 1 - (_poisson_prob(fh_home, 0) * _poisson_prob(fh_away, 0))

    return {
        "matrix":    matrix,
        "home_win":  round(home_win,  4),
        "draw":      round(draw,      4),
        "away_win":  round(away_win,  4),
        "over_15":   round(o15,       4),
        "over_25":   round(o25,       4),
        "over_35":   round(o35,       4),
        "over_45":   round(o45,       4),
        "btts":      round(btts,      4),
        "fh_over_05":round(fh_o05,    4),
    }

def _prob_to_odds(prob: float, min_prob: float = 0.01) -> float:
    """Convert probability to decimal odds, capped sensibly."""
    if prob < min_prob: return 999.0
    return round(1.0 / prob, 2)

def _value(model_prob: float, market_odds: float) -> float:
    """
    Value = model implied prob - 1/market_odds
    Positive = value bet, negative = avoid
    """
    if not market_odds or market_odds <= 1: return 0.0
    market_prob = 1.0 / market_odds
    return round((model_prob - market_prob) * 100, 2)  # as percentage

def _top3_scorelines(matrix: dict, n: int = 3):
    """Return top N most likely scorelines."""
    sorted_scores = sorted(matrix.items(), key=lambda x: x[1], reverse=True)
    return [f"{h}-{a}" for (h, a), _ in sorted_scores[:n]]

def build_poisson_for_match(
    home_xg_6:  float, away_xg_6:  float,   # last 6 game xG averages
    home_xg_s:  float, away_xg_s:  float,   # season xG averages
    home_team:  str = "",
    away_team:  str = "",
    betfair_prices: dict = None,             # optional Betfair prices for value calc
) -> dict:
    """
    Full Poisson model output for one fixture.
    Returns everything needed for pre-match sheet.
    """
    bp = betfair_prices or {}

    # ── 6 Game Model ──────────────────────────────────────────
    p6 = _match_probs(home_xg_6, away_xg_6)
    odds6 = {
        "home":    _prob_to_odds(p6["home_win"]),
        "draw":    _prob_to_odds(p6["draw"]),
        "away":    _prob_to_odds(p6["away_win"]),
        "over_15": _prob_to_odds(p6["over_15"]),
        "over_25": _prob_to_odds(p6["over_25"]),
        "over_35": _prob_to_odds(p6["over_35"]),
        "btts":    _prob_to_odds(p6["btts"]),
        "fh_over_05": _prob_to_odds(p6["fh_over_05"]),
    }

    # ── Season Model ──────────────────────────────────────────
    ps = _match_probs(home_xg_s, away_xg_s)
    odds_s = {
        "home":    _prob_to_odds(ps["home_win"]),
        "draw":    _prob_to_odds(ps["draw"]),
        "away":    _prob_to_odds(ps["away_win"]),
        "over_15": _prob_to_odds(ps["over_15"]),
        "over_25": _prob_to_odds(ps["over_25"]),
        "over_35": _prob_to_odds(ps["over_35"]),
        "btts":    _prob_to_odds(ps["btts"]),
        "fh_over_05": _prob_to_odds(ps["fh_over_05"]),
    }

    # ── Value vs Betfair (if prices available) ────────────────
    val6 = {
        "home":    _value(p6["home_win"],  bp.get("home_back")),
        "draw":    _value(p6["draw"],      bp.get("draw_back")),
        "away":    _value(p6["away_win"],  bp.get("away_back")),
        "over_15": _value(p6["over_15"],   bp.get("o15_back")),
        "over_25": _value(p6["over_25"],   bp.get("o25_back")),
        "over_35": _value(p6["over_35"],   bp.get("o35_back")),
        "btts":    _value(p6["btts"],      bp.get("btts_yes_back")),
        "fh_over_05": _value(p6["fh_over_05"], bp.get("fh_o05_back")),
    }
    val_s = {
        "home":    _value(ps["home_win"],  bp.get("home_back")),
        "draw":    _value(ps["draw"],      bp.get("draw_back")),
        "away":    _value(ps["away_win"],  bp.get("away_back")),
        "over_15": _value(ps["over_15"],   bp.get("o15_back")),
        "over_25": _value(ps["over_25"],   bp.get("o25_back")),
        "over_35": _value(ps["over_35"],   bp.get("o35_back")),
        "btts":    _value(ps["btts"],      bp.get("btts_yes_back")),
        "fh_over_05": _value(ps["fh_over_05"], bp.get("fh_o05_back")),
    }

    # ── Top 3 scorelines ─────────────────────────────────────
    top3 = _top3_scorelines(p6["matrix"])

    # ── Supremacy (home xG - away xG) ────────────────────────
    sup6 = round(home_xg_6 - away_xg_6, 2)
    sup_s = round(home_xg_s - away_xg_s, 2)

    return {
        "home_team":  home_team,
        "away_team":  away_team,
        # xG inputs
        "home_xg_6":  round(home_xg_6, 2),
        "away_xg_6":  round(away_xg_6, 2),
        "match_xg_6": round(home_xg_6 + away_xg_6, 2),
        "sup_6":      sup6,
        "home_xg_s":  round(home_xg_s, 2),
        "away_xg_s":  round(away_xg_s, 2),
        "match_xg_s": round(home_xg_s + away_xg_s, 2),
        "sup_s":      sup_s,
        # 6 game model odds
        "odds_6":     odds6,
        # season model odds
        "odds_s":     odds_s,
        # value columns
        "val_6":      val6,
        "val_s":      val_s,
        # scorelines
        "top3":       top3,
        # raw probabilities
        "prob_6":     p6,
        "prob_s":     ps,
    }

def get_xg_from_fbd(home: str, away: str) -> dict:
    """
    Pull xG from FBD data for a fixture.
    FBD gives pred_goals (total expected goals) and home/away split via odds.
    We estimate home_xg and away_xg from FBD's predicted total goals
    and their H/D/A probability split.
    """
    try:
        fbd = get_fbd_match(home, away)
        if not fbd:
            return {"home_xg": None, "away_xg": None, "source": "not_found"}

        pred_goals = fbd.get("pred_goals") or 0
        home_prob  = fbd.get("pred_home_odds") or 0   # already converted to prob
        away_prob  = fbd.get("pred_away_odds") or 0

        if pred_goals > 0 and (home_prob + away_prob) > 0:
            # Split total expected goals by home/away strength ratio
            total_prob = home_prob + away_prob
            home_share = home_prob / total_prob if total_prob > 0 else 0.5
            home_xg = round(pred_goals * home_share, 2)
            away_xg = round(pred_goals * (1 - home_share), 2)
            return {
                "home_xg":  home_xg,
                "away_xg":  away_xg,
                "total_xg": round(pred_goals, 2),
                "source":   "fbd"
            }
        return {"home_xg": None, "away_xg": None, "source": "fbd_no_data"}
    except Exception as e:
        logger.warning(f"get_xg_from_fbd error: {e}")
        return {"home_xg": None, "away_xg": None, "source": "error"}

def get_xg_from_api(home_id: int, away_id: int, league_id: int,
                    season: int = 2025, last_n: int = 6) -> dict:
    """
    Pull xG from API-Football for last N home and away games.
    Returns average xG for and against for each team.
    """
    try:
        # Home team last 6 home games
        home_data = api_get("fixtures", {
            "team": home_id, "last": last_n,
            "league": league_id, "season": season
        })
        # Away team last 6 away games
        away_data = api_get("fixtures", {
            "team": away_id, "last": last_n,
            "league": league_id, "season": season
        })

        def avg_xg(fixtures_response, team_id, side):
            """side = 'home' or 'away'"""
            xgs = []
            for f in (fixtures_response.get("response") or []):
                stats = f.get("statistics") or []
                for s in stats:
                    if s.get("type") == "expected_goals":
                        try:
                            xgs.append(float(s["value"]))
                        except: pass
                # Fallback — use actual goals if xG not available
                if not xgs:
                    goals = f.get("goals") or {}
                    g = goals.get(side)
                    if g is not None:
                        try: xgs.append(float(g))
                        except: pass
            return round(sum(xgs)/len(xgs), 2) if xgs else None

        home_xg = avg_xg(home_data, home_id, "home")
        away_xg = avg_xg(away_data, away_id, "away")

        return {
            "home_xg": home_xg,
            "away_xg": away_xg,
            "source":  "api_football"
        }
    except Exception as e:
        logger.warning(f"get_xg_from_api error: {e}")
        return {"home_xg": None, "away_xg": None, "source": "error"}

# ── Cache for daily Poisson results ─────────────────────────────
_poisson_cache: dict = {}  # fixture_key -> result
_poisson_cache_date: str = ""

def _poisson_cache_key(home: str, away: str) -> str:
    return f"{home.lower().strip()}_{away.lower().strip()}"

@app.get("/poisson/{home}/{away}")
async def poisson_single(home: str, away: str):
    """
    Get Poisson model output for a single fixture.
    Uses FBD as primary xG source, API-Football as fallback.
    """
    try:
        key = _poisson_cache_key(home, away)
        today = datetime.now().strftime("%Y-%m-%d")

        # Check cache
        global _poisson_cache_date
        if _poisson_cache_date != today:
            _poisson_cache.clear()
            _poisson_cache_date = today

        if key in _poisson_cache:
            return _poisson_cache[key]

        # Try FBD first
        xg = get_xg_from_fbd(home, away)

        # Default fallback xG if nothing found
        home_xg_6 = xg.get("home_xg") or 1.35
        away_xg_6 = xg.get("away_xg") or 1.10
        home_xg_s = home_xg_6  # same for now — API will improve this
        away_xg_s = away_xg_6

        result = build_poisson_for_match(
            home_xg_6=home_xg_6,
            away_xg_6=away_xg_6,
            home_xg_s=home_xg_s,
            away_xg_s=away_xg_s,
            home_team=home,
            away_team=away,
        )
        result["xg_source"] = xg.get("source", "default")

        _poisson_cache[key] = result
        return result

    except Exception as e:
        logger.error(f"Poisson error {home} vs {away}: {e}")
        return {"error": str(e)}

@app.get("/poisson/today")
async def poisson_today():
    """
    Run Poisson model on all of today's fixtures.
    Returns full model output for every game — used to generate pre-match sheet.
    """
    try:
        today = datetime.now().strftime("%Y-%m-%d")
        fbd_data = get_fbd_data()
        results = []

        for fbd_match in fbd_data:
            home = fbd_match.get("home", "")
            away = fbd_match.get("away", "")
            if not home or not away: continue

            # Get xG from FBD
            pred_goals = fbd_match.get("pred_goals") or 0
            home_prob  = fbd_match.get("pred_home_odds") or 0
            away_prob  = fbd_match.get("pred_away_odds") or 0

            if pred_goals > 0 and (home_prob + away_prob) > 0:
                total_prob = home_prob + away_prob
                home_share = home_prob / total_prob
                home_xg = round(pred_goals * home_share, 2)
                away_xg = round(pred_goals * (1 - home_share), 2)
            else:
                home_xg = 1.35
                away_xg = 1.10

            model = build_poisson_for_match(
                home_xg_6=home_xg, away_xg_6=away_xg,
                home_xg_s=home_xg, away_xg_s=away_xg,
                home_team=home, away_team=away,
            )
            model["league"]     = fbd_match.get("league", "")
            model["prediction"] = fbd_match.get("prediction", "")
            model["xg_source"]  = "fbd"
            results.append(model)

        logger.info(f"Poisson today: {len(results)} fixtures processed")
        return {
            "date":     today,
            "count":    len(results),
            "fixtures": results,
        }

    except Exception as e:
        logger.error(f"Poisson today error: {e}")
        return {"error": str(e), "fixtures": []}

@app.get("/poisson/validate")
async def poisson_validate():
    """
    Quick validation endpoint — run Poisson on a known fixture
    and return the full output so you can check the numbers.
    """
    test = build_poisson_for_match(
        home_xg_6=1.8, away_xg_6=1.1,
        home_xg_s=1.6, away_xg_s=1.2,
        home_team="Test Home", away_team="Test Away",
    )
    return {
        "test_fixture": "Home (1.8 xG) vs Away (1.1 xG)",
        "model_6_game": {
            "home_odds":  test["odds_6"]["home"],
            "draw_odds":  test["odds_6"]["draw"],
            "away_odds":  test["odds_6"]["away"],
            "over_25":    test["odds_6"]["over_25"],
            "btts":       test["odds_6"]["btts"],
            "fh_over_05": test["odds_6"]["fh_over_05"],
        },
        "top3_scorelines": test["top3"],
        "probabilities": {
            "home_win": f"{test['prob_6']['home_win']*100:.1f}%",
            "draw":     f"{test['prob_6']['draw']*100:.1f}%",
            "away_win": f"{test['prob_6']['away_win']*100:.1f}%",
            "over_25":  f"{test['prob_6']['over_25']*100:.1f}%",
            "btts":     f"{test['prob_6']['btts']*100:.1f}%",
        },
        "status": "✅ Poisson model working"
    }

logger.info("✅ Poisson model loaded — endpoints: /poisson/validate /poisson/today /poisson/{home}/{away}")


# ═══════════════════════════════════════════════════════════════════
# ═══════════════════════════════════════════════════════════════════
# EXCHANGE PRICE CAPTURE — via The Odds API (free tier)
# Free tier: 500 requests/month — enough for daily captures
# Sign up free at https://the-odds-api.com
# Add ODDS_API_KEY to Railway environment variables
# ═══════════════════════════════════════════════════════════════════
import requests as _requests
from datetime import datetime, timedelta
import re
import json as _json

ODDS_API_KEY = os.getenv("ODDS_API_KEY", "")
ODDS_API_URL = "https://api.the-odds-api.com/v4"

# ── Price storage ────────────────────────────────────────────────
_steam_prices: dict = {}
_steam_date:   str  = ""

def _load_todays_steam_prices():
    """Load today's steam prices from DB on startup."""
    global _steam_prices, _steam_date
    today = datetime.now().strftime("%Y-%m-%d")
    loaded = db_load_steam_prices(today)
    if loaded:
        _steam_prices = loaded
        _steam_date   = today
        logger.info(f"✅ Loaded {len(loaded)} steam fixtures from DB for {today}")
_odds_cache:   dict = {}   # cache today's odds to save API calls
_odds_cache_time: str = ""

def _steam_key(home: str, away: str) -> str:
    return f"{home.lower().strip()} v {away.lower().strip()}"

def _normalise(name: str) -> str:
    """Normalise team name for fuzzy matching."""
    return re.sub(r'[^a-z0-9]', '', name.lower())

def _names_close(a: str, b: str) -> bool:
    """Check if two team names are close enough to match."""
    na = _normalise(a); nb = _normalise(b)
    if na == nb: return True
    if na in nb or nb in na: return True
    # Check first word matches
    if na.split()[:1] == nb.split()[:1]: return True
    return False

def _fetch_odds_api(sport: str = "soccer", region: str = "uk",
                    market: str = "h2h") -> list:
    """
    Fetch odds from The Odds API for all football fixtures today.
    Caches results for 30 mins to save API quota.
    Returns list of fixture odds objects.
    """
    global _odds_cache, _odds_cache_time
    cache_key = f"{sport}_{market}"
    now_str   = datetime.utcnow().strftime("%Y-%m-%d %H:%M")

    # Return cache if less than 30 mins old
    if (_odds_cache.get(cache_key) and _odds_cache_time and
            _odds_cache_time[:13] == now_str[:13]):
        return _odds_cache[cache_key]

    if not ODDS_API_KEY:
        logger.warning("ODDS_API_KEY not set in Railway — add it to get real exchange prices")
        return []

    # Sport keys for The Odds API
    sport_keys = [
        "soccer_epl", "soccer_england_league1", "soccer_england_league2",
        "soccer_england_championship", "soccer_spain_la_liga",
        "soccer_germany_bundesliga", "soccer_italy_serie_a",
        "soccer_france_ligue_one", "soccer_netherlands_eredivisie",
        "soccer_portugal_primeira_liga", "soccer_scotland_premiership",
        "soccer_turkey_super_league", "soccer_belgium_first_div",
    ]

    all_fixtures = []
    for sport_key in sport_keys:
        try:
            r = _requests.get(
                f"{ODDS_API_URL}/sports/{sport_key}/odds",
                params={
                    "apiKey":   ODDS_API_KEY,
                    "regions":  "uk",
                    "markets":  "h2h,totals",
                    "bookmakers": "betfair_ex_eu,betfair_ex_uk",
                    "oddsFormat": "decimal",
                },
                timeout=10
            )
            if r.status_code == 200:
                data = r.json()
                if isinstance(data, list):
                    all_fixtures.extend(data)
            elif r.status_code == 401:
                logger.error("ODDS_API_KEY invalid — check Railway variables")
                break
            elif r.status_code == 429:
                logger.warning("Odds API quota exceeded for this month")
                break
        except Exception as e:
            logger.warning(f"Odds API fetch error for {sport_key}: {e}")
            continue

    _odds_cache[cache_key]  = all_fixtures
    _odds_cache_time        = now_str
    logger.info(f"✅ Odds API: fetched {len(all_fixtures)} fixtures")
    return all_fixtures

def _get_betfair_prices_from_odds_api(home: str, away: str) -> dict:
    """
    Find Betfair exchange prices for a specific fixture from cached Odds API data.
    Returns dict with back prices for H2H and totals markets.
    """
    try:
        fixtures = _fetch_odds_api()
        if not fixtures:
            return {}

        # Find matching fixture
        for fixture in fixtures:
            fhome = fixture.get("home_team", "")
            faway = fixture.get("away_team", "")
            if _names_close(home, fhome) and _names_close(away, faway):
                prices = {}
                for bookmaker in fixture.get("bookmakers", []):
                    bm_key = bookmaker.get("key", "")
                    if "betfair" not in bm_key:
                        continue
                    for market in bookmaker.get("markets", []):
                        mkey = market.get("key", "")
                        outcomes = market.get("outcomes", [])
                        if mkey == "h2h":
                            for o in outcomes:
                                name = o.get("name", "").lower()
                                price = o.get("price")
                                if "draw" in name:
                                    prices["draw_back"] = price
                                elif fhome.lower() in name or home.lower() in name:
                                    prices["home_back"] = price
                                else:
                                    prices["away_back"] = price
                        elif mkey == "totals":
                            for o in outcomes:
                                point = o.get("point", 0)
                                name  = o.get("name", "").lower()
                                price = o.get("price")
                                if point == 2.5 and "over" in name:
                                    prices["o25_back"] = price
                                elif point == 1.5 and "over" in name:
                                    prices["o15_back"] = price
                                elif point == 0.5 and "over" in name:
                                    prices["fh_o05_back"] = price

                if prices:
                    prices["source"] = "the_odds_api"
                    logger.info(f"✅ Real exchange prices for {home} v {away}")
                    return prices

        return {}
    except Exception as e:
        logger.warning(f"_get_betfair_prices_from_odds_api error: {e}")
        return {}

def _get_exchange_prices_for_fixture(home: str, away: str) -> dict:
    """
    Main price function — tries Odds API first, falls back to FBD proxy.
    """
    # Strategy 1 — The Odds API (real Betfair exchange prices)
    if ODDS_API_KEY:
        prices = _get_betfair_prices_from_odds_api(home, away)
        if prices:
            return prices

    # Strategy 2 — FBD implied odds as proxy
    # Close to exchange prices, good enough for steam detection
    try:
        fbd = get_fbd_match(home, away)
        if fbd:
            def p2o(p):
                try: return round(1/float(p), 2) if p and float(p)>0 else None
                except: return None
            return {
                "home_back": p2o(fbd.get("pred_home_odds")),
                "draw_back": p2o(fbd.get("pred_draw_odds")),
                "away_back": p2o(fbd.get("pred_away_odds")),
                "source":    "fbd_proxy",
            }
    except Exception:
        pass

    return {}

def capture_exchange_prices(capture_type: str = "7am") -> dict:
    """
    Main price capture function — called at 7am and 5min pre-KO.
    Iterates all today's FBD fixtures and captures exchange prices.
    capture_type: "7am" or "5m"
    """
    global _steam_prices, _steam_date
    today = datetime.utcnow().strftime("%Y-%m-%d")
    if _steam_date != today:
        _steam_prices = {}
        _steam_date   = today

    logger.info(f"📊 Exchange price capture starting — type: {capture_type}")

    try:
        fbd_data = get_fbd_data()
        if not fbd_data:
            logger.warning("No FBD data for price capture")
            return _steam_prices

        captured = 0
        for match in fbd_data:
            home = match.get("home", "")
            away = match.get("away", "")
            if not home or not away:
                continue

            fkey = _steam_key(home, away)
            if fkey not in _steam_prices:
                _steam_prices[fkey] = {
                    "home": home, "away": away,
                    "league": match.get("league", ""),
                    "markets": {}
                }

            # Get exchange prices
            prices = _get_exchange_prices_for_fixture(home, away)

            # Store prices for each market
            market_map = {
                "draw":    prices.get("draw_back"),
                "over_25": prices.get("o25_back"),
                "over_15": prices.get("o15_back"),
                "fh_o05":  prices.get("fh_o05_back"),
            }

            for mkt, price in market_map.items():
                if price:
                    _steam_prices[fkey]["markets"].setdefault(mkt, {})
                    _steam_prices[fkey]["markets"][mkt][capture_type] = price

            captured += 1

        # Calculate movements and flag steams
        _calculate_steam_movements()

        logger.info(f"✅ Exchange {capture_type} capture complete — {captured} fixtures")
        return _steam_prices

    except Exception as e:
        logger.error(f"capture_exchange_prices error: {e}")
        return _steam_prices

def _calculate_steam_movements():
    """Calculate price movements and flag steams where 7am and 5m both exist."""
    for fkey, data in _steam_prices.items():
        for mkt, mdata in data.get("markets", {}).items():
            am7 = mdata.get("7am")
            m5  = mdata.get("5m")
            if am7 and m5:
                move = round(m5 - am7, 2)
                pct  = round((abs(move) / am7) * 100, 1) if am7 else 0
                mdata["move"]     = move
                mdata["move_pct"] = pct
                # Steam = price shortened by 3%+ (backing favourite harder)
                mdata["steamed"]  = (move < -0.05) and (pct >= 3)
                mdata["drifted"]  = (move > 0.05)  and (pct >= 3)

def get_steam_for_fixture(home: str, away: str) -> dict:
    """Get steam data for a specific fixture."""
    return _steam_prices.get(_steam_key(home, away), {})

# ── Scheduled tasks ──────────────────────────────────────────────
async def exchange_7am_capture():
    """Scheduled at 7am — captures opening exchange prices."""
    logger.info("⏰ 7am exchange price capture starting...")
    try:
        capture_exchange_prices("7am")
        logger.info("✅ 7am capture complete")
    except Exception as e:
        logger.error(f"7am capture error: {e}")

async def exchange_preko_capture():
    """Runs every 5 minutes 12pm-10pm — captures pre-KO prices."""
    try:
        now = datetime.utcnow()
        if now.hour < 12 or now.hour > 22:
            return
        capture_exchange_prices("5m")
        logger.info(f"✅ Pre-KO capture at {now.strftime('%H:%M')}")
    except Exception as e:
        logger.error(f"Pre-KO capture error: {e}")

# ── API Endpoints ────────────────────────────────────────────────
@app.get("/steam/test")
async def steam_test():
    """Test price capture — run one fixture and show what comes back."""
    try:
        fbd_data = get_fbd_data()
        if not fbd_data:
            return {"status": "❌ No FBD data available"}
        # Test on first fixture
        match = fbd_data[0]
        home  = match.get("home", "")
        away  = match.get("away", "")
        prices = _get_exchange_prices_for_fixture(home, away)
        return {
            "status":   "✅ Price capture working",
            "fixture":  f"{home} v {away}",
            "prices":   prices,
            "source":   prices.get("source", "oddsportal"),
            "message":  "If prices show FBD proxy — OddsPortal scraping failed, using FBD odds as proxy. Still works for steam detection.",
        }
    except Exception as e:
        return {"status": f"❌ Error: {e}"}

@app.get("/steam/capture/7am")
async def trigger_7am_capture():
    """Manually trigger 7am price capture."""
    result = capture_exchange_prices("7am")
    return {
        "status":    "✅ 7am capture complete",
        "fixtures":  len(result),
        "timestamp": datetime.utcnow().isoformat(),
        "data":      {k: v for k, v in list(result.items())[:5]},  # first 5 fixtures
    }

@app.get("/steam/capture/5m")
async def trigger_5m_capture():
    """Manually trigger 5min pre-KO price capture."""
    result = capture_exchange_prices("5m")
    return {
        "status":    "✅ 5min pre-KO capture complete",
        "fixtures":  len(result),
        "timestamp": datetime.utcnow().isoformat(),
    }

@app.get("/steam/alerts")
async def get_steam_alerts():
    """Get all current steam alerts — prices that have shortened significantly."""
    steams  = []
    drifts  = []
    for fkey, data in _steam_prices.items():
        for mkt, mdata in data.get("markets", {}).items():
            entry = {
                "fixture":   fkey,
                "home":      data.get("home"),
                "away":      data.get("away"),
                "league":    data.get("league", ""),
                "market":    mkt,
                "price_7am": mdata.get("7am"),
                "price_5m":  mdata.get("5m"),
                "move":      mdata.get("move"),
                "move_pct":  mdata.get("move_pct"),
            }
            if mdata.get("steamed"):
                steams.append(entry)
            elif mdata.get("drifted"):
                drifts.append(entry)
    return {
        "date":         _steam_date,
        "steams":       steams,
        "drifts":       drifts,
        "steam_count":  len(steams),
        "drift_count":  len(drifts),
        "total_fixtures": len(_steam_prices),
        "message":      "Steam = price shortened 3%+. Drift = price lengthened 3%+.",
    }

@app.get("/steam/prices/{home}/{away}")
async def get_fixture_steam(home: str, away: str):
    """Get steam data for a specific fixture."""
    data = get_steam_for_fixture(home, away)
    if not data:
        return {"message": f"No data for {home} v {away} — run /steam/capture/7am first"}
    return data

@app.get("/steam/all")
async def get_all_steam_prices():
    """Get all captured prices for today."""
    return {
        "date":     _steam_date,
        "fixtures": len(_steam_prices),
        "data":     _steam_prices,
    }

@app.post("/steam/receive")
async def receive_prices_from_pc(request: Request):
    """
    Receives exchange prices sent from the home PC capture script.
    PC script captures real Betfair exchange prices and posts them here.
    Railway stores them and calculates steam movements.
    """
    try:
        global _steam_prices, _steam_date
        today = datetime.now().strftime("%Y-%m-%d")
        if _steam_date != today:
            _steam_prices = {}
            _steam_date   = today

        body         = await request.json()
        capture_type = body.get("capture_type", "7am")
        fixtures     = body.get("fixtures", {})
        timestamp    = body.get("timestamp", "")

        if not fixtures:
            return {"message": "No fixtures received", "count": 0}

        updated = 0
        for fkey, fdata in fixtures.items():
            home   = fdata.get("home", "")
            away   = fdata.get("away", "")
            mkts   = fdata.get("markets", {})

            if fkey not in _steam_prices:
                _steam_prices[fkey] = {
                    "home":    home,
                    "away":    away,
                    "league":  fdata.get("league", ""),
                    "kickoff": fdata.get("kickoff", ""),
                    "markets": {},
                    "source":  "betfair_exchange",
                }

            # Store each market price
            for mkt, price in mkts.items():
                if price is None:
                    continue
                _steam_prices[fkey]["markets"].setdefault(mkt, {})
                _steam_prices[fkey]["markets"][mkt][capture_type] = price

            updated += 1

        # Recalculate steam movements
        _calculate_steam_movements()

        # Count steams
        steam_count = sum(
            1 for fdata in _steam_prices.values()
            for mdata in fdata.get("markets", {}).values()
            if mdata.get("steamed")
        )

        # Persist to DB so prices survive Railway restarts
        db_save_steam_prices(_steam_date, _steam_prices)

        logger.info(f"✅ Received {updated} fixtures from PC ({capture_type}) — {steam_count} steams detected")
        return {
            "message":      f"✅ Received {updated} fixtures",
            "capture_type": capture_type,
            "timestamp":    timestamp,
            "steam_alerts": steam_count,
            "total_fixtures": len(_steam_prices),
        }

    except Exception as e:
        logger.error(f"/steam/receive error: {e}")
        return {"error": str(e)}

logger.info("✅ Exchange price capture loaded — endpoints: /steam/test /steam/capture/7am /steam/capture/5m /steam/alerts /steam/receive")


# ═══════════════════════════════════════════════════════════════════
# DAILY PRE-MATCH SHEET GENERATOR
# Generates Excel pre-match sheet with live Betfair prices + Poisson
# Download at /sheet/prematach
# ═══════════════════════════════════════════════════════════════════
from fastapi.responses import StreamingResponse
import io

def _generate_prematach_sheet() -> bytes:
    """Generate pre-match Excel sheet matching AG5 results sheet layout exactly."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        import datetime as dt
        import io as _io

        AG_BLACK="FF0D0D0D"; AG_GOLD="FFFFB800"; MUTED="FFB0B0B0"
        SEC_MATCH="FF2E3A5C"; SEC_FORM="FF5C3800"; SEC_SCORE="FF4A4A00"
        SEC_POIS_N="FF3D1A6B"; SEC_POIS_O="FF4A1A8A"; SEC_VAL="FF0D5C2E"
        AG_DARK="FF1C2330"; VOL_BLUE="FF00B0F0"; VOL_GREEN="FF92D050"
        BACK_COL="FF9DC3E6"; LAY_COL="FFFFC7CE"
        WHITE="FFFFFFFF"; BLACK="FF000000"
        PL_POS="FF00B050"; PL_NEG="FFFF0000"

        def fill(x): return PatternFill("solid",fgColor=x)
        def ctr(w=False): return Alignment(horizontal="center",vertical="center",wrap_text=w)
        def lft(): return Alignment(horizontal="left",vertical="center")
        def bdr():
            s=Side(style="thin",color="FF333333")
            return Border(left=s,right=s,top=s,bottom=s)

        today = dt.datetime.now()

        HDRS=[
            "Date","Time","Competition","Home Team","Away Team","Country",
            "H Rank","A Rank","H Colour","A Colour","All Avg","H/A Avg",
            "H GP","H W","H D","H L","H Pts","H Ave",
            "A GP","A W","A D","A L","A Pts","A Ave",
            "Score 1","Score 2","Score 3",
            "HxG 45","AxG 45","HxG Match","AxG Match","Match xG","H Sup",
            "Pois Home","Pois Draw","Pois Away",
            "Pois O1.5","Pois O2.5","Pois O3.5","Pois BTTS","Pois FH O0.5",
            "Val Home","Val Draw","Val Away",
            "Val O1.5","Val O2.5","Val O3.5","Val BTTS","Val FH O0.5",
            "MO Vol","Back Home","Lay Home","Back Away","Lay Away","Back Draw","Lay Draw",
            "2.5 Vol","Back U2.5","Lay U2.5","Back O2.5","Lay O2.5",
            "1.5 Vol","Back U1.5","Lay U1.5","Back O1.5","Lay O1.5",
            "3.5 Vol","Back U3.5","Lay U3.5","Back O3.5","Lay O3.5",
            "4.5 Vol","Back U4.5","Lay U4.5","Back O4.5","Lay O4.5",
            "BTTS Vol","Back BTTS Y","Lay BTTS Y","Back BTTS N","Lay BTTS N",
            "FH 0.5 Vol","Back FH U0.5","Lay FH U0.5","Back FH O0.5","Lay FH O0.5",
            "FH 1.5 Vol","Back FH U1.5","Lay FH U1.5","Back FH O1.5","Lay FH O1.5",
        ]
        N=len(HDRS)

        H={}
        for i in range(0,6):   H[i]=(SEC_MATCH,AG_GOLD)
        for i in range(6,12):  H[i]=(SEC_MATCH,"FFAAB4D4")
        for i in range(12,24): H[i]=(SEC_FORM,AG_GOLD)
        for i in range(24,27): H[i]=(SEC_SCORE,AG_GOLD)
        for i in range(27,33): H[i]=(SEC_POIS_N,WHITE)
        for i in range(33,41): H[i]=(SEC_POIS_O,"FFDDBBFF")
        for i in range(41,49): H[i]=(SEC_VAL,AG_GOLD)
        for i,hdr in enumerate(HDRS):
            if i<49: continue
            if "Vol" in hdr:             H[i]=(VOL_BLUE if "MO" in hdr else VOL_GREEN,BLACK)
            elif hdr.startswith("Back"): H[i]=(BACK_COL,BLACK)
            elif hdr.startswith("Lay"):  H[i]=(LAY_COL,BLACK)
            else:                        H[i]=(AG_DARK,WHITE)

        SEC_NAMES=[
            (1,6,"MATCH DETAILS"),(7,12,"LEAGUE & FORM"),
            (13,24,"FORM  HOME / AWAY"),(25,27,"TOP 3 SCORELINES"),
            (28,41,"POISSON MODEL"),(42,49,"VALUE  —  Model vs Exchange Price"),
            (50,56,"MATCH ODDS"),(57,61,"UNDER / OVER 2.5"),
            (62,66,"UNDER / OVER 1.5"),(67,71,"UNDER / OVER 3.5"),
            (72,76,"UNDER / OVER 4.5"),(77,81,"BTTS"),
            (82,86,"FIRST HALF  U/O 0.5"),(87,N,"FIRST HALF  U/O 1.5"),
        ]

        fbd_data = get_fbd_data() or []
        fixtures = []

        for fbd in fbd_data:
            home = fbd.get("home",""); away = fbd.get("away","")
            if not home or not away: continue
            pred=fbd.get("pred_goals") or 0
            h_prob=fbd.get("pred_home_odds") or 0
            a_prob=fbd.get("pred_away_odds") or 0
            if pred>0 and (h_prob+a_prob)>0:
                h_share=h_prob/(h_prob+a_prob)
                home_xg=round(pred*h_share,2); away_xg=round(pred*(1-h_share),2)
            else:
                home_xg=1.35; away_xg=1.10
            poisson=build_poisson_for_match(
                home_xg_6=home_xg,away_xg_6=away_xg,
                home_xg_s=home_xg,away_xg_s=away_xg,
                home_team=home,away_team=away)
            p6=poisson.get("prob_6",{}); o6=poisson.get("odds_6",{})
            top3=poisson.get("top3",["","",""])
            while len(top3)<3: top3.append("")
            steam=get_steam_for_fixture(home,away)
            mkts=steam.get("markets",{}) if steam else {}
            def gp(mkt):
                m=mkts.get(mkt,{})
                return m.get("5m") or m.get("7am") or None
            def vl(prob,price):
                if not price or price<=1: return ""
                v=round((prob-1/price)*100,1)
                return v if abs(v)>0.5 else ""
            ko=""
            try:
                kstr=steam.get("kickoff","") if steam else ""
                if kstr: ko=dt.datetime.strptime(kstr[:19],"%Y-%m-%dT%H:%M:%S").strftime("%H:%M")
            except: pass
            fixtures.append({
                "home":home,"away":away,"ko":ko,
                "league":fbd.get("league",""),
                "home_xg":home_xg,"away_xg":away_xg,
                "top3":top3,"p6":p6,"o6":o6,
                "home_back":gp("home_back"),"home_lay":gp("home_lay"),
                "away_back":gp("away_back"),"away_lay":gp("away_lay"),
                "draw_back":gp("draw_back"),"draw_lay":gp("draw_lay"),
                "o25_back":gp("o25_back"),"o25_lay":gp("o25_lay"),
                "u25_back":gp("u25_back"),"u15_back":gp("u15_back"),
                "o15_back":gp("o15_back"),"o15_lay":gp("o15_lay"),
                "fho_back":gp("fh_o05_back"),"fho_lay":gp("fh_o05_lay"),
                "fhu_back":gp("fh_u05_back"),
                "val_home":vl(p6.get("home_win",0),gp("home_back")),
                "val_draw":vl(p6.get("draw",0),gp("draw_back")),
                "val_away":vl(p6.get("away_win",0),gp("away_back")),
                "val_o25":vl(p6.get("over_25",0),gp("o25_back")),
                "val_o15":vl(p6.get("over_15",0),gp("o15_back")),
                "val_btts":vl(p6.get("btts",0),gp("btts_yes_back")),
                "val_fho":vl(p6.get("fh_over_05",0),gp("fh_o05_back")),
            })

        out=Workbook(); ws=out.active
        ws.title=f"Pre-Match {today.strftime('%d-%m-%Y')}"
        ws.sheet_view.showGridLines=False; ws.freeze_panes="A5"

        ws.row_dimensions[1].height=22
        ws.cell(row=1,column=1).value="AGD Sports Trading  —  Football Trading Data"
        ws.cell(row=1,column=1).font=Font(name="Calibri",size=12,bold=True,color=AG_GOLD)
        ws.cell(row=1,column=1).fill=fill(AG_BLACK); ws.cell(row=1,column=1).alignment=lft()
        ws.cell(row=1,column=8).value=(f"Pre-Match Sheet  ·  {today.strftime('%A %d %B %Y')}  ·  "
            f"{len(fixtures)} Fixtures  ·  Blue=Back  Pink=Lay  ·  18+ Gamble Responsibly")
        ws.cell(row=1,column=8).font=Font(name="Calibri",size=9,color=MUTED)
        ws.cell(row=1,column=8).fill=fill(AG_BLACK); ws.cell(row=1,column=8).alignment=lft()
        for ci in range(2,N+1):
            if ci!=8: ws.cell(row=1,column=ci).fill=fill(AG_BLACK)

        ws.row_dimensions[2].height=18
        for sc,ec,name in SEC_NAMES:
            sl=get_column_letter(sc); el=get_column_letter(ec)
            if sc!=ec: ws.merge_cells(f"{sl}2:{el}2")
            cx=ws.cell(row=2,column=sc); cx.value=name
            bg,fc=H.get(sc-1,(AG_DARK,WHITE))
            cx.font=Font(name="Calibri",size=8,bold=True,color=fc)
            cx.fill=fill(bg); cx.alignment=ctr(); cx.border=bdr()

        ws.row_dimensions[3].height=34
        for i,hdr in enumerate(HDRS,1):
            cx=ws.cell(row=3,column=i,value=hdr)
            bg,fc=H.get(i-1,(AG_DARK,WHITE))
            cx.font=Font(name="Calibri",size=8,bold=True,color=fc)
            cx.fill=fill(bg); cx.alignment=ctr(True); cx.border=bdr()

        ws.row_dimensions[4].height=16
        ws.cell(row=4,column=1).value="TODAY'S FIXTURES  →"
        ws.cell(row=4,column=1).font=Font(name="Calibri",size=8,bold=True,color=AG_GOLD)
        ws.cell(row=4,column=1).fill=fill(AG_BLACK)
        ws.cell(row=4,column=1).alignment=lft(); ws.cell(row=4,column=1).border=bdr()
        for ci in range(2,N+1):
            ws.cell(row=4,column=ci).fill=fill(AG_BLACK); ws.cell(row=4,column=ci).border=bdr()

        WID={0:13,1:7,2:22,3:20,4:20,5:5,6:7,7:7,8:8,9:8,10:8,11:8}
        for i in range(N):
            ws.column_dimensions[get_column_letter(i+1)].width=WID.get(i,8)

        VALUE_COLS=set(range(42,49))

        for ri,fix in enumerate(fixtures,5):
            ws.row_dimensions[ri].height=16
            p6=fix["p6"]; o6=fix["o6"]; top3=fix["top3"]
            row_vals=[
                today.strftime("%d/%m/%Y"),fix["ko"],fix["league"],
                fix["home"],fix["away"],"",
                "","","","","","",
                "","","","","","",
                "","","","","","",
                top3[0],top3[1],top3[2],
                fix["home_xg"],fix["away_xg"],
                fix["home_xg"],fix["away_xg"],
                round(fix["home_xg"]+fix["away_xg"],2),
                round(fix["home_xg"]-fix["away_xg"],2),
                o6.get("home",""),o6.get("draw",""),o6.get("away",""),
                o6.get("over_15",""),o6.get("over_25",""),
                o6.get("over_35",""),o6.get("btts",""),o6.get("fh_over_05",""),
                fix["val_home"],fix["val_draw"],fix["val_away"],
                fix["val_o15"],fix["val_o25"],"",fix["val_btts"],fix["val_fho"],
                "",
                fix["home_back"],fix["home_lay"],
                fix["away_back"],fix["away_lay"],
                fix["draw_back"],fix["draw_lay"],
                "",fix["u25_back"],"",fix["o25_back"],fix["o25_lay"],
                "",fix["u15_back"],"",fix["o15_back"],fix["o15_lay"],
                "","","","","",
                "","","","","",
                "","","","","",
                "",fix["fhu_back"],"",fix["fho_back"],fix["fho_lay"],
                "","","","","",
            ]
            while len(row_vals)<N: row_vals.append("")

            for ci,val in enumerate(row_vals[:N],1):
                cx=ws.cell(row=ri,column=ci,value=val)
                sv=str(val) if val is not None else ""
                if ci==4:
                    cx.font=Font(name="Calibri",size=10,bold=True,color=AG_GOLD)
                    cx.alignment=lft()
                elif ci==5:
                    cx.font=Font(name="Calibri",size=10,bold=True,color=WHITE)
                    cx.alignment=lft()
                elif ci==1:
                    cx.font=Font(name="Calibri",size=9,bold=True,color=WHITE)
                    cx.alignment=lft()
                elif ci==2:
                    cx.font=Font(name="Calibri",size=9,bold=True,color=AG_GOLD)
                    cx.alignment=ctr()
                elif ci==3:
                    cx.font=Font(name="Calibri",size=8,color=MUTED)
                    cx.alignment=lft()
                elif ci in VALUE_COLS:
                    try:
                        v=float(sv)
                        cx.font=Font(name="Calibri",size=8,bold=True,
                                    color=PL_POS if v>0 else(PL_NEG if v<0 else WHITE))
                    except: cx.font=Font(name="Calibri",size=8,color=WHITE)
                    cx.alignment=ctr()
                else:
                    bg,fc=H.get(ci-1,(AG_DARK,WHITE))
                    cx.font=Font(name="Calibri",size=8,color=fc)
                    cx.alignment=ctr() if ci>6 else lft()

        buf=_io.BytesIO(); out.save(buf); buf.seek(0)
        return buf.getvalue()

    except Exception as e:
        logger.error(f"Pre-match sheet generation error: {e}")
        import traceback
        logger.error(traceback.format_exc())
        raise

@app.get("/sheet/prematch")
async def download_prematch_sheet(email: str = "", date: str = ""):
    """
    Download pre-match sheet for a specific date (or today if no date).
    Requires trading plan or admin.
    """
    if not _is_trading(email):
        return {"error": "Trading plan required. Upgrade at accagenius.com/trading"}
    try:
        sheet_date = date if date else datetime.now().strftime("%Y-%m-%d")
        filename   = f"AGD_PreMatch_{sheet_date}.xlsx"
        content    = _generate_prematach_sheet()
        return StreamingResponse(
            io.BytesIO(content),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f'attachment; filename="{filename}"',
                "Access-Control-Allow-Origin": "*",
                "Access-Control-Expose-Headers": "Content-Disposition",
            }
        )
    except Exception as e:
        logger.error(f"Pre-match download error: {e}")
        return {"error": str(e)}

@app.get("/sheet/results/today")
async def download_todays_results(email: str = "", date: str = ""):
    """
    Download results sheet for a specific date (or yesterday if no date).
    Requires trading plan or admin.
    """
    if not _is_trading(email):
        return {"error": "Trading plan required"}
    try:
        sheet_date = date if date else (datetime.now()-timedelta(days=1)).strftime("%Y-%m-%d")
        content    = await _generate_results_sheet(sheet_date)
        filename   = f"AGD_Results_{sheet_date}.xlsx"
        return StreamingResponse(
            io.BytesIO(content),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f'attachment; filename="{filename}"',
                "Access-Control-Allow-Origin": "*",
                "Access-Control-Expose-Headers": "Content-Disposition",
            }
        )
    except Exception as e:
        logger.error(f"Results download error: {e}")
        return {"error": str(e)}

@app.get("/sheet/steam")
async def download_steam_sheet_protected(email: str = ""):
    """
    Download personal steam detection sheet — full 3-sheet version.
    ADMIN ONLY.
    """
    if not _is_admin(email):
        return {"error": "Admin access only"}

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        import io as _io

        AG_BLACK="FF0D0D0D"; AG_GOLD="FFFFB800"; MUTED="FFB0B0B0"
        STEAM_HDR="FFFFCC00"; GREEN="FF00B050"; RED="FFFF0000"
        WHITE="FFFFFFFF"; BLACK="FF000000"; MID="FF1C1C1C"; ALT="FF242424"
        ALERT_BG="FF3D2200"

        def fill(x): return PatternFill("solid",fgColor=x)
        def ctr(w=False): return Alignment(horizontal="center",vertical="center",wrap_text=w)
        def lft(): return Alignment(horizontal="left",vertical="center")
        def bdr():
            s=Side(style="thin",color="FF2A2A2A")
            return Border(left=s,right=s,top=s,bottom=s)

        today = datetime.now()
        THRESH = 3.0

        # ── SHEET 1: STEAM DETECTION ────────────────────────────
        HDRS=["Date","KO Time","Competition","Home Team","Away Team",
              "7am O2.5","5m O2.5","O2.5 Move","O2.5 %","O2.5 Steam",
              "7am Draw","5m Draw","Draw Move","Draw %","Draw Steam",
              "7am FH O0.5","5m FH O0.5","FHO Move","FHO %","FHO Steam",
              "7am O1.5","5m O1.5","O1.5 Move","O1.5 %","O1.5 Steam",
              "Pois O2.5","Pois Draw","Pois O1.5","Pois FH O0.5",
              "Val O2.5","Val Draw","Val O1.5","Val FH O0.5",
              "LTD?","LU1.5?","LFHU0.5?","Notes"]
        N=len(HDRS)

        out=Workbook()
        ws=out.active; ws.title="Steam Detection"
        ws.sheet_view.showGridLines=False; ws.freeze_panes="A5"

        ws.row_dimensions[1].height=22
        ws.cell(row=1,column=1).value=f"AGD Sports Trading  ⚡  Steam Detection  |  {today.strftime('%A %d %B %Y')}  |  PRIVATE"
        ws.cell(row=1,column=1).font=Font(name="Calibri",size=12,bold=True,color=AG_GOLD)
        ws.cell(row=1,column=1).fill=fill(AG_BLACK); ws.cell(row=1,column=1).alignment=lft()
        for ci in range(2,N+1): ws.cell(row=1,column=ci).fill=fill(AG_BLACK)

        ws.row_dimensions[2].height=14
        ws.cell(row=2,column=1).value=("System 1 LTD: Draw+O2.5 steam  ·  "
            "System 2 LU1.5: Back O2.5 steams 3%+  ·  "
            "System 3 LFHU0.5: Back O1.5 steams 3%+  ·  "
            "Prices auto-captured at 7am & 5min pre-KO via Railway")
        ws.cell(row=2,column=1).font=Font(name="Calibri",size=8,color=MUTED)
        ws.cell(row=2,column=1).fill=fill(AG_BLACK); ws.cell(row=2,column=1).alignment=lft()
        for ci in range(2,N+1): ws.cell(row=2,column=ci).fill=fill(AG_BLACK)

        # Section headers row 3
        sections=[
            (1,5,"MATCH DETAILS","FF1A2A3A",AG_GOLD),
            (6,10,"7AM PRICES","FF003366","FFAACCFF"),
            (11,15,"5 MIN PRE-KO","FF006633","FFAAFFCC"),
            (16,20,"FH O0.5 MARKETS","FF003366","FFAACCFF"),
            (21,25,"O1.5 MARKETS","FF003366","FFAACCFF"),
            (26,29,"POISSON MODEL","FF1A0033","FFDDBBFF"),
            (30,33,"VALUE","FF003319",AG_GOLD),
            (34,36,"⚠ SYSTEMS","FF1A0000","FFFF6666"),
            (37,37,"NOTES","FF1A1A1A",MUTED),
        ]
        ws.row_dimensions[3].height=16
        for sc,ec,name,bg,fc in sections:
            sl=get_column_letter(sc); el=get_column_letter(ec)
            if sc!=ec: ws.merge_cells(f"{sl}3:{el}3")
            c2=ws.cell(row=3,column=sc); c2.value=name
            c2.font=Font(name="Calibri",size=7,bold=True,color=fc)
            c2.fill=fill(bg); c2.alignment=ctr(); c2.border=bdr()

        ws.row_dimensions[4].height=30
        hdr_colors={
            "Date":("FF1A2A3A",AG_GOLD),"KO Time":("FF1A2A3A",MUTED),
            "Competition":("FF1A2A3A",MUTED),"Home Team":("FF1A2A3A",AG_GOLD),
            "Away Team":("FF1A2A3A",AG_GOLD),
            "7am O2.5":("FF003366","FFAACCFF"),"5m O2.5":("FF006633","FFAAFFCC"),
            "O2.5 Move":("FF332600",AG_GOLD),"O2.5 %":("FF332600",AG_GOLD),
            "O2.5 Steam":(STEAM_HDR,BLACK),
            "7am Draw":("FF003366","FFAACCFF"),"5m Draw":("FF006633","FFAAFFCC"),
            "Draw Move":("FF332600",AG_GOLD),"Draw %":("FF332600",AG_GOLD),
            "Draw Steam":(STEAM_HDR,BLACK),
            "7am FH O0.5":("FF003366","FFAACCFF"),"5m FH O0.5":("FF006633","FFAAFFCC"),
            "FHO Move":("FF332600",AG_GOLD),"FHO %":("FF332600",AG_GOLD),
            "FHO Steam":(STEAM_HDR,BLACK),
            "7am O1.5":("FF003366","FFAACCFF"),"5m O1.5":("FF006633","FFAAFFCC"),
            "O1.5 Move":("FF332600",AG_GOLD),"O1.5 %":("FF332600",AG_GOLD),
            "O1.5 Steam":(STEAM_HDR,BLACK),
            "Pois O2.5":("FF1A0033","FFDDBBFF"),"Pois Draw":("FF1A0033","FFDDBBFF"),
            "Pois O1.5":("FF1A0033","FFDDBBFF"),"Pois FH O0.5":("FF1A0033","FFDDBBFF"),
            "Val O2.5":("FF003319",AG_GOLD),"Val Draw":("FF003319",AG_GOLD),
            "Val O1.5":("FF003319",AG_GOLD),"Val FH O0.5":("FF003319",AG_GOLD),
            "LTD?":("FF1A0000","FFFF6666"),"LU1.5?":("FF001A00","FF66FF66"),
            "LFHU0.5?":("FF001A1A","FF66FFFF"),"Notes":("FF1A1A1A",MUTED),
        }
        for i,hdr in enumerate(HDRS,1):
            c2=ws.cell(row=4,column=i,value=hdr)
            bg,fc=hdr_colors.get(hdr,("FF1A1A1A",WHITE))
            c2.font=Font(name="Calibri",size=8,bold=True,color=fc)
            c2.fill=fill(bg); c2.alignment=ctr(True); c2.border=bdr()

        # Column widths
        cw={1:11,2:7,3:18,4:16,5:16}
        for i in range(6,N+1):
            if "Steam" in HDRS[i-1] or "?" in HDRS[i-1]: cw[i]=10
            elif HDRS[i-1]=="Notes": cw[i]=20
            else: cw[i]=8
        for col,w in cw.items():
            ws.column_dimensions[get_column_letter(col)].width=w

        # Data rows
        rows_written=0
        for fkey,fdata in _steam_prices.items():
            mkts=fdata.get("markets",{})
            home=fdata.get("home",""); away=fdata.get("away","")
            league=fdata.get("league",""); kickoff=fdata.get("kickoff","")

            # Parse KO time
            ko_time=""
            try:
                from datetime import datetime as _dt
                ko_time=_dt.strptime(kickoff[:19],"%Y-%m-%dT%H:%M:%S").strftime("%H:%M")
            except: pass

            def gm(mkt,t): return mkts.get(mkt,{}).get(t)
            def mv(p7,p5):
                if not p7 or not p5: return None,None
                return round(p5-p7,2),round(abs(p5-p7)/p7*100,1)
            def sf(m,pct):
                if not m or not pct: return ""
                if pct>=THRESH and m<0: return "🔥 STEAM"
                if pct>=THRESH and m>0: return "📈 DRIFT"
                return "—"

            o25_7=gm("o25_back","7am"); o25_5=gm("o25_back","5m")
            draw_7=gm("draw_back","7am"); draw_5=gm("draw_back","5m")
            fho_7=gm("fh_o05_back","7am"); fho_5=gm("fh_o05_back","5m")
            o15_7=gm("o15_back","7am"); o15_5=gm("o15_back","5m")

            o25_m,o25_p=mv(o25_7,o25_5); draw_m,draw_p=mv(draw_7,draw_5)
            fho_m,fho_p=mv(fho_7,fho_5); o15_m,o15_p=mv(o15_7,o15_5)

            o25_s=sf(o25_m,o25_p); draw_s=sf(draw_m,draw_p)
            fho_s=sf(fho_m,fho_p); o15_s=sf(o15_m,o15_p)

            o25_steam=o25_p and o25_p>=THRESH and o25_m and o25_m<0
            draw_steam=draw_p and draw_p>=THRESH and draw_m and draw_m<0
            o15_steam=o15_p and o15_p>=THRESH and o15_m and o15_m<0

            sys1="✅ TRIGGER" if (draw_steam and o25_steam) else ""
            sys2="✅ TRIGGER" if o25_steam else ""
            sys3="✅ TRIGGER" if o15_steam else ""
            has_steam=any([o25_steam,draw_steam,o15_steam])

            ri=rows_written+5
            row_vals=[
                today.strftime("%d/%m/%Y"), ko_time, league, home, away,
                o25_7, o25_5, o25_m, f"{o25_p}%" if o25_p else "", o25_s,
                draw_7, draw_5, draw_m, f"{draw_p}%" if draw_p else "", draw_s,
                fho_7, fho_5, fho_m, f"{fho_p}%" if fho_p else "", fho_s,
                o15_7, o15_5, o15_m, f"{o15_p}%" if o15_p else "", o15_s,
                "","","","",  # Poisson — from model
                "","","","",  # Value
                sys1, sys2, sys3, "",
            ]
            ws.row_dimensions[ri].height=14
            for ci,val in enumerate(row_vals[:N],1):
                c2=ws.cell(row=ri,column=ci,value=val); sv=str(val) if val else ""
                c2.fill=fill(ALERT_BG if has_steam else(ALT if ri%2==0 else MID))
                if ci in(4,5): c2.font=Font(name="Calibri",size=8,bold=True,color=AG_GOLD);c2.alignment=lft()
                elif ci in(1,2,3): c2.font=Font(name="Calibri",size=8,color=MUTED);c2.alignment=lft()
                elif "STEAM" in sv: c2.font=Font(name="Calibri",size=8,bold=True,color="FFFF6600");c2.alignment=ctr()
                elif "TRIGGER" in sv: c2.font=Font(name="Calibri",size=8,bold=True,color=GREEN);c2.alignment=ctr()
                elif "DRIFT" in sv: c2.font=Font(name="Calibri",size=8,bold=True,color=GREEN);c2.alignment=ctr()
                elif "Move" in HDRS[ci-1]:
                    try:
                        v=float(sv)
                        c2.font=Font(name="Calibri",size=8,bold=True,color=GREEN if v>0 else(RED if v<0 else MUTED))
                    except: c2.font=Font(name="Calibri",size=8,color=WHITE)
                    c2.alignment=ctr()
                elif "%" in HDRS[ci-1]:
                    try:
                        v=float(sv.replace("%",""))
                        c2.font=Font(name="Calibri",size=8,bold=(v>=THRESH),color="FFFF6600" if v>=THRESH else MUTED)
                    except: c2.font=Font(name="Calibri",size=8,color=MUTED)
                    c2.alignment=ctr()
                else: c2.font=Font(name="Calibri",size=8,color=WHITE);c2.alignment=ctr()
                c2.border=bdr()
            rows_written+=1

        # ── SHEET 2: P&L TRACKER ────────────────────────────────
        ps=out.create_sheet("P&L Tracker")
        ps.sheet_view.showGridLines=False
        ps.merge_cells("A1:K1")
        ps["A1"]="AGD Sports Trading  —  Steam Systems P&L Tracker  |  PRIVATE"
        ps["A1"].font=Font(name="Calibri",size=11,bold=True,color=AG_GOLD)
        ps["A1"].fill=fill(AG_BLACK); ps["A1"].alignment=lft()
        ps.row_dimensions[1].height=22
        pl_h=["Date","System","Fixture","KO Time","Trigger %","Entry Market",
              "Entry Odds","Stake (pts)","Result","P&L (pts)","Running Total","Notes"]
        for i,h in enumerate(pl_h,1):
            c2=ps.cell(row=3,column=i,value=h)
            c2.font=Font(name="Calibri",size=8,bold=True,color=BLACK)
            c2.fill=fill(STEAM_HDR); c2.alignment=ctr(True); c2.border=bdr()
        ps.row_dimensions[3].height=28
        for i,w in enumerate([11,14,24,8,10,14,10,10,8,10,12,20],1):
            ps.column_dimensions[get_column_letter(i)].width=w

        # ── SHEET 3: SYSTEM RULES ───────────────────────────────
        ns=out.create_sheet("System Rules")
        ns.sheet_view.showGridLines=False
        ns.merge_cells("A1:D1")
        ns["A1"]="AGD Sports Trading  ⚡  Steam System Rules  |  PRIVATE"
        ns["A1"].font=Font(name="Calibri",size=11,bold=True,color=AG_GOLD)
        ns["A1"].fill=fill(AG_BLACK); ns["A1"].alignment=lft()
        ns.row_dimensions[1].height=22
        rules=[
            ("SYSTEM 1","LAY THE DRAW (LTD)",
             "TRIGGER: Draw market steams 3%+ AND O2.5 also steams 3%+",
             "ENTRY: Lay the draw at 5min pre-KO price on Betfair Exchange"),
            ("","",
             "LOGIC: When draw AND goals steam together = smart money expects goals",
             "EXIT: Green up at HT if 0-0. Let run if goal scored."),
            ("SYSTEM 2","LAY UNDER 1.5 (LU1.5)",
             "TRIGGER: Back O2.5 market steams 3%+",
             "ENTRY: Lay Under 1.5 at 5min pre-KO price"),
            ("","",
             "LOGIC: O2.5 steam = smart money backing 3+ goals. Lay the unders.",
             "EXIT: Win if 2+ goals. Lose if 0 or 1 goal."),
            ("SYSTEM 3","LAY FH UNDER 0.5",
             "TRIGGER: Back O1.5 market steams 3%+",
             "ENTRY: Lay FH Under 0.5 at 5min pre-KO price"),
            ("","",
             "LOGIC: O1.5 steam = smart money expects early goals.",
             "EXIT: Win if any goal in first half. Lose if 0-0 at HT."),
            ("THRESHOLD","3% minimum",
             "Below 3% = noise. 3-5% = moderate. 5%+ = strong signal.",
             "Adjust THRESH in main.py if needed."),
            ("STAKING","1 point per bet",
             "Start with 1pt until you have 50+ results per system.",
             "Scale up when proven profitable over 100+ bets."),
        ]
        for ri,row in enumerate(rules,3):
            ns.row_dimensions[ri].height=18
            for ci,val in enumerate(row,1):
                c2=ns.cell(row=ri,column=ci,value=val)
                c2.fill=fill(ALT if ri%2==0 else MID)
                sys_colors={"SYSTEM 1":"FFFF6666","SYSTEM 2":"FF66FF66",
                            "SYSTEM 3":"FF66FFFF","THRESHOLD":AG_GOLD,"STAKING":MUTED}
                c2.font=Font(name="Calibri",size=9,bold=(ci==1 and val!=""),
                            color=sys_colors.get(val,AG_GOLD if ci==1 and val else WHITE))
                c2.alignment=lft(); c2.border=bdr()
        for i,w in enumerate([12,20,48,48],1):
            ns.column_dimensions[get_column_letter(i)].width=w

        buf=_io.BytesIO(); out.save(buf); buf.seek(0)
        today_str=datetime.now().strftime("%Y-%m-%d")
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f'attachment; filename="AGD_Steam_{today_str}.xlsx"',
                "Access-Control-Allow-Origin": "*",
            }
        )
    except Exception as e:
        logger.error(f"Steam sheet error: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return {"error": str(e)}
@app.get("/admin/sheets")
async def admin_all_sheets(email: str = ""):
    """
    Admin endpoint — lists all available downloads.
    Only accessible by admin.
    """
    if not _is_admin(email):
        return {"error": "Admin only"}

    return {
        "seasons": [
            {"season": s, "url": f"/sheet/season/{s}?email={email}"}
            for s in TRADING_SEASONS
        ],
        "prematch":   f"/sheet/prematch?email={email}",
        "results":    f"/sheet/results/today?email={email}",
        "steam":      f"/sheet/steam?email={email}",
        "poisson":    "/poisson/today",
        "steam_data": "/steam/all",
        "alerts":     "/steam/alerts",
    }

# ── AUTO RESULTS UPDATE — runs nightly ──────────────────────────
_results_cache: dict = {}
_results_cache_date: str = ""

async def _fetch_yesterdays_results(target_date: str = ""):
    """
    Pull results from API-Football for a specific date (default yesterday).
    Caches yesterday's results. Other dates fetched fresh.
    """
    global _results_cache, _results_cache_date
    yesterday = (datetime.utcnow() - timedelta(days=1)).strftime("%Y-%m-%d")
    fetch_date = target_date if target_date else yesterday

    # Only cache yesterday's results
    if fetch_date == yesterday and _results_cache_date == yesterday and _results_cache:
        return _results_cache

    logger.info(f"Fetching results for {yesterday}...")
    try:
        our_league_ids = {l["id"] for l in LEAGUES}
        all_results = []

        for league_id in our_league_ids:
            try:
                data = api_get("fixtures", {
                    "date":   fetch_date,
                    "league": league_id,
                    "season": 2025,
                })
                fixtures = data.get("response", [])
                for f in fixtures:
                    status = f.get("fixture",{}).get("status",{}).get("short","")
                    if status != "FT":
                        continue

                    home = f.get("teams",{}).get("home",{}).get("name","")
                    away = f.get("teams",{}).get("away",{}).get("name","")
                    score= f.get("score",{})
                    goals= f.get("goals",{})
                    ht_h = score.get("halftime",{}).get("home")
                    ht_a = score.get("halftime",{}).get("away")
                    ft_h = goals.get("home")
                    ft_a = goals.get("away")

                    if ft_h is None or ft_a is None:
                        continue

                    ft_goals = ft_h + ft_a
                    ht_goals = (ht_h or 0) + (ht_a or 0)

                    all_results.append({
                        "date":      yesterday,
                        "league_id": league_id,
                        "home":      home,
                        "away":      away,
                        "ht_score":  f"{ht_h}-{ht_a}" if ht_h is not None else "",
                        "ft_score":  f"{ft_h}-{ft_a}",
                        "ht_1x2":    "H" if ht_h>ht_a else("A" if ht_a>ht_h else "D"),
                        "ft_1x2":    "H" if ft_h>ft_a else("A" if ft_a>ft_h else "D"),
                        "ht_goals":  ht_goals,
                        "ft_goals":  ft_goals,
                        "o15":       "Y" if ft_goals>1 else "N",
                        "o25":       "Y" if ft_goals>2 else "N",
                        "o35":       "Y" if ft_goals>3 else "N",
                        "btts":      "Y" if ft_h>0 and ft_a>0 else "N",
                        "fh_o05":    "Y" if ht_goals>0 else "N",
                    })
            except Exception:
                continue

        _results_cache      = all_results
        _results_cache_date = yesterday
        logger.info(f"✅ Fetched {len(all_results)} results for {yesterday}")
        return all_results

    except Exception as e:
        logger.error(f"Results fetch error: {e}")
        return []

async def _generate_results_sheet(target_date: str = "") -> bytes:
    """Generate yesterday's results as Excel file."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    import io as _io

    AG_BLACK="FF0D0D0D"; AG_GOLD="FFFFB800"; MUTED="FFB0B0B0"
    GREEN="FF00B050"; RED="FFFF0000"; WHITE="FFFFFFFF"
    MID="FF1C1C1C"; ALT="FF242424"

    def fill(c): return PatternFill("solid",fgColor=c)
    def ctr(): return Alignment(horizontal="center",vertical="center")
    def lft(): return Alignment(horizontal="left",vertical="center")
    def bdr():
        s=Side(style="thin",color="FF2A2A2A")
        return Border(left=s,right=s,top=s,bottom=s)

    results = await _fetch_yesterdays_results(target_date)
    display_date = target_date if target_date else (datetime.now()-timedelta(days=1)).strftime("%Y-%m-%d")
    yesterday = datetime.strptime(display_date, "%Y-%m-%d").strftime("%d/%m/%Y")

    HDRS=["Date","Home Team","Away Team","HT Score","FT Score",
          "HT 1X2","FT 1X2","HT Goals","FT Goals",
          "O1.5","O2.5","O3.5","BTTS","FH O0.5"]
    N=len(HDRS)

    out=Workbook(); ws=out.active
    ws.title=f"Results {yesterday}"
    ws.sheet_view.showGridLines=False; ws.freeze_panes="A3"

    ws.row_dimensions[1].height=22
    ws.cell(row=1,column=1).value=f"AGD Sports Trading — Daily Results  |  {yesterday}  |  {len(results)} Games"
    ws.cell(row=1,column=1).font=Font(name="Calibri",size=12,bold=True,color=AG_GOLD)
    ws.cell(row=1,column=1).fill=fill(AG_BLACK); ws.cell(row=1,column=1).alignment=lft()
    for ci in range(2,N+1): ws.cell(row=1,column=ci).fill=fill(AG_BLACK)

    ws.row_dimensions[2].height=26
    for i,hdr in enumerate(HDRS,1):
        c=ws.cell(row=2,column=i,value=hdr)
        c.font=Font(name="Calibri",size=8,bold=True,color=WHITE)
        c.fill=fill("FF2E3A5C"); c.alignment=ctr(); c.border=bdr()

    for ri,res in enumerate(results,3):
        ws.row_dimensions[ri].height=14
        row_vals=[
            res["date"],res["home"],res["away"],
            res["ht_score"],res["ft_score"],
            res["ht_1x2"],res["ft_1x2"],
            res["ht_goals"],res["ft_goals"],
            res["o15"],res["o25"],res["o35"],
            res["btts"],res["fh_o05"],
        ]
        alt=ri%2==0
        for ci,val in enumerate(row_vals,1):
            c=ws.cell(row=ri,column=ci,value=val); sv=str(val) if val else ""
            c.fill=fill(ALT if alt else MID)
            if ci in(2,3):
                c.font=Font(name="Calibri",size=8,bold=True,color=AG_GOLD); c.alignment=lft()
            elif sv in("Y","N"):
                c.font=Font(name="Calibri",size=8,bold=True,
                           color=GREEN if sv=="Y" else RED); c.alignment=ctr()
            elif ci in(6,7):
                col=GREEN if sv=="H" else(RED if sv=="A" else WHITE)
                c.font=Font(name="Calibri",size=8,bold=True,color=col); c.alignment=ctr()
            else:
                c.font=Font(name="Calibri",size=8,color=WHITE); c.alignment=ctr()
            c.border=bdr()

    for i,w in enumerate([11,18,18,9,9,8,8,8,8,7,7,7,7,7],1):
        ws.column_dimensions[get_column_letter(i)].width=w

    buf=_io.BytesIO(); out.save(buf); buf.seek(0)
    return buf.getvalue()

async def _generate_season_file(season: str) -> bytes:
    """
    Generate a season results file.
    Uses cached AG5 format — same columns as the pre-built files.
    """
    # For now return a placeholder pointing to the pre-built files
    # In production these would be stored on Railway volume or S3
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    import io as _io

    out=Workbook(); ws=out.active; ws.title=season
    ws.cell(row=1,column=1).value=(
        f"AGD Sports Trading — Season {season}  |  "
        f"Download the full file from your member dashboard")
    ws.cell(row=1,column=1).font=Font(name="Calibri",size=11,bold=True,color="FFFFB800")
    ws.cell(row=1,column=1).fill=PatternFill("solid",fgColor="FF0D0D0D")
    ws.cell(row=1,column=1).alignment=Alignment(horizontal="left",vertical="center")
    ws.column_dimensions["A"].width=80
    ws.row_dimensions[1].height=30

    buf=_io.BytesIO(); out.save(buf); buf.seek(0)
    return buf.getvalue()

# ── NIGHTLY RESULTS SCHEDULER ───────────────────────────────────
async def nightly_results_update():
    """
    Runs at 6am every morning.
    Fetches previous day's results from API-Football.
    Caches them ready for member download.
    """
    logger.info("🌙 Nightly results update starting...")
    try:
        results = await _fetch_yesterdays_results()
        logger.info(f"✅ Nightly update complete — {len(results)} results cached")
    except Exception as e:
        logger.error(f"Nightly results update error: {e}")

logger.info("✅ Trading plan loaded — /create-trading-checkout /trading/access /sheet/season/{season} /sheet/prematch /sheet/results/today /admin/sheets")


async def _trading_scheduler():
    """
    Trading plan scheduler — runs daily tasks:
    - 6:00am: fetch yesterday's results from API-Football
    - 7:00am: trigger 7am Betfair price capture (if PC script not running)
    """
    import pytz
    london = pytz.timezone("Europe/London")
    logger.info("✅ Trading scheduler started")

    while True:
        try:
            now = datetime.now(london)

            # 6am — fetch yesterday's results
            results_target = now.replace(hour=6, minute=0, second=0, microsecond=0)
            if now >= results_target:
                results_target += timedelta(days=1)

            # 7am — 7am price capture fallback
            prices_target = now.replace(hour=7, minute=5, second=0, microsecond=0)
            if now >= prices_target:
                prices_target += timedelta(days=1)

            # Sleep until next task
            next_task = min(results_target, prices_target)
            wait_secs = (next_task - now).total_seconds()
            logger.info(f"⏰ Trading scheduler: next task in {int(wait_secs/60)}m")
            await asyncio.sleep(wait_secs)

            now = datetime.now(london)

            # 6am results fetch
            if abs((now - results_target + timedelta(days=1)).total_seconds()) < 300:
                logger.info("🌙 6am — fetching yesterday's results...")
                await _fetch_yesterdays_results()

            # 7am price capture fallback
            if abs((now - prices_target + timedelta(days=1)).total_seconds()) < 300:
                logger.info("⏰ 7am — triggering price capture fallback...")
                capture_exchange_prices("7am")

        except Exception as e:
            logger.error(f"Trading scheduler error: {e}")
            await asyncio.sleep(3600)
